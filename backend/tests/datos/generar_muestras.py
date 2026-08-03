"""Genera las muestras reducidas de Excel que usan las pruebas.

Se versiona el generador **y** los .xlsx que produce: las pruebas tienen que poder correr sin
los archivos reales (5.4 MB, no versionados), y a la vez hay que poder rehacer las muestras
cuando el cliente actualice su Excel.

    python tests/datos/generar_muestras.py --origen /ruta/a/data/layers/data

Las muestras no son un recorte al azar. Cada anomalía verificada en la auditoría del 02/08 está
representada, porque son exactamente los casos que las pruebas protegen:

nivel_peligro_muestra.xlsx
  · 4 hojas: Sismo, Lluvias (la columna PELIGRO dice "Lluvias intensas"), Incendios Forestales
    (idem, con minúscula) —las dos discrepancias hoja/columna— e Inundación, que es la única
    hoja donde aparece la segunda grafía de Fuente
  · SICUANI (0806010001) con DISTRITO en blanco en una hoja y lleno en otra
  · filas con PELIGRO y sin NIVEL_PELI
  · las dos filas huérfanas sin CODIGO al final de Incendios Forestales
  · los dos valores de Fuente (SIGRID_CENEPRED y SINAGERD_CENEPRED)

frecuencia_muestra.xlsx
  · OLLANTAYTAMBO: desglose normal
  · CUSCO: los cuatro TOT_* llenos y ninguna columna de evento (ADR-D1)
  · SANGARARA: descuadre entre subtotal y desglose
  · MOLLEPATA: descuadre en otra categoría
  · sin fila de ACOMAYO, que es el distrito ausente del padrón
  · una fila con FUENTE invertida (CENEPRED_SIGRID) y un RANGO FECHA con espacios
"""
import argparse
import pathlib

import openpyxl

AQUI = pathlib.Path(__file__).resolve().parent

HOJAS_NIVEL = ["Sismo", "Lluvias", "Inundación", "Incendios Forestales"]
CCPP_OBLIGATORIOS = {
    "0806010001",  # SICUANI: DISTRITO vacío en una de las hojas
    "0801010001",  # CUSCO ciudad: el CCPP más poblado (111,930)
    "0803020001",  # PISAC: distrito de la medida demo
}
DISTRITOS_FRECUENCIA = ["OLLANTAYTAMBO", "CUSCO", "SANGARARA", "MOLLEPATA", "ACOPIA"]
FILAS_POR_HOJA = 60


def muestra_niveles(origen: pathlib.Path, destino: pathlib.Path) -> None:
    wb = openpyxl.load_workbook(origen, read_only=True, data_only=True)
    salida = openpyxl.Workbook()
    salida.remove(salida.active)

    for hoja in HOJAS_NIVEL:
        if hoja not in wb.sheetnames:
            raise SystemExit(f"la hoja '{hoja}' no está en {origen.name}")
        ws = wb[hoja]
        it = ws.iter_rows(values_only=True)
        header = next(it)
        obligatorias, con_nivel, sin_nivel, huerfanas = [], [], [], []
        # Al menos una fila de cada grafía de Fuente: en Inundación las 49 de
        # SINAGERD_CENEPRED están al final, y un recorte por posición se las dejaría fuera.
        por_fuente: dict[str, list] = {}
        for fila in it:
            if not fila or all(c is None for c in fila):
                continue
            codigo = str(fila[3]).strip() if fila[3] is not None else ""
            if not codigo:
                huerfanas.append(fila)
            elif codigo in CCPP_OBLIGATORIOS:
                obligatorias.append(fila)
            elif fila[10] and fila[12] is None:
                sin_nivel.append(fila)
            elif fila[12] is not None:
                fuente = str(fila[13] or "").strip()
                muestras = por_fuente.setdefault(fuente, [])
                if len(muestras) < 3:
                    muestras.append(fila)
                else:
                    con_nivel.append(fila)

        por_fuente_planas = [f for filas in por_fuente.values() for f in filas]
        hueco = FILAS_POR_HOJA - len(obligatorias) - 5 - len(por_fuente_planas)
        elegidas = (
            obligatorias
            + sin_nivel[:5]
            + por_fuente_planas
            + con_nivel[: max(hueco, 0)]
            + huerfanas[:2]
        )
        destino_ws = salida.create_sheet(hoja)
        destino_ws.append(list(header[:15]))
        for fila in elegidas:
            destino_ws.append(list(fila[:15]))
        print(f"  {hoja}: {len(elegidas)} filas "
              f"({len(obligatorias)} clave, {len(sin_nivel[:5])} sin nivel, "
              f"{len(huerfanas[:2])} sin código)")
    wb.close()
    salida.save(destino)


def muestra_frecuencia(origen: pathlib.Path, destino: pathlib.Path) -> None:
    wb = openpyxl.load_workbook(origen, read_only=True, data_only=True)
    ws = wb["NºEMERGENCIAS"]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    filas = []
    for fila in it:
        if not fila or not fila[2]:
            continue
        nombre = str(fila[2]).strip().upper()
        if nombre in DISTRITOS_FRECUENCIA or nombre.startswith("SANGARAR"):
            filas.append(fila)
    wb.close()

    salida = openpyxl.Workbook()
    ws_out = salida.active
    ws_out.title = "NºEMERGENCIAS"
    ws_out.append(list(header))
    for fila in filas:
        ws_out.append(list(fila))
    print(f"  NºEMERGENCIAS: {len(filas)} distritos "
          f"({', '.join(str(f[2]).strip() for f in filas)})")
    salida.save(destino)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--origen", required=True, help="Carpeta con los Excel reales.")
    args = parser.parse_args()
    origen = pathlib.Path(args.origen)

    print("nivel_peligro_muestra.xlsx")
    muestra_niveles(
        origen / "Base_Nivel Peligro_CCPP_Cusco.xlsx", AQUI / "nivel_peligro_muestra.xlsx"
    )
    print("frecuencia_muestra.xlsx")
    muestra_frecuencia(
        origen / "Base_Frecuencia_Peligro_Cusco.xlsx", AQUI / "frecuencia_muestra.xlsx"
    )
