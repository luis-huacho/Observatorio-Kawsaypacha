"""API del contenido editorial (spec 02): medidas, normativa, noticias, videos, biblioteca.

La regla que ordena el módulo: **fuera del estado publicado no existe nada**. Es la garantía que
sostiene todo el flujo de revisión — si un borrador se colara por el listado, por el detalle o por
un export, la revisión sería decorativa.
"""
import io

import openpyxl
import pytest

pytestmark = pytest.mark.django_db


@pytest.fixture
def medida(db):
    """Una medida publicada, con lo mínimo obligatorio y sin imagen propia."""
    from apps.medidas.models import Medida
    from apps.peligros.models import TipoPeligro

    return Medida.objects.create(
        slug="terrazas-en-pisac",
        titulo="Recuperación de terrazas en Písac",
        tipo_peligro=TipoPeligro.objects.get(slug="lluvias_intensas"),
        ambito=Medida.Ambito.COMUNAL,
        resultado=Medida.Resultado.EXITO,
        resumen_corto="Andenes rehabilitados con faenas comunales.",
        palabras_clave=["andenes", "lluvias"],
        estado=Medida.Estado.PUBLICADO,
    )


@pytest.fixture
def norma(db):
    import datetime

    from apps.normativa.models import Norma

    return Norma.objects.create(
        slug="ordenanza-grd-cusco-2024",
        titulo="Ordenanza regional de gestión del riesgo",
        tipo=Norma.Tipo.choices[0][0],
        ambito=Norma.Ambito.choices[0][0],
        fecha=datetime.date(2024, 5, 20),
        resumen="Crea el grupo de trabajo regional de GRD.",
        palabras_clave=["grd"],
        estado=Norma.Estado.PUBLICADO,
    )


# --- Solo lo publicado se sirve ---------------------------------------------


@pytest.mark.parametrize(
    "estado", ["borrador", "archivado"]
)
def test_lo_no_publicado_no_existe_para_el_api(api, medida, estado):
    """Ni en el listado, ni en el detalle por slug. Un 404, no un 403.

    Que el detalle responda 404 y no «no autorizado» es intencional: el API público no confirma
    la existencia de contenido que aún no se ha publicado.
    """
    medida.estado = estado
    medida.save()

    listado = api.get("/api/medidas/").json()
    detalle = api.get(f"/api/medidas/{medida.slug}/")

    assert listado["count"] == 0
    assert detalle.status_code == 404


def test_una_medida_publicada_sale_con_su_forma_completa(api, medida):
    listado = api.get("/api/medidas/").json()
    detalle = api.get(f"/api/medidas/{medida.slug}/").json()

    assert listado["count"] == 1
    assert set(listado["results"][0]) >= {
        "slug", "titulo", "peligro", "peligro_slug", "ambito", "resultado", "distrito",
        "resumen_corto", "imagen_portada", "imagen_titulo", "palabras_clave", "publicado_en",
    }
    assert set(detalle) >= {"contenido", "galeria", "enlaces", "video_url"}


def test_la_portada_llega_resuelta_por_el_servidor(api, medida):
    """La regla del default institucional vive en el backend (spec 01).

    Ningún cliente la replica: el día que cambie la ilustración se toca un solo sitio. Y el pie
    dice que es una ilustración, para no hacer pasar el gráfico por una fotografía del hecho.
    """
    fila = api.get("/api/medidas/").json()["results"][0]

    assert fila["imagen_portada"].endswith("/img/default/peligro-lluvias_intensas.svg")
    assert fila["imagen_titulo"] == "Ilustración del Observatorio Kallpachakuy"


def test_filtros_de_medidas(api, medida):
    from apps.medidas.models import Medida
    from apps.peligros.models import TipoPeligro

    Medida.objects.create(
        slug="cortina-rompeviento",
        titulo="Cortinas rompeviento en Canchis",
        tipo_peligro=TipoPeligro.objects.get(slug="heladas"),
        ambito=Medida.Ambito.DISTRITAL,
        resultado=Medida.Resultado.LECCION,
        resumen_corto="Barreras vivas contra las heladas.",
        palabras_clave=["heladas"],
        estado=Medida.Estado.PUBLICADO,
    )

    assert api.get("/api/medidas/?peligro=heladas").json()["count"] == 1
    assert api.get("/api/medidas/?resultado=exito").json()["count"] == 1
    assert api.get("/api/medidas/?ambito=comunal").json()["count"] == 1
    assert api.get("/api/medidas/?tema=heladas").json()["count"] == 1
    assert api.get("/api/medidas/").json()["count"] == 2


def test_el_tema_filtra_por_coincidencia_exacta_no_parcial(api, medida):
    """`?tema=` alimenta los chips navegables: «lluvia» no debe traer «lluvias»."""
    assert api.get("/api/medidas/?tema=lluvias").json()["count"] == 1
    assert api.get("/api/medidas/?tema=lluvia").json()["count"] == 0


# --- Paginación -------------------------------------------------------------


def test_paginacion_por_defecto_y_techo(api, medida):
    """El techo de 200 impide que `?page_size=100000` convierta un listado en un export.

    Un export sin límite se saltaría el throttling de descargas, que es donde está el control.
    """
    respuesta = api.get("/api/medidas/?page_size=100000").json()

    assert set(respuesta) == {"count", "next", "previous", "results"}
    assert len(respuesta["results"]) <= 200


# --- Normativa y exports ----------------------------------------------------


def test_normativa_expone_anio_y_acceso_al_documento(api, norma):
    fila = api.get("/api/normativa/").json()["results"][0]

    assert fila["anio"] == 2024
    assert "documento_url" in fila


def test_normativa_se_filtra_por_anio(api, norma):
    assert api.get("/api/normativa/?anio=2024").json()["count"] == 1
    assert api.get("/api/normativa/?anio=2023").json()["count"] == 0


def test_el_export_de_normativa_solo_lleva_lo_publicado(api, norma, sin_throttling):
    """El Excel es una copia del listado: si arrastrara borradores, sería una fuga de contenido."""
    from apps.normativa.models import Norma

    Norma.objects.create(
        slug="borrador-interno",
        titulo="Proyecto de ordenanza en discusión",
        tipo=Norma.Tipo.choices[0][0],
        ambito=Norma.Ambito.choices[0][0],
        fecha=norma.fecha,
        resumen="No debe salir del admin.",
        estado=Norma.Estado.BORRADOR,
    )

    respuesta = api.get("/api/normativa/export.xlsx")
    libro = openpyxl.load_workbook(io.BytesIO(respuesta.content))
    titulos = [
        fila[0] for fila in libro.active.iter_rows(min_row=2, values_only=True) if fila[0]
    ]

    assert respuesta["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml"
    )
    assert "attachment" in respuesta["Content-Disposition"]
    assert any("ordenanza regional" in t.lower() for t in titulos)
    assert not any("discusión" in t.lower() for t in titulos)


def test_el_export_de_centros_poblados_respeta_los_filtros(api, datos_muestra, sin_throttling):
    respuesta = api.get("/api/ccpp/export.xlsx?clasificados=1&peligro=sismo&nivel_min=4")
    libro = openpyxl.load_workbook(io.BytesIO(respuesta.content))
    filas = list(libro.active.iter_rows(min_row=2, values_only=True))
    esperado = api.get("/api/ccpp/?clasificados=1&peligro=sismo&nivel_min=4").json()["count"]

    assert len(filas) == esperado


def _hoja_ccpp(api, consulta=""):
    """Descarga el Excel de centros poblados y devuelve `(cabecera, filas)`.

    Las celdas vacías vuelven de openpyxl como `None` —una cadena vacía no se guarda—, y aquí
    se normalizan a `""` para poder compararlas con lo que escribe el export.
    """
    respuesta = api.get(f"/api/ccpp/export.xlsx?clasificados=1{consulta}")
    libro = openpyxl.load_workbook(io.BytesIO(respuesta.content))
    filas = [
        [c if c is not None else "" for c in f]
        for f in libro.active.iter_rows(values_only=True)
    ]
    return filas[0], filas[1:]


def _catalogo():
    from apps.peligros.models import TipoPeligro

    return list(TipoPeligro.objects.all())


def test_el_export_de_centros_poblados_no_lleva_altitud_ni_coordenadas(
    api, datos_muestra, sin_throttling
):
    """Salieron de la ficha por ser ruido para quien consulta exposición; el Excel iba detrás.

    En su lugar entra una columna por peligro del catálogo, y salen **del catálogo**: si mañana
    se clasifica un décimo peligro, el Excel lo lleva sin tocar código.
    """
    cabecera, filas = _hoja_ccpp(api)

    assert not [c for c in cabecera if c.startswith(("Altitud", "Latitud", "Longitud"))]
    esperadas = [f"{t.nombre} (nivel)" for t in _catalogo()]
    assert cabecera[-len(esperadas):] == esperadas
    # Una desalineación entre cabecera y fila es el fallo probable de una tabla con columnas
    # dinámicas, y sin esto pasaría desapercibida: Excel no se queja de una fila más corta.
    assert {len(f) for f in filas} == {len(cabecera)}


def test_cada_peligro_del_export_cae_en_su_columna(api, datos_muestra, sin_throttling):
    """Un centro poblado promedia 3.4 peligros y el Excel solo llevaba su nivel máximo.

    Se compara contra lo que el API sirve a la tabla, que es lo que el usuario tenía delante.
    """
    listado = api.get("/api/ccpp/?clasificados=1&page_size=100").json()["results"]
    ficha = max(listado, key=lambda c: len(c["peligros"]))
    assert len(ficha["peligros"]) > 1, "la muestra debería tener algún CCPP con varios peligros"

    cabecera, filas = _hoja_ccpp(api)
    fila = dict(zip(cabecera, next(f for f in filas if f[0] == ficha["codigo"])))
    niveles = {p["nombre"]: p["nivel"] for p in ficha["peligros"]}

    for tipo in _catalogo():
        assert fila[f"{tipo.nombre} (nivel)"] == (niveles.get(tipo.nombre) or "")
    for nombre, nivel in niveles.items():
        assert f"{nombre} ({nivel} ·" in fila["Peligros"]


def test_los_peligros_del_export_respetan_el_filtro(api, datos_muestra, sin_throttling):
    """El archivo sale de una consulta filtrada, así que no puede hablar de lo que el mapa oculta.

    Sin esto, un Excel de «solo sismo» listaría también las heladas de esos mismos puntos y
    contradiría a la pantalla de la que salió.
    """
    cabecera, filas = _hoja_ccpp(api, "&peligros=sismo")
    otros = [
        cabecera.index(f"{t.nombre} (nivel)") for t in _catalogo() if t.slug != "sismo"
    ]
    columna_texto = cabecera.index("Peligros")

    assert filas, "el filtro no debería vaciar la muestra"
    assert not [f for f in filas if any(f[i] != "" for i in otros)]
    assert {f[columna_texto].split(" (")[0] for f in filas} == {"Sismo"}


def test_el_export_de_centros_poblados_no_consulta_uno_a_uno(
    api, datos_muestra, sin_throttling, django_assert_max_num_queries
):
    """El prefetch es lo que hace viable exportar miles de filas, y perderlo no se ve.

    El export seguiría dando el archivo correcto con una consulta por centro poblado; solo se
    notaría en producción, con 3,238 filas en vez de las de la muestra.
    """
    with django_assert_max_num_queries(10):
        api.get("/api/ccpp/export.xlsx?clasificados=1")


def test_las_descargas_estan_limitadas(api, norma, monkeypatch, settings):
    """30/hora en exports y PDF: son consultas caras y el resto del API va a 1000/hora.

    Se comprueba el ajuste **y** que el límite se aplica de verdad. Bajarlo a 2/hora se hace
    parcheando la clase, no el ajuste: DRF liga las tasas al definir la clase (ver
    `sin_throttling` en conftest).

    Se afirma sobre `THROTTLE_PRODUCCION` y no sobre `DEFAULT_THROTTLE_RATES`, que es lo que la
    aplicación acaba usando: desde que las tasas se pueden vaciar por variable de entorno, la
    efectiva vale `None` en cualquier entorno de desarrollo, y una prueba que la mirase pasaría a
    fallar según dónde se corra en vez de según lo que se haya cambiado.
    """
    from django.core.cache import cache

    from apps.api.throttling import DescargaThrottle

    assert settings.THROTTLE_PRODUCCION["descarga"] == "30/hour"
    monkeypatch.setattr(DescargaThrottle, "get_rate", lambda self: "2/hour")
    cache.clear()

    assert api.get("/api/normativa/export.xlsx").status_code == 200
    assert api.get("/api/normativa/export.xlsx").status_code == 200
    assert api.get("/api/normativa/export.xlsx").status_code == 429


# --- Contenidos y biblioteca ------------------------------------------------


def test_noticias_videos_eventos_y_biblioteca_responden_vacios_sin_contenido(api):
    """Un listado vacío es `results: []`, no un 404 ni un error.

    La portada pide varios de estos a la vez: si uno respondiera error con la base recién
    sembrada, la página entraría en estado de fallo por no tener contenido todavía.
    """
    for ruta in ("/api/noticias/", "/api/videos/", "/api/biblioteca/"):
        datos = api.get(ruta).json()
        assert datos["count"] == 0
        assert datos["results"] == []

    for ruta in ("/api/eventos/", "/api/biblioteca/categorias/"):
        assert api.get(ruta).status_code == 200


def _noticia(slug, fecha, destacada=False):
    """Una noticia publicada con lo mínimo obligatorio, para las pruebas de orden."""
    import datetime

    from apps.contenidos.models import Noticia

    return Noticia.objects.create(
        slug=slug,
        titulo=slug.replace("-", " ").capitalize(),
        tipo=Noticia.Tipo.NOTICIA,
        fecha=datetime.date(*fecha),
        bajada="…",
        destacada=destacada,
        estado=Noticia.Estado.PUBLICADO,
    )


def test_las_noticias_destacadas_encabezan_el_listado(api):
    """Destacadas primero y, dentro de cada grupo, lo más reciente arriba.

    La corriente del medio es la que da valor a la prueba: con el orden viejo —solo `-fecha`—
    se colaría entre las dos destacadas, que es justo lo que la portada no debe mostrar.
    """
    _noticia("destacada-antigua", (2026, 2, 17), destacada=True)
    _noticia("corriente-reciente", (2026, 6, 12))
    _noticia("destacada-reciente", (2026, 7, 28), destacada=True)

    slugs = [n["slug"] for n in api.get("/api/noticias/").json()["results"]]

    assert slugs == ["destacada-reciente", "destacada-antigua", "corriente-reciente"]


def test_el_listado_de_noticias_no_repite_ni_se_salta_filas_al_paginar(api):
    """El orden tiene que ser TOTAL: `fecha` es un DateField y los empates son la norma.

    Con `LIMIT`/`OFFSET` sobre un orden parcial PostgreSQL puede devolver la misma fila en dos
    páginas y omitir otra, sin error de por medio. `/noticias` acumula páginas con
    `useApiPaginado`, así que el visitante lo vería como noticias duplicadas al pulsar «Ver más».
    """
    esperados = {f"misma-fecha-{i}" for i in range(5)}
    for i in range(5):
        _noticia(f"misma-fecha-{i}", (2026, 5, 20))

    vistos = []
    for pagina in (1, 2, 3):
        datos = api.get(f"/api/noticias/?page_size=2&page={pagina}").json()
        vistos += [n["slug"] for n in datos["results"]]

    assert len(vistos) == len(set(vistos)), f"filas repetidas entre páginas: {vistos}"
    assert set(vistos) == esperados


def test_una_noticia_publicada_sale_en_el_listado_y_en_su_ficha(api):
    import datetime

    from apps.contenidos.models import Noticia

    noticia = Noticia.objects.create(
        slug="taller-en-canchis",
        titulo="Taller de preparación en Canchis",
        tipo=Noticia.Tipo.choices[0][0],
        fecha=datetime.date(2026, 3, 10),
        bajada="Autoridades locales y comunidades.",
        estado=Noticia.Estado.PUBLICADO,
    )

    assert api.get("/api/noticias/").json()["count"] == 1
    assert api.get(f"/api/noticias/{noticia.slug}/").json()["titulo"] == noticia.titulo
