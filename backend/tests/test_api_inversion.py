"""Ventana Inversión (PP 0068) — contrato del API y derivados.

Los derivados son el punto delicado: saldo, variación PIA-PIM, % de ejecución y % sobre el
institucional se ven bien aunque estén mal, porque son números plausibles en cualquier caso.
Cada prueba de aquí fija uno contra cifras que se pueden comprobar a mano en
`tests/datos/inversion_serie_muestra.csv`.
"""
import pytest

from tests.rutas import MUESTRA_INVERSION, MUESTRA_INVERSION_INSTITUCIONAL

pytestmark = pytest.mark.django_db


@pytest.fixture
def inversion_cargada(importar, datos_muestra):
    """Las dos series importadas y los ejercicios publicados."""
    from apps.datasets.models import DatasetUpload
    from apps.inversion.models import Ejercicio

    importar(tipo=DatasetUpload.Tipo.INVERSION, archivo=MUESTRA_INVERSION)
    importar(tipo=DatasetUpload.Tipo.INVERSION, archivo=MUESTRA_INVERSION_INSTITUCIONAL)
    Ejercicio.objects.update(visible=True)
    return Ejercicio.objects.get(anio=2026)


def test_sin_ejercicio_visible_responde_el_contrato_de_sin_datos(api, inversion_cargada):
    """Ocultar los ejercicios devuelve la ventana a su estado vacío, sin romper el cliente.

    Es el mismo contrato que servía la ventana cuando estaba diferida (ADR-D3): el frontend no
    necesita un caso especial para «hay datos pero no publicados».
    """
    from apps.inversion.models import Ejercicio

    Ejercicio.objects.update(visible=False)

    cuerpo = api.get("/api/inversion/").json()

    assert cuerpo["disponible"] is False
    assert cuerpo["motivo"]
    assert "agregados" not in cuerpo

    # El listado sirve el mismo estado vacío como una página sin resultados: el cliente no
    # necesita un caso especial para «hay datos pero todavía sin publicar».
    listado = api.get("/api/inversion/entidades/").json()
    assert listado["count"] == 0
    assert listado["results"] == []


def test_un_anio_oculto_no_cae_al_ultimo_visible(api, inversion_cargada):
    """Pedir un ejercicio que no se publica devuelve vacío, no el de otro año.

    Servir 2026 cuando piden 2025 se vería perfecto y todas las cifras serían del año
    equivocado, que es peor que no mostrar nada.
    """
    from apps.inversion.models import Ejercicio

    Ejercicio.objects.filter(anio=2025).update(visible=False)

    assert api.get("/api/inversion/?anio=2025").json()["disponible"] is False
    assert api.get("/api/inversion/?anio=2026").json()["anio"] == 2026


def test_por_defecto_sirve_el_ejercicio_mas_reciente(api, inversion_cargada):
    cuerpo = api.get("/api/inversion/").json()

    assert cuerpo["disponible"] is True
    assert cuerpo["anio"] == 2026
    assert cuerpo["ambito"] == "municipal"


def test_declara_el_corte_parcial_y_la_fuente(api, inversion_cargada):
    """Un % de ejecución de medio año no es comparable con uno de año cerrado, y el payload
    tiene que decirlo: la advertencia no puede depender de que la interfaz se acuerde."""
    cuerpo = api.get("/api/inversion/?anio=2026").json()

    assert cuerpo["es_parcial"] is True
    assert cuerpo["corte"] == "2026-06"
    assert "PREDES" in cuerpo["fuente"]

    cerrado = api.get("/api/inversion/?anio=2025").json()
    assert cerrado["es_parcial"] is False

    # La tendencia arrastra la marca punto a punto: la serie mezcla dos fuentes.
    assert [(p["anio"], p["es_parcial"]) for p in cuerpo["tendencia"]] == [
        (2025, False),
        (2026, True),
    ]


def test_el_payload_nombra_el_ejercicio_en_curso_y_su_corte_en_palabras(api, inversion_cargada):
    """`es_parcial` dice qué NO es el dato; `en_curso` y `corte_legible` dicen qué ES.

    La pantalla solo tenía la advertencia («no comparable con un ejercicio cerrado») y de ahí
    había que deducir por descarte que 2026 es el año en curso. La etiqueta la calcula el
    servidor para que pantalla, PDF y cualquier otro cliente digan lo mismo.
    """
    cuerpo = api.get("/api/inversion/?anio=2026").json()

    assert cuerpo["en_curso"] is True
    assert cuerpo["corte_legible"] == "junio de 2026"

    cerrado = api.get("/api/inversion/?anio=2025").json()
    assert cerrado["en_curso"] is False
    # Un año completo no tiene corte que nombrar: cadena vacía, no "anual" ni None.
    assert cerrado["corte_legible"] == ""


def test_un_corte_parcial_de_un_anio_pasado_no_esta_en_curso(api, inversion_cargada):
    """La trampa que obliga a que `en_curso` sea propia y no un alias de `es_parcial`.

    «Parcial» es *el devengado no cubre el año*; «en curso» es *el año no ha terminado*. Hoy
    coinciden porque el único parcial es el del año corriente, pero el día que PREDES cargue un
    corte a junio de un año ya pasado, llamarlo «en curso» sería mentir en pantalla y en el PDF.
    """
    from apps.inversion.models import Ejercicio

    Ejercicio.objects.filter(anio=2025).update(es_parcial=True, corte="2025-06")

    cuerpo = api.get("/api/inversion/?anio=2025").json()

    assert cuerpo["es_parcial"] is True
    assert cuerpo["en_curso"] is False
    assert cuerpo["corte_legible"] == "junio de 2025"


def test_un_corte_con_formato_inesperado_no_revienta(api, inversion_cargada):
    """`corte` es un CharField libre que llena una persona en el admin.

    Se prefiere devolver el valor crudo antes que una excepción: el peor caso de una etiqueta
    mal formateada es que se lea raro, y el de un 500 es que la ventana entera desaparezca.
    Los 10 caracteres de `max_length` acotan el destrozo, pero no obligan al formato.
    """
    from apps.inversion.models import Ejercicio

    Ejercicio.objects.filter(anio=2026).update(corte="jun 2026")

    cuerpo = api.get("/api/inversion/?anio=2026").json()

    assert cuerpo["corte_legible"] == "jun 2026"


def test_la_etiqueta_del_corte_viaja_en_los_cinco_sitios(api, inversion_cargada):
    """Las claves iban copiadas a mano en cinco payloads; ahora salen de un solo helper.

    Si un día falta en uno, el cliente de ese punto se queda sin poder nombrar el ejercicio y
    vuelve a la advertencia por descarte, que es justo lo que se está corrigiendo.
    """
    claves = {"anio", "corte", "corte_legible", "es_parcial", "en_curso"}

    cuerpo = api.get("/api/inversion/?anio=2026&comparar_con=2025").json()
    assert claves <= set(cuerpo)
    assert all(claves <= set(e) for e in cuerpo["ejercicios"])
    assert all(claves <= set(p) for p in cuerpo["tendencia"])
    assert claves <= set(cuerpo["comparacion"])

    ficha = api.get("/api/inversion/entidades/300684/?anio=2026").json()
    assert claves <= set(ficha)
    assert all(claves <= set(p) for p in ficha["serie"])


def test_los_derivados_de_una_entidad_cuadran_a_mano(api, inversion_cargada):
    """Cusco 2026: PIA 100 000, PIM 220 000, devengado 65 000, institucional 2 200 000."""
    filas = api.get("/api/inversion/entidades/?anio=2026").json()["results"]
    cusco = next(f for f in filas if f["codigo"] == "300684")

    assert cusco["pia"] == 100_000
    assert cusco["pim"] == 220_000
    assert cusco["devengado"] == 65_000
    assert cusco["pct_ejecucion"] == pytest.approx(65_000 / 220_000)
    assert cusco["saldo"] == 155_000
    assert cusco["variacion_pia_pim"] == 120_000
    assert cusco["pct_0068_institucional"] == pytest.approx(220_000 / 2_200_000)
    # El presupuesto de la entidad entera, que es lo que pone el 0068 en contexto. Los tres
    # importes salen tal cual del archivo, no solo el PIM del que cuelga el porcentaje.
    assert cusco["pia_institucional"] == 1_500_000
    assert cusco["pim_institucional"] == 2_200_000
    assert cusco["devengado_institucional"] == 900_000
    # Cusco solo tiene actividades en 2026; Poroy es quien aporta el proyecto.
    assert cusco["pct_proyectos"] == 0


def test_sin_total_institucional_el_porcentaje_es_nulo_y_no_cero(api, inversion_cargada):
    """`None` y `0` se dibujan distinto, y una municipalidad sin denominador no tiene un 0 %.

    Poroy no trae fila institucional de 2026 en la muestra: su porcentaje no se puede calcular.
    """
    filas = api.get("/api/inversion/entidades/?anio=2026").json()["results"]
    poroy = next(f for f in filas if f["codigo"] == "300686")

    assert poroy["pia_institucional"] is None
    assert poroy["pim_institucional"] is None
    assert poroy["devengado_institucional"] is None
    assert poroy["pct_0068_institucional"] is None


def test_el_total_institucional_y_su_porcentaje_salen_del_mismo_universo(api, inversion_cargada):
    """El total institucional del ámbito y el % sobre él suman las **mismas** entidades.

    Si el numerador sumara las dos municipalidades y el denominador solo la que tiene total, el
    porcentaje saldría inflado sin que nada lo dijera; y publicar un total institucional que no
    cuadre con el porcentaje de al lado es el mismo problema por otra vía. Solo entra Cusco.
    """
    agregados = api.get("/api/inversion/?anio=2026").json()["agregados"]

    assert agregados["entidades_con_institucional"] == 1
    assert agregados["pia_institucional"] == 1_500_000
    assert agregados["pim_institucional"] == 2_200_000
    assert agregados["devengado_institucional"] == 900_000
    assert agregados["pct_0068_institucional"] == pytest.approx(220_000 / 2_200_000)


def test_sin_ninguna_entidad_con_institucional_el_total_es_nulo(api, inversion_cargada):
    """La suma de un conjunto vacío se leería como «estas municipalidades no tienen presupuesto»."""
    from apps.inversion.models import PresupuestoEntidad

    PresupuestoEntidad.objects.filter(ejercicio__anio=2026).update(
        pia_institucional=None, pim_institucional=None, devengado_institucional=None
    )

    agregados = api.get("/api/inversion/?anio=2026").json()["agregados"]

    assert agregados["entidades_con_institucional"] == 0
    assert agregados["pia_institucional"] is None
    assert agregados["pim_institucional"] is None
    assert agregados["pct_0068_institucional"] is None
    # El presupuesto del propio programa no se ve afectado.
    assert agregados["pim"] == 535_000


def test_el_ambito_municipal_deja_fuera_al_gobierno_regional(api, inversion_cargada):
    """El GORE tiene 900 000 en la muestra: colarlo en el ranking municipal lo dominaría."""
    municipal = api.get("/api/inversion/?anio=2026").json()
    codigos = {f["codigo"] for f in api.get("/api/inversion/entidades/?anio=2026").json()["results"]}

    assert "446" not in codigos
    assert municipal["agregados"]["pim"] == 535_000  # 220 000 + 310 000 + 5 000

    regional = api.get("/api/inversion/entidades/?anio=2026&ambito=regional").json()
    assert {f["codigo"] for f in regional["results"]} == {"446"}


def test_el_reparto_por_procesos_sale_del_catalogo_vigente(api, inversion_cargada):
    """Editar el catálogo cambia el gráfico sin volver a importar ni recalcular nada."""
    from apps.inversion.models import ClasificacionActividad, ProcesoGRD

    antes = {p["slug"]: p["pim"] for p in api.get("/api/inversion/?anio=2026").json()["procesos"]}
    assert antes["preparacion"] == 200_000
    assert antes["estimacion"] == 20_000

    ClasificacionActividad.objects.filter(codigo="5005561").update(
        proceso=ProcesoGRD.objects.get(slug="respuesta"), automatico=False
    )

    despues = {p["slug"]: p["pim"] for p in api.get("/api/inversion/?anio=2026").json()["procesos"]}
    assert despues["preparacion"] == 0
    assert despues["respuesta"] == 200_000


def test_lo_que_el_catalogo_no_clasifica_se_declara_aparte(api, inversion_cargada):
    """El código desconocido de la muestra (5009999, 10 000) no se reparte ni desaparece."""
    cuerpo = api.get("/api/inversion/?anio=2026").json()

    assert cuerpo["sin_clasificar"]["pim"] == 10_000
    assert sum(p["pim"] for p in cuerpo["procesos"]) + cuerpo["sin_clasificar"]["pim"] == (
        cuerpo["agregados"]["pim"]
    )


def test_el_desglose_de_proyectos_no_pierde_ni_inventa_un_sol(api, inversion_cargada):
    """La suma del desglose es exactamente el agregado que la barra pinta.

    Es el mismo problema del mapa: un desglose al que le falta dinero se ve idéntico a uno
    correcto. La barra dice «X % en proyectos» y el cuadro de debajo dice de quién es; si las
    dos cifras se calcularan por caminos que un día divergen, nadie lo notaría en pantalla.
    """
    cuerpo = api.get("/api/inversion/?anio=2026").json()
    proyectos = cuerpo["proyectos"]

    assert sum(e["pim_proyectos"] for e in proyectos["entidades"]) == proyectos["pim"]
    assert proyectos["pim"] == cuerpo["agregados"]["pim_proyectos"]
    assert proyectos["con_proyectos"] == len(proyectos["entidades"])
    assert proyectos["de"] == cuerpo["agregados"]["entidades_en_ambito"]


def test_el_desglose_de_proyectos_solo_trae_a_quien_tiene_obra(api, inversion_cargada):
    """Una municipalidad sin proyectos no aparece con un cero: no aparece.

    «24 de 116» es la frase que la ventana quiere poder decir, y solo se sostiene si la lista
    son las que sí tienen. Una fila en cero la haría contar como si tuviera obra.
    """
    proyectos = api.get("/api/inversion/?anio=2026").json()["proyectos"]

    assert proyectos["entidades"], "la muestra tiene al menos un proyecto de inversión"
    assert all(e["pim_proyectos"] > 0 for e in proyectos["entidades"])
    assert proyectos["con_proyectos"] <= proyectos["de"]
    # Orden total: importe descendente, y el código desempata para que no baile entre peticiones.
    llaves = [(-e["pim_proyectos"], e["codigo"]) for e in proyectos["entidades"]]
    assert llaves == sorted(llaves)


def test_el_desglose_de_proyectos_respeta_el_filtro_de_provincia(api, inversion_cargada):
    """Filtrar por provincia acota el desglose y su denominador, no solo la tabla."""
    cuerpo = api.get("/api/inversion/?anio=2026&provincia=0801").json()
    proyectos = cuerpo["proyectos"]

    assert proyectos["de"] == cuerpo["agregados"]["entidades_en_ambito"]
    assert proyectos["pim"] == cuerpo["agregados"]["pim_proyectos"]
    assert all(e["provincia"] == "CUSCO" for e in proyectos["entidades"])


def test_filtra_por_provincia(api, inversion_cargada):
    filas = api.get("/api/inversion/entidades/?anio=2026&provincia=CUSCO").json()["results"]

    assert {f["provincia"] for f in filas} == {"CUSCO"}
    assert api.get("/api/inversion/entidades/?anio=2026&provincia=9999").json()["results"] == []


def test_el_export_lleva_el_corte_en_cada_fila(api, inversion_cargada, sin_throttling):
    """El Excel viaja suelto por correo: sin el corte, un 30 % de ejecución parece de un año."""
    import io

    import openpyxl

    respuesta = api.get("/api/inversion/export.xlsx?anio=2026")

    assert respuesta.status_code == 200
    hoja = openpyxl.load_workbook(io.BytesIO(respuesta.content)).active
    filas = list(hoja.iter_rows(values_only=True))
    assert filas[0][:3] == ("Ejercicio", "Corte", "Fuente")
    assert all(f[1] == "2026-06" for f in filas[1:])

    # Los tres importes institucionales viajan en el archivo: es el contexto sin el cual una
    # cifra del 0068 suelta no dice si la municipalidad prioriza la gestión del riesgo.
    cabeceras = list(filas[0])
    assert {"PIA institucional", "PIM institucional", "Devengado institucional"} <= set(cabeceras)
    cusco = next(f for f in filas[1:] if f[3] == "300684")
    assert cusco[cabeceras.index("PIA institucional")] == 1_500_000


def test_el_export_sin_datos_explica_por_que(api, inversion_cargada, sin_throttling):
    """Un Excel vacío se abre igual y parece que no hay presupuesto."""
    import io

    import openpyxl

    from apps.inversion.models import Ejercicio

    Ejercicio.objects.update(visible=False)
    respuesta = api.get("/api/inversion/export.xlsx")

    hoja = openpyxl.load_workbook(io.BytesIO(respuesta.content)).active
    assert [c.value for c in hoja[1]] == ["Motivo"]
    assert "PREDES" in hoja.cell(row=2, column=1).value


# --- Listado paginado y ranking --------------------------------------------


def test_paginacion_por_defecto_y_techo(api, inversion_cargada):
    respuesta = api.get("/api/inversion/entidades/?anio=2026&page_size=100000").json()

    assert set(respuesta) == {"count", "next", "previous", "results"}
    assert len(respuesta["results"]) <= 200


def test_el_orden_es_estable_entre_paginas(api, inversion_cargada):
    """Recorrer todas las páginas tiene que dar cada municipalidad una vez, ni más ni menos.

    Sin un orden total, dos filas empatadas pueden salir en distinto orden en dos consultas y
    la paginación repite unas y se salta otras. No se ve a simple vista y por eso se prueba:
    el desempate por código existe exactamente para esto.
    """
    codigos, pagina = [], 1
    while True:
        cuerpo = api.get(f"/api/inversion/entidades/?anio=2026&page_size=1&page={pagina}").json()
        codigos += [f["codigo"] for f in cuerpo["results"]]
        if not cuerpo["next"]:
            break
        pagina += 1

    assert len(codigos) == cuerpo["count"]
    assert len(set(codigos)) == len(codigos), "la paginación repitió alguna municipalidad"


@pytest.mark.parametrize(
    "clave, campo",
    [("pim", "pim"), ("saldo", "saldo"), ("ejecucion", "pct_ejecucion")],
)
def test_cada_ranking_ordena_de_verdad(api, inversion_cargada, clave, campo):
    """El orden lo resuelve el servidor: ordenar en el cliente ordenaría solo la página cargada."""
    filas = api.get(
        f"/api/inversion/entidades/?anio=2026&ambito=todos&ordenar={clave}&page_size=200"
    ).json()["results"]
    valores = [f[campo] for f in filas]

    # Los nulos van al final y no se comparan: «no se puede calcular» no es un valor pequeño.
    # La muestra incluye a Ccorca con PIM 0 justo para que este caso no pase de vacío.
    conocidos = [v for v in valores if v is not None]
    assert valores[: len(conocidos)] == conocidos, "los nulos tienen que quedar al final"
    assert conocidos == sorted(conocidos, reverse=True)


def test_el_listado_busca_por_nombre(api, inversion_cargada):
    filas = api.get("/api/inversion/entidades/?anio=2026&buscar=poroy").json()["results"]

    assert [f["codigo"] for f in filas] == ["300686"]


# --- Comparación entre ejercicios ------------------------------------------


def test_la_comparacion_marca_los_cortes_distintos(api, inversion_cargada):
    """2026 es un corte a junio y 2025 un año cerrado: el Δ de % de ejecución no es comparable.

    La decisión fue **mostrarlo marcado**, no ocultarlo, así que lo que se prueba es que la
    marca viaja en el dato: sin `comparable` en el payload, cada cliente tendría que
    redescubrir la regla y el Excel saldría sin advertencia.
    """
    filas = api.get(
        "/api/inversion/entidades/?anio=2026&comparar_con=2025&ambito=todos"
    ).json()["results"]
    cusco = next(f for f in filas if f["codigo"] == "300684")

    assert cusco["comparacion"]["comparable"] is False
    assert cusco["comparacion"]["anio"] == 2025
    # Cusco 2026: PIM 220 000; 2025: 170 000 (120 000 + 50 000).
    assert cusco["comparacion"]["pim"] == 170_000
    assert cusco["comparacion"]["delta_pim"] == 50_000
    assert cusco["comparacion"]["pct_delta_pim"] == pytest.approx(50_000 / 170_000)
    assert cusco["comparacion"]["delta_pct_ejecucion"] is not None


def test_una_entidad_sin_presupuesto_el_otro_ano_no_delta_a_cero(api, inversion_cargada):
    """Aparecer de la nada no es lo mismo que no haber cambiado: los deltas son null.

    El gobierno regional solo tiene presupuesto en 2026 en la muestra.
    """
    filas = api.get(
        "/api/inversion/entidades/?anio=2026&comparar_con=2025&ambito=regional"
    ).json()["results"]
    gore = next(f for f in filas if f["codigo"] == "446")

    assert gore["comparacion"]["sin_presupuesto"] is True
    assert gore["comparacion"]["delta_pim"] is None


def test_comparar_con_el_mismo_ejercicio_se_ignora(api, inversion_cargada):
    """Compararse consigo mismo daría una columna de ceros que parece un dato."""
    filas = api.get("/api/inversion/entidades/?anio=2026&comparar_con=2026").json()["results"]

    assert "comparacion" not in filas[0]


def test_la_cabecera_de_comparacion_trae_los_agregados_del_otro_ejercicio(api, inversion_cargada):
    cuerpo = api.get("/api/inversion/?anio=2026&comparar_con=2025&ambito=todos").json()

    assert cuerpo["comparacion"]["anio"] == 2025
    assert cuerpo["comparacion"]["comparable"] is False
    assert cuerpo["comparacion"]["agregados"]["pim"] == 570_000  # 120k + 50k + 400k
    assert cuerpo["comparacion"]["deltas"]["pim"] == cuerpo["agregados"]["pim"] - 570_000


# --- Ficha de una municipalidad --------------------------------------------


def test_la_ficha_trae_la_serie_y_sus_actividades(api, inversion_cargada):
    cuerpo = api.get("/api/inversion/entidades/300684/?anio=2026").json()

    assert cuerpo["entidad"]["nombre"] == "MUNICIPALIDAD PROVINCIAL DEL CUZCO"
    assert cuerpo["entidad"]["ambito"] == "provincial"
    assert [p["anio"] for p in cuerpo["serie"]] == [2025, 2026]
    assert cuerpo["serie"][-1]["es_parcial"] is True
    # Las actividades de la ficha suman exactamente el PIM de la municipalidad: si no, la
    # pantalla estaría enseñando un desglose que no cuadra con su propio total.
    assert sum(a["pim"] for a in cuerpo["actividades"]) == 220_000
    assert {a["codigo"] for a in cuerpo["actividades"]} == {"5005561", "5005571"}


def test_la_serie_omite_los_ejercicios_sin_presupuesto(api, inversion_cargada):
    """No participar del programa un año no es participar con cero soles."""
    cuerpo = api.get("/api/inversion/entidades/446/?anio=2026").json()

    assert [p["anio"] for p in cuerpo["serie"]] == [2026]


def test_la_ficha_de_una_entidad_sin_territorio_se_sirve_igual(api, inversion_cargada):
    """Cuenta en los totales, así que su ficha existe; lo que hace es declarar el hueco."""
    cuerpo = api.get("/api/inversion/entidades/309999/?anio=2026").json()

    assert cuerpo["entidad"]["sin_territorio"] is True
    assert cuerpo["entidad"]["ubigeo_distrito"] is None


def test_una_municipalidad_que_no_existe_responde_404(api, inversion_cargada):
    assert api.get("/api/inversion/entidades/000000/").status_code == 404


def test_el_export_comparado_lleva_la_advertencia_en_cada_fila(
    api, inversion_cargada, sin_throttling
):
    """En pantalla la leyenda está al lado; el Excel viaja solo por correo.

    Es la mitigación de haber elegido mostrar el Δ marcado en vez de suprimirlo: sin esta
    columna, la cifra circula sin su advertencia en cuanto alguien reenvía el archivo.
    """
    import io

    import openpyxl

    respuesta = api.get("/api/inversion/export.xlsx?anio=2026&comparar_con=2025")

    hoja = openpyxl.load_workbook(io.BytesIO(respuesta.content)).active
    filas = list(hoja.iter_rows(values_only=True))
    cabeceras = list(filas[0])
    assert "Comparabilidad" in cabeceras
    assert "Δ PIM" in cabeceras
    columna = cabeceras.index("Comparabilidad")
    assert all("no es comparable" in f[columna] for f in filas[1:])


# --- El mapa (ADR-D6) -----------------------------------------------------------------------
#
# El riesgo del coroplético no es que se vea mal: es que se vea bien y falte dinero. Pintar el
# presupuesto de una municipalidad provincial sobre su distrito capital, o dejar caer en
# silencio a las entidades sin territorio, produce un mapa impecable con cifras que no suman.
# Las dos primeras pruebas son la contabilidad completa del mapa, y son las que hacen cumplir
# ADR-D6: lo pintado más lo declarado es exactamente el total del ámbito.


def test_el_mapa_distrital_solo_pinta_lo_que_puede_atribuir(api, inversion_cargada):
    cuerpo = api.get("/api/inversion/mapa/?anio=2026&nivel=distrital").json()

    assert cuerpo["disponible"] is True
    assert cuerpo["nivel"] == "distrital"

    # Un polígono por distrito con municipalidad distrital, sin repetir.
    ubigeos = [f["ubigeo"] for f in cuerpo["filas"]]
    assert len(ubigeos) == len(set(ubigeos))
    assert all(len(u) == 6 for u in ubigeos)

    # La provincial de Cusco y la entidad sin territorio no se pintan; su importe se declara.
    assert "300684" not in {f["codigo_entidad"] for f in cuerpo["filas"]}
    assert cuerpo["no_ubicado"]["entidades"] == 2
    assert cuerpo["no_ubicado"]["motivo"]


def test_el_mapa_no_pierde_ni_inventa_un_sol(api, inversion_cargada):
    """La prueba de ADR-D6, en las dos direcciones.

    A nivel provincial el mapa cubre todo el ámbito; a nivel distrital deja fuera lo que no
    puede ubicar, y eso sale en `no_ubicado`. En los dos casos, pintado + declarado == total.
    """
    total = api.get("/api/inversion/?anio=2026").json()["agregados"]["pim"]

    distrital = api.get("/api/inversion/mapa/?anio=2026&nivel=distrital").json()
    assert (
        sum(f["pim"] for f in distrital["filas"]) + distrital["no_ubicado"]["pim"]
        == pytest.approx(total)
    )

    provincial = api.get("/api/inversion/mapa/?anio=2026&nivel=provincial").json()
    assert (
        sum(f["pim"] for f in provincial["filas"]) + provincial["no_ubicado"]["pim"]
        == pytest.approx(total)
    )
    # Un ubigeo de provincia tiene cuatro dígitos, y la fila agrega varias municipalidades.
    assert all(len(f["ubigeo"]) == 4 for f in provincial["filas"])
    assert any(f["entidades"] > 1 for f in provincial["filas"])


def test_el_mapa_arrastra_las_cuatro_metricas_en_cada_fila(api, inversion_cargada):
    """Conmutar entre PIA, PIM, devengado y % de ejecución no puede disparar otra petición: si
    lo hiciera, dos métricas del mismo mapa podrían venir de ejercicios distintos."""
    fila = next(
        f
        for f in api.get("/api/inversion/mapa/?anio=2026&nivel=distrital").json()["filas"]
        if f["pim"] > 0
    )

    assert {"pia", "pim", "devengado", "saldo", "pct_ejecucion"} <= set(fila)
    assert fila["saldo"] == pytest.approx(fila["pim"] - fila["devengado"])
    assert fila["pct_ejecucion"] == pytest.approx(fila["devengado"] / fila["pim"])


def test_una_municipalidad_con_pim_cero_no_tiene_porcentaje_de_ejecucion(api, inversion_cargada):
    """Sin denominador no hay avance. Un 0 % aquí diría «no gastó lo que tenía», y no tenía."""
    fila = next(
        f
        for f in api.get("/api/inversion/mapa/?anio=2026&nivel=distrital").json()["filas"]
        if f["pim"] == 0
    )

    assert fila["pct_ejecucion"] is None


def test_un_ambito_sin_geografia_lo_dice_en_vez_de_quedarse_en_blanco(api, inversion_cargada):
    """Las provinciales no se pueden pintar por distrito: el mapa lo explica."""
    cuerpo = api.get("/api/inversion/mapa/?anio=2026&nivel=distrital&ambito=provincial").json()

    assert cuerpo["filas"] == []
    assert cuerpo["no_ubicado"]["pim"] > 0
    assert cuerpo["no_ubicado"]["motivo"]


def test_los_cortes_de_la_escala_son_crecientes_y_no_revientan_con_pocas_filas(
    api, inversion_cargada
):
    """Los quintiles se calculan sobre lo pintado. Con dos filas siguen saliendo cuatro cortes:
    la leyenda tiene que poder dibujarse igual, aunque varios tramos queden vacíos."""
    cortes = api.get("/api/inversion/mapa/?anio=2026&nivel=distrital").json()["cortes"]

    assert set(cortes) == {"pia", "pim", "devengado"}
    for metrica, valores in cortes.items():
        assert len(valores) == 4, metrica
        assert valores == sorted(valores), metrica


def test_los_cuartiles_salen_por_indice_igual_que_los_quintiles_del_mapa(api, inversion_cargada):
    """La caja y la leyenda del mapa tienen que medir con la misma regla.

    Son estadísticos distintos —quintiles el coroplético, cuartiles la caja— y no pueden dar el
    mismo número. Pero calculados con métodos distintos, uno por índice y otro interpolando,
    nadie sabría si la diferencia es del dato o del método. `_cortes` no interpola; esto tampoco.
    """
    from apps.inversion import consultas

    filas = [{"nombre": f"D{i}", "pim": float(i)} for i in range(1, 9)]  # 1..8

    d = consultas.distribucion(filas, "pim")

    # Por índice sobre [1..8]: n*k//4 = 2, 4, 6 ⇒ los valores 3, 5 y 7.
    assert (d["q1"], d["mediana"], d["q3"]) == (3.0, 5.0, 7.0)
    assert d["n"] == 8


def test_una_serie_constante_no_produce_atipicos_ni_revienta(api, inversion_cargada):
    """IQR cero es el caso que rompe un boxplot escrito a la ligera.

    Con todos los valores iguales, el rango intercuartílico vale 0 y la regla de Tukey deja los
    bigotes pegados a la caja: **cero atípicos**, no todos ellos.
    """
    from apps.inversion import consultas

    d = consultas.distribucion([{"nombre": f"D{i}", "pim": 500.0} for i in range(6)], "pim")

    assert d["q1"] == d["mediana"] == d["q3"] == 500.0
    assert d["bigote_min"] == d["bigote_max"] == 500.0
    assert d["atipicos"] == []


def test_los_ceros_se_cuentan_en_vez_de_desaparecer(api, inversion_cargada):
    """Un cero no cabe en un eje logarítmico, y descartarlo en silencio falsea la caja.

    Los cuartiles se calculan sobre **todos** los valores, ceros incluidos, porque un distrito
    sin presupuesto es un dato; lo que el dibujo no puede es colocarlo, así que se cuenta aparte
    para poder declararlo.
    """
    from apps.inversion import consultas

    filas = [{"nombre": "A", "pim": 0.0}, {"nombre": "B", "pim": 0.0}] + [
        {"nombre": f"D{i}", "pim": float(i * 100)} for i in range(1, 7)
    ]

    d = consultas.distribucion(filas, "pim")

    assert d["n"] == 8
    assert d["ceros"] == 2


def test_el_porcentaje_de_ejecucion_ignora_los_nulos_y_no_los_toma_por_cero(
    api, inversion_cargada
):
    """`pct_ejecucion` es `None` cuando el PIM es 0: no hay avance que calcular.

    Contarlo como 0 % metería en la caja distritos que no tienen porcentaje, y la mediana
    bajaría sin que nada fallara — el fallo silencioso de este endpoint.
    """
    from apps.inversion import consultas

    filas = [
        {"nombre": "A", "pct_ejecucion": None},
        {"nombre": "B", "pct_ejecucion": 0.4},
        {"nombre": "C", "pct_ejecucion": 0.6},
    ]

    d = consultas.distribucion(filas, "pct_ejecucion")

    assert d["n"] == 2
    assert d["ceros"] == 0


def test_los_atipicos_vienen_con_nombre_y_de_mayor_a_menor(api, inversion_cargada):
    """Un punto suelto en el gráfico no dice nada; con el nombre, dice quién es PICHARI."""
    from apps.inversion import consultas

    filas = [{"nombre": f"D{i}", "pim": 10.0} for i in range(10)]
    filas += [{"nombre": "GRANDE", "pim": 5_000.0}, {"nombre": "ENORME", "pim": 90_000.0}]

    d = consultas.distribucion(filas, "pim")

    assert [a["nombre"] for a in d["atipicos"]] == ["ENORME", "GRANDE"]
    assert d["atipicos"][0]["valor"] == 90_000.0


def test_el_mapa_publica_la_distribucion_de_las_cuatro_metricas(api, inversion_cargada):
    """Cambiar de métrica no dispara otra petición, así que las cuatro cajas viajan juntas.

    Y `n + ceros` no sale de la nada: es la misma cuenta de polígonos que el mapa dice haber
    pintado. Si divergieran, la caja describiría un conjunto distinto del que enseña el mapa.
    """
    cuerpo = api.get("/api/inversion/mapa/?anio=2026&nivel=distrital").json()

    assert set(cuerpo["distribucion"]) == {"pia", "pim", "devengado", "pct_ejecucion"}
    pim = cuerpo["distribucion"]["pim"]
    assert set(pim) >= {"n", "ceros", "q1", "mediana", "q3", "bigote_min", "bigote_max",
                        "atipicos", "frase"}
    assert pim["n"] == cuerpo["poligonos"]["pintados"]
    assert pim["q1"] <= pim["mediana"] <= pim["q3"]
    assert pim["frase"]


def test_los_pies_del_mapa_no_dicen_dos_veces_lo_mismo(api, inversion_cargada):
    """Los 13 distritos capital salían explicados en los dos pies, con palabras distintas.

    Junto al párrafo de quintiles y a la entradilla eran ~150 palabras alrededor de un mapa, y
    el lector no llegaba al final. Cada pie dice ahora **su** hecho: uno el dinero que no se
    pinta, otro los polígonos en blanco.
    """
    cuerpo = api.get("/api/inversion/mapa/?anio=2026&nivel=distrital").json()
    no_ubicado = cuerpo["no_ubicado"]["motivo"]
    poligonos = cuerpo["poligonos"]["motivo"]

    # El plural entre paréntesis leía como una circular administrativa.
    assert "(es)" not in no_ubicado and "(es)" not in poligonos
    # Qué gestiona una municipalidad provincial se explica UNA vez, en el pie del dinero.
    assert "provincia" in no_ubicado
    assert "gestiona" not in poligonos
    assert len(no_ubicado.split()) <= 32
    assert len(poligonos.split()) <= 20


def test_el_mapa_se_acota_por_provincia(api, inversion_cargada):
    cuerpo = api.get("/api/inversion/mapa/?anio=2026&nivel=distrital&provincia=CUSCO").json()

    assert cuerpo["filas"]
    assert all(f["ubigeo"].startswith("0801") for f in cuerpo["filas"])


def test_el_mapa_sin_ejercicio_visible_sirve_el_mismo_estado_vacio(api, inversion_cargada):
    from apps.inversion.models import Ejercicio

    Ejercicio.objects.update(visible=False)

    cuerpo = api.get("/api/inversion/mapa/").json()

    assert cuerpo["disponible"] is False
    assert cuerpo["motivo"]
