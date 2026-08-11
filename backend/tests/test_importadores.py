"""Importadores de Excel (spec 03 y plan de pruebas 08).

Cada caso protege una decisión que salió de la auditoría de los datos reales y que **un refactor
puede deshacer sin que nada falle a la vista**: los datos siguen entrando, el importador sigue
diciendo «activo», y la cifra que se publica es otra. Esa es la razón de ser de este módulo.
"""
from django.utils.text import slugify

import openpyxl
import pytest

from tests.rutas import (
    MUESTRA_FRECUENCIA,
    MUESTRA_INVERSION,
    MUESTRA_INVERSION_INSTITUCIONAL,
    MUESTRA_NIVEL,
)

pytestmark = pytest.mark.django_db


def _advertencias(upload) -> str:
    """Todas las advertencias en un solo texto, para buscar por contenido."""
    return "\n".join(upload.log.get("advertencias", []))


# --- Nivel de peligro ------------------------------------------------------


def test_importa_y_queda_activo(importar):
    from apps.datasets.models import DatasetUpload
    from apps.peligros.models import ClasificacionPeligro

    upload = importar()

    assert upload.estado == DatasetUpload.Estado.ACTIVO
    assert ClasificacionPeligro.objects.exists()
    assert upload.log["centros_poblados"] > 0


@pytest.mark.parametrize(
    "hoja, nombre, slug",
    [
        ("Lluvias", "Lluvias intensas", "lluvias_intensas"),
        ("Incendios Forestales", "Incendios forestales", "incendios_forestales"),
    ],
)
def test_el_nombre_del_peligro_sale_del_catalogo_no_del_titulo_de_la_hoja(
    importar, hoja, nombre, slug
):
    """Dos hojas se llaman distinto de lo que dice su columna PELIGRO.

    El catálogo canónico resuelve las dos. Si un refactor derivara el peligro del título de la
    hoja, aparecerían un «Lluvias» y un «Incendios Forestales» fantasma, con slug distinto, y el
    visor perdería esas dos capas del semáforo sin que ningún otro sitio fallara.
    """
    from apps.peligros.models import ClasificacionPeligro, TipoPeligro

    importar()

    assert TipoPeligro.objects.filter(slug=slug, nombre=nombre).exists()
    assert ClasificacionPeligro.objects.filter(tipo_peligro__slug=slug).exists()
    # Ni el nombre tal como lo escribe la hoja, ni el slug que saldría de aplicarle `slugify`.
    assert not TipoPeligro.objects.filter(nombre=hoja).exists()
    assert not TipoPeligro.objects.filter(slug=slugify(hoja)).exists()


def test_un_peligro_fuera_del_catalogo_se_descarta_citando_la_fila(importar, tmp_path):
    """La discrepancia hoja/columna está en el catálogo; un valor **desconocido** no.

    Se descarta la clasificación y se dice dónde estaba, que es lo que permite corregir el Excel.
    """
    inventado = tmp_path / "inventado.xlsx"
    wb = openpyxl.load_workbook(MUESTRA_NIVEL)
    wb["Sismo"].cell(row=2, column=11).value = "Terremoto"
    wb.save(inventado)

    upload = importar(archivo=inventado)

    assert "PELIGRO 'Terremoto' no está en el catálogo" in _advertencias(upload)


def test_ningun_slug_de_peligro_lleva_guion():
    """Es la clave de las propiedades `nivel_<slug>` del tile.

    Con guion medio el visor deja de pintar el semáforo y **ninguna otra prueba lo nota**: el
    API responde, la tabla lista y el mapa sale en gris.
    """
    from apps.peligros.models import TipoPeligro

    slugs = list(TipoPeligro.objects.values_list("slug", flat=True))

    assert slugs, "el seed de catálogos no dejó tipos de peligro"
    assert not [s for s in slugs if "-" in s]


def test_filas_sin_nivel_se_descartan_y_se_cuentan(importar):
    """Sin nivel no hay semáforo, y asumir «1» sería inventar el dato."""
    from apps.peligros.models import ClasificacionPeligro

    upload = importar()

    assert upload.log["descartadas_sin_nivel"] > 0
    assert "sin NIVEL_PELI" in _advertencias(upload)
    assert not ClasificacionPeligro.objects.filter(nivel=None).exists()
    assert set(ClasificacionPeligro.objects.values_list("nivel", flat=True)) <= {1, 2, 3, 4}


def test_filas_sin_codigo_se_descartan_con_aviso(importar):
    upload = importar()

    assert upload.log["descartadas_sin_codigo"] == 2
    assert "sin CODIGO" in _advertencias(upload)


def test_al_deduplicar_gana_el_valor_no_vacio(importar):
    """SICUANI trae DISTRITO en blanco en una hoja y lleno en otra.

    Quedarse con el primero que aparece deja el resultado a merced del orden de las hojas: el
    centro poblado más poblado de Canchis se quedaría sin distrito.
    """
    from apps.territorio.models import CentroPoblado

    importar()
    sicuani = CentroPoblado.objects.get(codigo="0806010001")

    assert sicuani.distrito is not None
    assert sicuani.distrito.nombre.upper().startswith("SICUANI")


def test_se_normalizan_las_dos_grafias_de_la_fuente(importar):
    """`CENEPRED_SIGRID` y `SIGRID_CENEPRED` son la misma fuente escrita al revés.

    Sin normalizar quedan dos filas en el catálogo de fuentes y la ficha de un centro poblado
    cita una u otra según de qué hoja salió su clasificación.
    """
    from apps.peligros.models import Fuente

    importar()
    nombres = set(Fuente.objects.values_list("nombre", flat=True))

    assert "CENEPRED_SIGRID" not in nombres
    assert "SIGRID_CENEPRED" in nombres


def test_ubigeo_derivado_del_codigo_inei(importar):
    """El Excel no trae ubigeo: sale de los 6 primeros dígitos del código de 10."""
    from apps.territorio.models import CentroPoblado

    importar()

    for ccpp in CentroPoblado.objects.select_related("distrito__provincia")[:20]:
        assert ccpp.codigo.startswith(ccpp.distrito.ubigeo)
        assert ccpp.distrito.ubigeo.startswith(ccpp.distrito.provincia.ubigeo)


# --- Frecuencia de emergencias ---------------------------------------------


def test_adr_d1_cusco_se_guarda_como_declarado(datos_muestra):
    """El caso que motivó ADR-D1.

    CUSCO trae los cuatro subtotales llenos (TOTAL 134) y **todas** las columnas de evento
    vacías. Descartar los `TOT_*` —lo que hacía la primera versión— dejaba a la capital regional
    publicando «0 emergencias», que es peor que no publicar nada.
    """
    from apps.peligros.models import FrecuenciaEmergencia, TotalDeclaradoEmergencias
    from apps.territorio.models import Distrito

    cusco = Distrito.objects.get(ubigeo="080101")
    declarados = TotalDeclaradoEmergencias.objects.filter(distrito=cusco)

    assert not FrecuenciaEmergencia.objects.filter(distrito=cusco).exists()
    assert declarados.count() == 4
    assert sum(declarados.values_list("total", flat=True)) == 134


def test_cuando_hay_desglose_el_declarado_no_se_guarda(datos_muestra):
    """Regla literal del ADR-D1: el desglose manda y el declarado solo se registra si descuadra.

    Guardar los dos haría ambiguo el `desglose_disponible` del API, que es lo que la interfaz usa
    para decidir si dibuja el gráfico o explica que la fuente no desagrega.
    """
    from apps.peligros.models import FrecuenciaEmergencia, TotalDeclaradoEmergencias
    from apps.territorio.models import Distrito

    ollanta = Distrito.objects.get(nombre__iexact="OLLANTAYTAMBO")

    assert FrecuenciaEmergencia.objects.filter(distrito=ollanta).exists()
    assert not TotalDeclaradoEmergencias.objects.filter(distrito=ollanta).exists()


def test_descuadre_de_subtotal_prevalece_el_desglose(datos_muestra):
    """SANGARARÁ declara 18 meteorológicos y su desglose suma 27.

    Se conserva el desglose —es el dato con detalle— y la diferencia queda en el log para que
    PREDES pueda llevársela a la fuente. Silenciarla sería publicar una cifra que la propia
    fuente contradice.
    """
    from apps.peligros.models import FrecuenciaEmergencia
    from apps.territorio.models import Distrito

    upload = datos_muestra["frecuencia"]
    sangarara = Distrito.objects.get(nombre__iexact="SANGARARA")
    meteorologicos = FrecuenciaEmergencia.objects.filter(
        distrito=sangarara, tipo_evento__categoria__slug="meteorologico"
    )

    assert sum(meteorologicos.values_list("conteo", flat=True)) == 27
    assert "el subtotal declarado de 'meteorologico' es 18" in _advertencias(upload)
    assert "Prevalece el desglose" in _advertencias(upload)


def test_distrito_del_padron_sin_fila_avisa_y_no_aborta(datos_muestra):
    """En el archivo real es ACOMAYO; en la muestra, cualquier distrito del padrón sin fila.

    Lo que se prueba es la política, no el nombre: un distrito ausente **avisa y no aborta**.
    Abortar dejaría los otros 111 sin importar, y asumir 0 lo publicaría como distrito sin
    emergencias. El caso concreto de ACOMAYO va en la prueba `lento` sobre el archivo completo.
    """
    from apps.datasets.models import DatasetUpload
    from apps.territorio.models import Distrito

    upload = datos_muestra["frecuencia"]
    con_fila = {"CUSCO", "SANGARARA", "MOLLEPATA", "OLLANTAYTAMBO", "ACOPIA"}
    esperados = [
        d.nombre for d in Distrito.objects.all() if d.nombre.upper() not in con_fila
    ]

    assert upload.estado == DatasetUpload.Estado.ACTIVO
    assert esperados, "la muestra no dejó ningún distrito sin fila que comprobar"
    assert "no tienen fila en el Excel" in _advertencias(upload)
    assert len(upload.log["distritos_sin_fila"]) == len(esperados)
    for nombre in esperados:
        assert any(nombre in fila for fila in upload.log["distritos_sin_fila"])


def test_fila_vacia_no_es_lo_mismo_que_declarar_subtotales(datos_muestra):
    """ACOPIA tiene fila y ni un solo número: 21 distritos así en el archivo real.

    No es «declara subtotales sin desagregar» (ADR-D1) ni «declara cero». No hay dato, y el API
    responderá 404 igual que para ACOMAYO — así que tiene que quedar contado aparte, o 21
    distritos sin información se esconden detrás de un aviso que dice otra cosa.
    """
    from apps.peligros.models import FrecuenciaEmergencia, TotalDeclaradoEmergencias
    from apps.territorio.models import Distrito

    upload = datos_muestra["frecuencia"]
    acopia = Distrito.objects.get(nombre__iexact="ACOPIA")

    assert not FrecuenciaEmergencia.objects.filter(distrito=acopia).exists()
    assert not TotalDeclaradoEmergencias.objects.filter(distrito=acopia).exists()
    assert any("ACOPIA" in a for a in upload.log["distritos_sin_dato"])
    assert "ni un solo dato" in _advertencias(upload)


def test_el_rango_de_fechas_se_normaliza_pero_no_se_reinterpreta(datos_muestra):
    """`2003 - 2022` y `2003-2022` son el mismo periodo escrito de dos formas (23 variantes).

    Se quitan los espacios y nada más: convertirlo a dos enteros perdería los valores que no son
    un rango simple, y el periodo es lo que impide comparar distritos a ciegas.
    """
    from apps.peligros.models import TotalDeclaradoEmergencias

    declarado = TotalDeclaradoEmergencias.objects.filter(distrito__ubigeo="080101").first()

    assert declarado.rango_fecha == "2003-2022"


def test_la_fuente_invertida_se_normaliza_y_se_registra(datos_muestra):
    from apps.peligros.models import TotalDeclaradoEmergencias

    upload = datos_muestra["frecuencia"]
    declarado = TotalDeclaradoEmergencias.objects.filter(distrito__ubigeo="080101").first()

    assert declarado.fuente == "SIGRID_CENEPRED"
    assert "se normaliza a 'SIGRID_CENEPRED'" in _advertencias(upload)


# --- Garantías transversales ----------------------------------------------


def test_todo_o_nada_un_archivo_invalido_no_toca_los_datos_activos(importar, tmp_path):
    """Un Excel con las columnas cambiadas deja los datos anteriores intactos.

    Es la garantía que permite a PREDES probar un archivo sin miedo: o entra completo, o no entra.
    """
    from apps.datasets.models import DatasetUpload
    from apps.peligros.models import ClasificacionPeligro

    importar()
    antes = ClasificacionPeligro.objects.count()
    assert antes > 0

    roto = tmp_path / "roto.xlsx"
    wb = openpyxl.load_workbook(MUESTRA_NIVEL)
    wb["Sismo"].cell(row=1, column=4).value = "COD"  # se esperaba CODIGO
    wb.save(roto)

    fallido = importar(archivo=roto)

    assert fallido.estado == DatasetUpload.Estado.ERROR
    assert "columnas inesperadas" in fallido.log["error"]
    assert ClasificacionPeligro.objects.count() == antes


def test_un_archivo_sin_hojas_conocidas_falla_con_mensaje_util(importar, tmp_path):
    from apps.datasets.models import DatasetUpload

    ajeno = tmp_path / "ajeno.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Hoja1"
    wb.save(ajeno)

    upload = importar(archivo=ajeno)

    assert upload.estado == DatasetUpload.Estado.ERROR
    assert "ninguna de las 9 hojas esperadas" in upload.log["error"]


def test_reimportar_reemplaza_en_vez_de_duplicar(importar):
    """Dos importaciones del mismo archivo dan los mismos conteos, y la anterior queda marcada."""
    from apps.datasets.models import DatasetUpload
    from apps.peligros.models import ClasificacionPeligro

    primera = importar()
    conteo = ClasificacionPeligro.objects.count()

    segunda = importar()
    primera.refresh_from_db()

    assert ClasificacionPeligro.objects.count() == conteo
    assert primera.estado == DatasetUpload.Estado.REEMPLAZADO
    assert segunda.reemplaza_a_id == primera.pk
    assert not ClasificacionPeligro.objects.filter(dataset_upload=primera).exists()


def test_el_padron_no_se_borra_al_reimportar(importar):
    """Los centros poblados son el padrón: medidas y frecuencias los referencian.

    Las clasificaciones se reemplazan; los centros poblados se actualizan en su sitio.
    """
    from apps.territorio.models import CentroPoblado

    importar()
    pks = set(CentroPoblado.objects.values_list("pk", flat=True))

    importar()

    assert set(CentroPoblado.objects.values_list("pk", flat=True)) == pks


def test_frecuencia_sin_la_hoja_esperada_falla_citandola(importar, tmp_path):
    from apps.datasets.models import DatasetUpload

    sin_hoja = tmp_path / "sin_hoja.xlsx"
    wb = openpyxl.load_workbook(MUESTRA_FRECUENCIA)
    wb["NºEMERGENCIAS"].title = "Datos"
    wb.save(sin_hoja)

    upload = importar(tipo=DatasetUpload.Tipo.FRECUENCIA, archivo=sin_hoja)

    assert upload.estado == DatasetUpload.Estado.ERROR
    assert "NºEMERGENCIAS" in upload.log["error"]


# --- Inversión (PP 0068) ---------------------------------------------------


@pytest.fixture
def importar_inversion(importar, datos_muestra):
    """Importa una serie de inversión sobre el territorio ya sembrado por las muestras.

    Depende de `datos_muestra` porque el importador resuelve la municipalidad contra el padrón
    de distritos, y ese padrón lo crea la importación de niveles de peligro.
    """
    from apps.datasets.models import DatasetUpload

    def _importar(archivo=MUESTRA_INVERSION):
        return importar(tipo=DatasetUpload.Tipo.INVERSION, archivo=archivo)

    return _importar


def test_inversion_importa_las_dos_series(importar_inversion):
    from apps.datasets.models import DatasetUpload
    from apps.inversion.models import PresupuestoEntidad

    serie = importar_inversion()
    institucional = importar_inversion(MUESTRA_INVERSION_INSTITUCIONAL)

    assert serie.estado == DatasetUpload.Estado.ACTIVO
    assert serie.log["forma"] == "programa"
    assert institucional.estado == DatasetUpload.Estado.ACTIVO
    assert institucional.log["forma"] == "institucional"

    cusco = PresupuestoEntidad.objects.get(entidad__codigo="300684", ejercicio__anio=2026)
    assert cusco.pim == 220000  # 200000 + 20000
    assert cusco.pim_institucional == 2200000


def test_inversion_no_pisa_lo_que_no_trae_el_archivo(importar_inversion):
    """Cargar los tres archivos en cualquier orden tiene que dejar el mismo resultado.

    La serie de totales institucionales no lleva provincia ni distrito, y escribir `None` desde
    ahí borraba el territorio que había resuelto la serie del programa: 120 municipalidades se
    quedaban «sin territorio» solo por el orden de carga. Al revés pasa lo mismo con el PIM.
    """
    from apps.inversion.models import EntidadEjecutora, PresupuestoEntidad

    importar_inversion(MUESTRA_INVERSION_INSTITUCIONAL)
    importar_inversion()

    entidad = EntidadEjecutora.objects.get(codigo="300684")
    assert entidad.distrito_id is not None, "la serie del programa tiene que resolver el distrito"

    presupuesto = PresupuestoEntidad.objects.get(entidad=entidad, ejercicio__anio=2026)
    assert presupuesto.pim == 220000
    assert presupuesto.pim_institucional == 2200000, (
        "la segunda carga borró el total institucional de la primera"
    )


def test_inversion_reemplaza_el_ejercicio_sin_duplicar(importar_inversion):
    """La misma carga dos veces deja los mismos números, no el doble."""
    from apps.inversion.models import PresupuestoActividad, PresupuestoEntidad

    importar_inversion()
    filas = PresupuestoActividad.objects.count()
    importar_inversion()

    assert PresupuestoActividad.objects.count() == filas
    assert PresupuestoEntidad.objects.get(
        entidad__codigo="300684", ejercicio__anio=2026
    ).pim == 220000


def test_inversion_clasifica_por_actividad_y_deja_ver_lo_que_no_conoce(importar_inversion):
    """El catálogo se descubre solo, pero un código desconocido no se reparte: se declara."""
    from apps.inversion.models import ClasificacionActividad

    importar_inversion()

    # Actividad conocida: la semilla ya le dio su proceso.
    brigadas = ClasificacionActividad.objects.get(codigo="5005561")
    assert brigadas.proceso.slug == "preparacion"

    # Proyecto: se clasifica por el proyecto, no por su acción de obra (4000122, que se repite
    # en obras de procesos distintos), y arranca en el proceso propuesto por defecto.
    proyecto = ClasificacionActividad.objects.get(codigo="2534565")
    assert proyecto.origen == ClasificacionActividad.Origen.PROYECTO
    assert proyecto.proceso.slug == "prevencion_reduccion"
    assert proyecto.automatico is True
    assert not ClasificacionActividad.objects.filter(codigo="4000122").exists()

    # Actividad desconocida: entra al catálogo sin proceso, para que se vea que falta.
    desconocida = ClasificacionActividad.objects.get(codigo="5009999")
    assert desconocida.proceso is None


def test_inversion_avisa_de_las_municipalidades_fuera_del_padron(importar_inversion):
    """Una municipalidad que no casa con el padrón cuenta en los totales, pero se declara.

    Descartarla en silencio restaría presupuesto sin que nada lo dijera; asignarla a un
    distrito cualquiera contaminaría el cruce con peligros.
    """
    from apps.inversion.models import EntidadEjecutora

    upload = importar_inversion()

    fantasma = EntidadEjecutora.objects.get(codigo="309999")
    assert fantasma.distrito_id is None
    assert fantasma.sin_territorio is True
    assert "sin territorio" in _advertencias(upload)
    assert "PUEBLO INEXISTENTE" in _advertencias(upload)


def test_inversion_no_publica_sola(importar_inversion):
    """Importar no enciende la ventana: publicar es una decisión editorial, y se avisa."""
    from apps.inversion.models import Ejercicio

    upload = importar_inversion()

    assert not Ejercicio.objects.filter(visible=True).exists()
    assert "visible" in _advertencias(upload)


def test_inversion_marca_el_corte_parcial_y_su_fuente(importar_inversion):
    """2026 llega a mitad de año y de la base del cliente; 2025, cerrado y del MEF.

    Si la serie se etiqueta entera con una sola fuente, el último punto de la tendencia dice
    que viene del MEF cuando no es así.
    """
    from apps.inversion.models import Ejercicio

    importar_inversion()

    parcial = Ejercicio.objects.get(anio=2026)
    assert parcial.es_parcial is True
    assert parcial.corte == "2026-06"
    assert parcial.fuente == Ejercicio.Fuente.CLIENTE

    cerrado = Ejercicio.objects.get(anio=2025)
    assert cerrado.es_parcial is False
    assert cerrado.fuente == Ejercicio.Fuente.MEF


def test_inversion_rechaza_un_archivo_que_no_es_ninguna_de_las_series(importar_inversion):
    """El error tiene que decir qué columnas faltan, no un «no se pudo importar» mudo."""
    from apps.datasets.models import DatasetUpload

    upload = importar_inversion(MUESTRA_FRECUENCIA)

    assert upload.estado == DatasetUpload.Estado.ERROR
    assert "Base 2026" not in upload.log["error"]
    assert "hoja" in upload.log["error"].lower()


def test_inversion_rechaza_un_devengado_mayor_que_el_pim(importar_inversion, tmp_path):
    """El SIAF bloquea devengar por encima del PIM, así que un archivo así está mal en origen.

    Sin esta comprobación entraría sin protestar y produciría un avance de ejecución superior al
    100 % — una cifra que se ve plausible en una tabla y que nadie sabría de dónde salió.
    """
    from apps.datasets.models import DatasetUpload
    from apps.inversion.models import PresupuestoEntidad

    lineas = MUESTRA_INVERSION.read_text(encoding="utf-8").splitlines()
    # La segunda fila lleva PIA 100000, PIM 120000 y devengado 90000: se le pasa el devengado.
    lineas[1] = lineas[1].replace(",100000.00,120000.00,90000.00,", ",100000.00,120000.00,130000.00,")
    roto = tmp_path / "devengado_sobre_pim.csv"
    roto.write_text("\n".join(lineas), encoding="utf-8")

    upload = importar_inversion(roto)

    assert upload.estado == DatasetUpload.Estado.ERROR
    assert "devengado" in upload.log["error"].lower()
    assert "Fila 2" in upload.log["error"]
    # El reemplazo es atómico: un archivo rechazado no deja nada a medio escribir.
    assert not PresupuestoEntidad.objects.exists()


def test_inversion_rechaza_importes_negativos(importar_inversion, tmp_path):
    """Un PIM negativo no existe: es un error de extracción, no un recorte presupuestal."""
    from apps.datasets.models import DatasetUpload
    from apps.inversion.models import PresupuestoEntidad

    lineas = MUESTRA_INVERSION.read_text(encoding="utf-8").splitlines()
    lineas[1] = lineas[1].replace(",100000.00,120000.00,90000.00,", ",100000.00,-120000.00,0.00,")
    roto = tmp_path / "pim_negativo.csv"
    roto.write_text("\n".join(lineas), encoding="utf-8")

    upload = importar_inversion(roto)

    assert upload.estado == DatasetUpload.Estado.ERROR
    assert "negativo" in upload.log["error"].lower()
    assert not PresupuestoEntidad.objects.exists()


def test_inversion_enumera_varias_filas_malas_en_un_solo_mensaje(importar_inversion, tmp_path):
    """Quien sube el archivo lo corrige una vez, no una fila por intento."""
    from apps.datasets.models import DatasetUpload

    lineas = MUESTRA_INVERSION.read_text(encoding="utf-8").splitlines()
    lineas[1] = lineas[1].replace(",100000.00,120000.00,90000.00,", ",100000.00,120000.00,130000.00,")
    lineas[3] = lineas[3].replace(",200000.00,400000.00,100000.00,", ",200000.00,400000.00,500000.00,")
    roto = tmp_path / "dos_filas_malas.csv"
    roto.write_text("\n".join(lineas), encoding="utf-8")

    upload = importar_inversion(roto)

    assert upload.estado == DatasetUpload.Estado.ERROR
    assert "Fila 2" in upload.log["error"]
    assert "Fila 4" in upload.log["error"]
