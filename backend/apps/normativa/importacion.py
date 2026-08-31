"""Importación de normativa desde un Excel de 7 columnas.

Segundo caso de **ADR-D9**, y por lo mismo que el primero: la carga es **aditiva** (no borra lo
que ya está), **parcial por diseño** (importa lo válido y omite el resto diciendo por qué) y
**síncrona**, porque la pantalla de confirmación tiene que responder en el momento. `DatasetUpload`
es la otra vía y no sirve aquí: es reemplazo total y va por el worker.

**En qué se aparta del importador de fichas ACC.** Allí las 17 columnas *son* los 17 `TextField`
del modelo, así que basta con recortar el texto y la cabecera se deriva de los `verbose_name`. Aquí
la hoja la trae el cliente con sus propios nombres y hay que **deducir** tres campos que el modelo
exige y la hoja no trae:

- **`tipo`** es una lista cerrada de cinco. La columna es texto libre, así que hay tabla de
  sinónimos; lo que no case **omite la fila**. Replegar a una opción por defecto dejaría la norma
  clasificada como algo que nadie decidió, y con aspecto de dato bueno.
- **`entidad_emisora`** casa contra el catálogo de ADR-D11, por nombre o por sigla. **Nunca se
  crea una entidad**: ese ADR ya decidió que el catálogo lo mantiene una persona, porque escritas
  a mano «PCM», «P.C.M.» y «Presidencia del Consejo de Ministros» son tres filas y el filtro del
  listado deja de servir. Un importador tiene todavía menos derecho que la IA a inventarlas.
- **`ambito`** no está en la hoja y es obligatorio. Se deduce del **nombre canónico de la entidad
  ya casada**, no del texto que escribió el usuario —que puede venir como sigla, y «MPC» no dice
  que sea una municipalidad—. Lo que no case, fuera: inventar el nivel de gobierno de una norma es
  peor que no importarla.

Y **`fecha`** es un `DateField` obligatorio contra una columna de año: `2019 → 2019-01-01`, que es
lo que ya decidió ADR-D10 para el mismo problema. Un año ilegible omite la fila y **no cae a hoy**,
que sería la única fecha del sistema que parece cierta sin serlo.

Aquí vive solo la lógica —leer, clasificar, escribir—; la vista está en `admin.py` sobre
`core.importacion_admin`. Separarlas es lo que permite probar la clasificación sin pasar por HTTP.
"""
from __future__ import annotations

import datetime as dt
import io
import unicodedata
from dataclasses import dataclass, field

from django.db import transaction
from django.utils.text import slugify

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .models import EntidadEmisora, Norma

#: La cabecera exacta, en orden. Constante literal y no `verbose_name` derivados —al revés que en
#: fichas ACC—: esta hoja la define el cliente y no tiene por qué parecerse a los campos.
COLUMNAS = (
    "N",
    "Tipo de normativa",
    "Nombre",
    "Descripción",
    "Entidad autora",
    "Año de publicación",
    "Link",
)

#: La columna que identifica la norma y sobre la que se comprueba que no haya repetidos.
COLUMNA_NOMBRE = "Nombre"

#: Sin estas no hay norma que crear. «N» es un correlativo del cliente y no se usa; «Link» puede
#: faltar, porque `url_oficial` es opcional en el modelo.
COLUMNAS_OBLIGATORIAS = ("Tipo de normativa", "Nombre", "Descripción", "Entidad autora",
                         "Año de publicación")

HOJA_DATOS = "Normativa"
HOJA_INSTRUCCIONES = "Instrucciones"

#: Cómo se escribe cada tipo por ahí fuera. La clave va normalizada (sin tildes, en mayúsculas y
#: con los espacios colapsados) por `_clave`.
SINONIMOS_TIPO = {
    "LEY": Norma.Tipo.LEY,
    "DECRETO SUPREMO": Norma.Tipo.DS,
    "DS": Norma.Tipo.DS,
    "D.S.": Norma.Tipo.DS,
    "RESOLUCION MINISTERIAL": Norma.Tipo.RM,
    "RM": Norma.Tipo.RM,
    "R.M.": Norma.Tipo.RM,
    "RESOLUCION JEFATURAL": Norma.Tipo.RJ,
    "RJ": Norma.Tipo.RJ,
    "R.J.": Norma.Tipo.RJ,
    "ORDENANZA": Norma.Tipo.ORDENANZA,
    "ORDENANZA REGIONAL": Norma.Tipo.ORDENANZA,
    "ORDENANZA MUNICIPAL": Norma.Tipo.ORDENANZA,
    "ORDENANZA PROVINCIAL": Norma.Tipo.ORDENANZA,
    "ORDENANZA DISTRITAL": Norma.Tipo.ORDENANZA,
}

#: Qué prefijo del nombre canónico de la entidad la sitúa en cada nivel de gobierno. Se recorre en
#: orden y gana el primero que aparezca dentro del nombre.
REGLAS_AMBITO = (
    ("MUNICIPALIDAD", Norma.Ambito.LOCAL),
    ("GOBIERNO REGIONAL", Norma.Ambito.REGIONAL),
    ("GOBIERNO LOCAL", Norma.Ambito.LOCAL),
    ("CONGRESO", Norma.Ambito.NACIONAL),
    ("PRESIDENCIA DEL CONSEJO", Norma.Ambito.NACIONAL),
    ("MINISTERIO", Norma.Ambito.NACIONAL),
    ("INSTITUTO NACIONAL", Norma.Ambito.NACIONAL),
    ("SERVICIO NACIONAL", Norma.Ambito.NACIONAL),
    ("AUTORIDAD NACIONAL", Norma.Ambito.NACIONAL),
    ("CENTRO NACIONAL", Norma.Ambito.NACIONAL),
    ("SUPERINTENDENCIA", Norma.Ambito.NACIONAL),
)

#: Fuera de esto, un año no es un año: es una celda mal llenada o un número de otra cosa.
ANIO_MINIMO, ANIO_MAXIMO = 1900, 2100

_TOPE_RESUMEN = Norma._meta.get_field("resumen").max_length
_TOPE_TITULO = Norma._meta.get_field("titulo").max_length
_TOPE_SLUG = Norma._meta.get_field("slug").max_length


def clave_comparacion(texto: str) -> str:
    """Clave para decidir si dos nombres de norma son el mismo.

    Recorta y sube a mayúsculas. Es **solo para comparar**: lo que se guarda es el texto tal como
    vino en el Excel.
    """
    return (texto or "").strip().upper()


def _clave(texto: str) -> str:
    """Como la anterior, más colapsar espacios y quitar tildes.

    Se usa para la cabecera, los tipos y las entidades: son cosas que una persona reescribe a
    mano, y una tilde perdida no puede costar la importación. **Los datos que se guardan no se
    tocan.**
    """
    plano = unicodedata.normalize("NFKD", (texto or "").strip().upper())
    plano = "".join(c for c in plano if not unicodedata.combining(c))
    return " ".join(plano.split())


def _texto(celda) -> str:
    if celda is None:
        return ""
    if isinstance(celda, str):
        return celda.strip()
    if isinstance(celda, float) and celda.is_integer():
        # openpyxl devuelve 2011.0 para un entero escrito en una celda con formato numérico, y
        # «2011.0» no casa con ningún año.
        return str(int(celda))
    return str(celda).strip()


def _extracto(texto: str, limite: int = 80) -> str:
    texto = (texto or "").strip()
    return texto if len(texto) <= limite else texto[: limite - 1] + "…"


class ExcelInvalido(Exception):
    """El archivo no se puede leer o su cabecera no es la esperada: no se importa nada."""


@dataclass
class FilaValida:
    numero: int
    #: Ya convertidos y listos para construir la `Norma`.
    valores: dict
    #: Lo que se hizo con el dato sin llegar a rechazarlo (p. ej. recortar el resumen).
    avisos: list[str] = field(default_factory=list)

    @property
    def nombre(self) -> str:
        return self.valores["titulo"]


@dataclass
class FilaOmitida:
    numero: int
    nombre: str
    motivo: str


@dataclass
class Analisis:
    validas: list[FilaValida] = field(default_factory=list)
    omitidas: list[FilaOmitida] = field(default_factory=list)

    @property
    def total(self) -> int:
        return len(self.validas) + len(self.omitidas)

    @property
    def entidades_desconocidas(self) -> list[str]:
        """Las que hay que dar de alta en el catálogo para que esas filas entren.

        Se agrupan para poder crearlas de una vez y volver a subir el archivo, en vez de
        descubrirlas de una en una.
        """
        vistas: list[str] = []
        for fila in self.omitidas:
            marca = "no está en el catálogo de entidades emisoras: "
            if marca in fila.motivo:
                nombre = fila.motivo.split(marca, 1)[1].strip(" «».")
                if nombre and nombre not in vistas:
                    vistas.append(nombre)
        return vistas


def _catalogo_entidades() -> dict[str, EntidadEmisora]:
    """El catálogo indexado por nombre **y** por sigla, ambos normalizados."""
    catalogo: dict[str, EntidadEmisora] = {}
    for entidad in EntidadEmisora.objects.all():
        catalogo[_clave(entidad.nombre)] = entidad
        if entidad.sigla:
            catalogo.setdefault(_clave(entidad.sigla), entidad)
    return catalogo


def ambito_de(entidad: EntidadEmisora) -> str | None:
    """El nivel de gobierno de una entidad, o `None` si sus reglas no la cubren.

    Se mira el **nombre canónico** del catálogo y no lo que escribió el usuario: la hoja puede
    traer la sigla, y «MPC» no dice que sea una municipalidad.

    Devolver `None` y omitir la fila es deliberado. El catálogo lo amplía PREDES, así que tarde o
    temprano entrará una entidad que estas reglas no cubran; colar un «nacional» por defecto la
    clasificaría mal y ya no se distinguiría de una correcta.
    """
    nombre = _clave(entidad.nombre)
    for marca, ambito in REGLAS_AMBITO:
        if marca in nombre:
            return ambito
    return None


def _fecha_de(anio_texto: str) -> dt.date | None:
    """`"2019"` ⇒ 1 de enero de 2019. Lo que no sea un año creíble, `None`.

    El 1 de enero es la convención que ya usa el proyecto cuando solo hay año (ADR-D10). Lo que
    **no** se hace es replegar a hoy: sería la única fecha del sistema que parece cierta sin
    serlo, indistinguible de una cargada a mano.
    """
    try:
        anio = int(float(anio_texto))
    except (TypeError, ValueError):
        return None
    if not (ANIO_MINIMO <= anio <= ANIO_MAXIMO):
        return None
    return dt.date(anio, 1, 1)


def _slug_libre(titulo: str, reservados: set[str]) -> str:
    """Un slug único, reservándolo en memoria.

    `core.slug_unico()` consulta la base por cada candidato, así que no vería las colisiones
    **dentro del mismo archivo**: con `bulk_create` la transacción entera reventaría contra el
    índice único y no entraría ninguna norma.
    """
    base = slugify(titulo)[: _TOPE_SLUG - 10] or "norma"
    candidato, n = base, 1
    while candidato in reservados:
        n += 1
        candidato = f"{base}-{n}"
    reservados.add(candidato)
    return candidato


def analizar(origen) -> Analisis:
    """Lee el Excel y reparte sus filas entre las que se importarán y las que no.

    `origen` es cualquier cosa que openpyxl acepte: una ruta o un archivo abierto en binario.

    Una fila mala nunca aborta el archivo; una cabecera equivocada sí. Con las columnas corridas
    cada valor entraría en el campo de al lado —el tipo en el nombre, la entidad en la
    descripción— y la norma quedaría plausible y mal, sin que nada lo delatara después.
    """
    try:
        wb = openpyxl.load_workbook(origen, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl lanza de todo ante un archivo que no es .xlsx
        raise ExcelInvalido(f"No se pudo leer el archivo como Excel (.xlsx): {exc}") from exc

    try:
        hoja = wb.worksheets[0]

        cabecera = next(hoja.iter_rows(max_row=1, values_only=True), None)
        leidas = [_texto(c) for c in (cabecera or [])][: len(COLUMNAS)]
        if [_clave(c) for c in leidas] != [_clave(c) for c in COLUMNAS]:
            raise ExcelInvalido(
                f"La hoja «{hoja.title}» no tiene las {len(COLUMNAS)} columnas esperadas, así "
                "que no se importó nada.\n\n"
                "Se esperaba:\n"
                + "\n".join(f"  {i}. {c}" for i, c in enumerate(COLUMNAS, 1))
                + "\n\nSe encontró:\n"
                + ("\n".join(f"  {i}. {c}" for i, c in enumerate(leidas, 1)) or "  (vacío)")
                + "\n\nDescarga la plantilla desde el listado para partir de la cabecera correcta."
            )

        analisis = Analisis()
        catalogo = _catalogo_entidades()
        # Lo que ya está en la base y lo que va entrando del propio archivo: un nombre repetido
        # dentro del Excel es tan duplicado como uno que choca contra una norma existente.
        vistos: dict[str, str] = {
            clave_comparacion(titulo): f"ya hay una norma registrada como «{_extracto(titulo)}»"
            for titulo in Norma.objects.values_list("titulo", flat=True)
        }
        reservados: set[str] = set(Norma.objects.values_list("slug", flat=True))

        for numero, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            celdas = {
                columna: _texto(fila[i]) if i < len(fila) else ""
                for i, columna in enumerate(COLUMNAS)
            }
            if not any(celdas.values()):
                continue  # fila entera en blanco: el relleno del final del Excel, no un error

            nombre = celdas[COLUMNA_NOMBRE]

            def omitir(motivo: str) -> None:
                analisis.omitidas.append(FilaOmitida(numero, nombre, motivo))

            vacias = [c for c in COLUMNAS_OBLIGATORIAS if not celdas[c]]
            if vacias:
                omitir("faltan columnas obligatorias: " + ", ".join(f"«{c}»" for c in vacias) + ".")
                continue

            clave = clave_comparacion(nombre)
            if clave in vistos:
                omitir(f"el nombre está repetido: {vistos[clave]}.")
                continue

            tipo = SINONIMOS_TIPO.get(_clave(celdas["Tipo de normativa"]))
            if tipo is None:
                omitir(
                    f"el tipo «{celdas['Tipo de normativa']}» no está en el catálogo "
                    "(Ley, Decreto Supremo, Resolución Ministerial, Resolución Jefatural, "
                    "Ordenanza)."
                )
                continue

            entidad = catalogo.get(_clave(celdas["Entidad autora"]))
            if entidad is None:
                omitir(
                    "no está en el catálogo de entidades emisoras: "
                    f"«{celdas['Entidad autora']}»."
                )
                continue

            ambito = ambito_de(entidad)
            if ambito is None:
                omitir(
                    f"no se pudo deducir el ámbito de «{entidad.nombre}». Complétala a mano o "
                    "ajusta el nombre de la entidad en su catálogo."
                )
                continue

            fecha = _fecha_de(celdas["Año de publicación"])
            if fecha is None:
                omitir(
                    f"el año «{celdas['Año de publicación']}» no es un año entre "
                    f"{ANIO_MINIMO} y {ANIO_MAXIMO}."
                )
                continue

            avisos: list[str] = []
            resumen = celdas["Descripción"]
            if len(resumen) > _TOPE_RESUMEN:
                resumen = resumen[:_TOPE_RESUMEN]
                avisos.append(f"la descripción se recortó a {_TOPE_RESUMEN} caracteres")
            titulo = nombre[:_TOPE_TITULO]
            if len(nombre) > _TOPE_TITULO:
                avisos.append(f"el nombre se recortó a {_TOPE_TITULO} caracteres")

            vistos[clave] = f"lo trajo antes la fila {numero} de este mismo archivo"
            analisis.validas.append(
                FilaValida(
                    numero,
                    {
                        "titulo": titulo,
                        "slug": _slug_libre(titulo, reservados),
                        "tipo": tipo,
                        "ambito": ambito,
                        "entidad_emisora": entidad,
                        "fecha": fecha,
                        "resumen": resumen,
                        "url_oficial": celdas["Link"] or "",
                    },
                    avisos,
                )
            )

        return analisis
    finally:
        wb.close()


@transaction.atomic
def importar(analisis: Analisis) -> int:
    """Crea las normas válidas **en borrador** y devuelve cuántas entraron.

    En borrador y no publicadas: el importador trae datos, no decide qué se publica. El estado lo
    mueve una persona por `WorkflowMixin.transicionar()`, que es donde vive el permiso.
    """
    normas = [Norma(**fila.valores) for fila in analisis.validas]
    Norma.objects.bulk_create(normas, batch_size=500)
    return len(normas)


def plantilla_xlsx() -> bytes:
    """Excel vacío con la cabecera exacta y una hoja de instrucciones.

    Se genera desde `COLUMNAS`, así que la plantilla que se descarga y el validador que la lee no
    pueden separarse. Hay una prueba que la descarga, la llena y la reimporta.
    """
    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.title = HOJA_DATOS

    ayuda = {
        "N": "Correlativo del archivo. No se importa: está para que puedas referirte a una fila.",
        "Tipo de normativa": "Ley, Decreto Supremo, Resolución Ministerial, Resolución Jefatural "
                             "u Ordenanza. Se aceptan abreviaturas (DS, RM, RJ).",
        "Nombre": "Título de la norma. NO puede repetirse: es lo que decide si una fila ya está "
                  "cargada.",
        "Descripción": f"De qué trata, en un párrafo. Máximo {_TOPE_RESUMEN} caracteres.",
        "Entidad autora": "Institución que la dicta, tal como está en el catálogo de entidades "
                          "emisoras del admin. Vale el nombre completo o la sigla.",
        "Año de publicación": f"Solo el año, entre {ANIO_MINIMO} y {ANIO_MAXIMO}.",
        "Link": "Enlace a la publicación oficial. Es la única columna que puede ir vacía.",
    }

    hoja.append(list(COLUMNAS))
    relleno = PatternFill("solid", fgColor="FFE8F2EC")
    for i, columna in enumerate(COLUMNAS, 1):
        celda = hoja.cell(row=1, column=i)
        celda.font = Font(bold=True)
        celda.fill = relleno
        celda.alignment = Alignment(wrap_text=True, vertical="top")
        hoja.column_dimensions[celda.column_letter].width = 14 if columna == "N" else 38
    hoja.row_dimensions[1].height = 34
    hoja.freeze_panes = "A2"

    instrucciones = wb.create_sheet(HOJA_INSTRUCCIONES)
    instrucciones.append(["#", "Columna", "¿Es obligatoria?", "Cómo llenarla"])
    for i, columna in enumerate(COLUMNAS, 1):
        obligatoria = "Obligatoria" if columna in COLUMNAS_OBLIGATORIAS else "Opcional"
        instrucciones.append([i, columna, obligatoria, ayuda[columna]])
    instrucciones.append([])
    instrucciones.append([
        "", "Nombres repetidos", "",
        "El «Nombre» debe ser distinto en cada norma, y distinto de las que ya están cargadas. Al "
        "comparar no se distinguen mayúsculas ni espacios sobrantes: las filas que repitan un "
        "nombre no se cargan, y antes de confirmar verás cuáles son.",
    ])
    instrucciones.append([
        "", "Entidades emisoras", "",
        "El importador no da de alta entidades: si una no está en el catálogo, su fila se omite y "
        "te dice cuál falta. Créala en «Entidades emisoras» y vuelve a subir el archivo.",
    ])
    instrucciones.append([
        "", "Ámbito y fecha", "",
        "No se piden en el archivo: el ámbito (nacional, regional o local) se deduce de la "
        "entidad, y la fecha se fija al 1 de enero del año indicado. Los verás en la pantalla de "
        "confirmación antes de que se guarde nada.",
    ])
    for columna, ancho in zip("ABCD", (5, 26, 16, 90)):
        instrucciones.column_dimensions[columna].width = ancho
    for fila in instrucciones.iter_rows(min_row=2):
        fila[3].alignment = Alignment(wrap_text=True, vertical="top")
    for celda in instrucciones[1]:
        celda.font = Font(bold=True)
        celda.fill = relleno

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
