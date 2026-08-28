"""Importación de fichas ACC desde un Excel de 17 columnas.

**Por qué no pasa por `DatasetUpload`.** Esa vía existe para los datasets de reemplazo total
—peligros, frecuencia, inversión—: subida, acción masiva, worker, «todo o nada por dataset».
Las fichas ACC son lo contrario en las tres dimensiones: la carga es **aditiva** (no borra las
fichas que ya están), **parcial por diseño** (importa lo válido y omite lo demás diciendo por
qué) y **síncrona**, porque la pantalla de confirmación tiene que responder en el momento.

Aquí vive solo la lógica —leer, clasificar, escribir—; las vistas están en `admin.py`. Separarlas
es lo que permite probar la clasificación sin pasar por HTTP.
"""
from __future__ import annotations

import io
import unicodedata
from dataclasses import dataclass, field

from django.db import transaction

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from .models import MedidaFichaACC, extracto

#: Los 17 campos de la ficha, en el orden en que se declaran en el modelo.
CAMPOS = [f"value_{n:03d}" for n in range(1, 18)]

#: El campo que identifica la ficha y sobre el que se comprueba que no haya repetidos.
CAMPO_NOMBRE = "value_001"

HOJA_DATOS = "Fichas ACC"
HOJA_INSTRUCCIONES = "Instrucciones"


def _campo(nombre: str):
    return MedidaFichaACC._meta.get_field(nombre)


def columnas_esperadas() -> list[str]:
    """Las 17 cabeceras, tomadas del modelo.

    No se copian a una constante de texto a propósito: si alguien retoca un `verbose_name`, la
    plantilla que se descarga y el validador que la lee siguen de acuerdo por construcción.
    """
    return [str(_campo(nombre).verbose_name) for nombre in CAMPOS]


def campos_obligatorios() -> list[str]:
    """Los que el modelo no deja en blanco. Hoy son 14: quedan fuera 002, 004 y 008."""
    return [nombre for nombre in CAMPOS if not _campo(nombre).blank]


def clave_comparacion(texto: str) -> str:
    """Clave para decidir si dos nombres de experiencia son el mismo.

    Recorta y sube a mayúsculas. Es **solo para comparar**: lo que se guarda es el texto tal
    como vino en el Excel.
    """
    return (texto or "").strip().upper()


def _clave_cabecera(texto: str) -> str:
    """Igual que la anterior, pero además colapsa espacios y quita tildes.

    Las cabeceras son preguntas largas que el usuario copia y pega; un doble espacio o una tilde
    perdida al reescribirlas no puede costarle la importación entera. Los datos no se tocan.
    """
    plano = unicodedata.normalize("NFKD", (texto or "").strip().upper())
    plano = "".join(c for c in plano if not unicodedata.combining(c))
    return " ".join(plano.split())


def _texto(celda) -> str:
    if celda is None:
        return ""
    if isinstance(celda, str):
        return celda.strip()
    return str(celda).strip()


class ExcelInvalido(Exception):
    """El archivo no se puede leer o su cabecera no es la esperada: no se importa nada."""


@dataclass
class FilaValida:
    numero: int
    valores: dict[str, str]

    @property
    def nombre(self) -> str:
        return self.valores[CAMPO_NOMBRE]


@dataclass
class FilaOmitida:
    numero: int
    nombre: str
    motivo: str


@dataclass
class Analisis:
    validas: list[FilaValida] = field(default_factory=list)
    omitidas: list[FilaOmitida] = field(default_factory=list)

    @property
    def total(self) -> int:
        return len(self.validas) + len(self.omitidas)


def analizar(origen) -> Analisis:
    """Lee el Excel y reparte sus filas entre las que se importarán y las que no.

    `origen` es cualquier cosa que openpyxl acepte: una ruta o un archivo abierto en binario.

    Una fila mala nunca aborta el archivo —ese es el patrón de los importadores de datasets—,
    pero una cabecera equivocada sí: si las columnas no son las que se creen, cada fila se
    guardaría con los textos cambiados de campo y nada lo delataría después.
    """
    try:
        wb = openpyxl.load_workbook(origen, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl lanza de todo ante un archivo que no es .xlsx
        raise ExcelInvalido(
            f"No se pudo leer el archivo como Excel (.xlsx): {exc}"
        ) from exc

    try:
        hoja = wb.worksheets[0]
        esperadas = columnas_esperadas()

        cabecera = next(hoja.iter_rows(max_row=1, values_only=True), None)
        leidas = [_texto(c) for c in (cabecera or [])][: len(esperadas)]
        if [_clave_cabecera(c) for c in leidas] != [_clave_cabecera(c) for c in esperadas]:
            raise ExcelInvalido(
                f"La hoja «{hoja.title}» no tiene las 17 columnas esperadas, así que no se "
                "importó nada.\n\n"
                "Se esperaba:\n"
                + "\n".join(f"  {i}. {c}" for i, c in enumerate(esperadas, 1))
                + "\n\nSe encontró:\n"
                + ("\n".join(f"  {i}. {c}" for i, c in enumerate(leidas, 1)) or "  (vacío)")
                + "\n\nDescarga la plantilla desde el listado para partir de la cabecera correcta."
            )

        analisis = Analisis()
        obligatorios = campos_obligatorios()
        # Lo que ya está en la base, y lo que va entrando del propio archivo: un nombre repetido
        # dentro del Excel es tan duplicado como uno que choca contra una ficha existente.
        vistos: dict[str, str] = {
            clave_comparacion(nombre): f"ya hay una ficha registrada como «{extracto(nombre)}»"
            for nombre in MedidaFichaACC.objects.values_list(CAMPO_NOMBRE, flat=True)
        }

        for numero, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            valores = {
                campo: _texto(fila[i]) if i < len(fila) else ""
                for i, campo in enumerate(CAMPOS)
            }
            if not any(valores.values()):
                continue  # fila entera en blanco: es el relleno del final del Excel, no un error

            nombre = valores[CAMPO_NOMBRE]
            vacios = [c for c in obligatorios if not valores[c]]
            if vacios:
                etiquetas = ", ".join(f"«{_campo(c).verbose_name}»" for c in vacios)
                analisis.omitidas.append(
                    FilaOmitida(numero, nombre, f"faltan campos obligatorios: {etiquetas}.")
                )
                continue

            clave = clave_comparacion(nombre)
            if clave in vistos:
                analisis.omitidas.append(
                    FilaOmitida(
                        numero,
                        nombre,
                        f"el nombre de la experiencia está repetido: {vistos[clave]}.",
                    )
                )
                continue

            vistos[clave] = f"lo trajo antes la fila {numero} de este mismo archivo"
            analisis.validas.append(FilaValida(numero, valores))

        return analisis
    finally:
        wb.close()


@transaction.atomic
def importar(analisis: Analisis) -> int:
    """Crea las fichas válidas y devuelve cuántas entraron."""
    fichas = [MedidaFichaACC(**fila.valores) for fila in analisis.validas]
    MedidaFichaACC.objects.bulk_create(fichas, batch_size=500)
    return len(fichas)


def plantilla_xlsx() -> bytes:
    """Excel vacío con la cabecera exacta y las ayudas del formulario.

    Los `help_text` van como comentario de celda **y** en una segunda hoja: el comentario se ve
    al llenar, la hoja se ve al imprimir. El importador lee siempre la primera hoja, así que la
    de instrucciones no estorba.
    """
    wb = openpyxl.Workbook()

    datos = wb.active
    datos.title = HOJA_DATOS
    relleno = PatternFill("solid", start_color="FFE8F2EC")
    for i, campo in enumerate(CAMPOS, start=1):
        meta = _campo(campo)
        celda = datos.cell(row=1, column=i, value=str(meta.verbose_name))
        celda.font = Font(bold=True)
        celda.fill = relleno
        celda.alignment = Alignment(wrap_text=True, vertical="top")
        ayuda = str(meta.help_text or "")
        ayuda = f"{ayuda}\n\n(Campo opcional)".strip() if meta.blank else ayuda
        if ayuda:
            celda.comment = Comment(ayuda, "Observatorio Kallpachakuy")
        datos.column_dimensions[celda.column_letter].width = 38
    datos.row_dimensions[1].height = 60
    datos.freeze_panes = "A2"

    guia = wb.create_sheet(HOJA_INSTRUCCIONES)
    guia.append(["#", "Columna", "¿Es obligatoria?", "Cómo llenarla"])
    for celda in guia[1]:
        celda.font = Font(bold=True)
        celda.fill = relleno
    for i, campo in enumerate(CAMPOS, start=1):
        meta = _campo(campo)
        guia.append([
            i,
            str(meta.verbose_name),
            "Opcional" if meta.blank else "Obligatoria",
            str(meta.help_text) or "—",
        ])
    guia.append([])
    guia.append([
        "",
        "Nombres repetidos",
        "",
        "El «Nombre de la experiencia, práctica proyecto o programa» debe ser distinto en cada "
        "ficha. Al importar se compara sin distinguir mayúsculas ni espacios sobrantes: las "
        "filas que repitan un nombre no se cargan, y antes de confirmar verás cuáles son.",
    ])
    for columna, ancho in (("A", 5), ("B", 42), ("C", 16), ("D", 90)):
        guia.column_dimensions[columna].width = ancho
    for fila in guia.iter_rows(min_row=2):
        for celda in fila:
            celda.alignment = Alignment(wrap_text=True, vertical="top")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
