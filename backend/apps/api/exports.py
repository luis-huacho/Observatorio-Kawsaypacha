"""Exports a Excel (requisito 3 del TDR).

Se usa `write_only=True` de openpyxl: escribe fila a fila sin mantener el libro en memoria, que
es lo que hace viable exportar los 8,968 centros poblados sin que el worker se hinche.
"""
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Verde de la paleta PREDES, para que el Excel se reconozca como del observatorio.
RELLENO_CABECERA = PatternFill("solid", fgColor="0B3B26")
FUENTE_CABECERA = Font(color="FFFFFF", bold=True)


def libro_excel(hoja: str, cabeceras: list[str], filas, anchos: list[int] | None = None):
    """Devuelve un `HttpResponse` con el .xlsx listo para descargar."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(hoja[:31])  # Excel corta los títulos de hoja en 31 caracteres

    if anchos:
        from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

        holder = DimensionHolder(worksheet=ws)
        for i, ancho in enumerate(anchos, start=1):
            holder[get_column_letter(i)] = ColumnDimension(
                ws, index=get_column_letter(i), width=ancho
            )
        ws.column_dimensions = holder

    from openpyxl.cell import WriteOnlyCell

    celdas = []
    for texto in cabeceras:
        celda = WriteOnlyCell(ws, value=texto)
        celda.fill = RELLENO_CABECERA
        celda.font = FUENTE_CABECERA
        celda.alignment = Alignment(vertical="center", wrap_text=True)
        celdas.append(celda)
    ws.append(celdas)

    for fila in filas:
        ws.append(list(fila))

    bufer = BytesIO()
    wb.save(bufer)
    bufer.seek(0)
    return bufer


def respuesta_excel(nombre_archivo: str, hoja: str, cabeceras, filas, anchos=None):
    bufer = libro_excel(hoja, list(cabeceras), filas, anchos)
    respuesta = HttpResponse(
        bufer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    respuesta["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return respuesta


NIVELES = {1: "Bajo", 2: "Medio", 3: "Alto", 4: "Muy alto"}


def filas_ccpp(queryset, tipos):
    """Una fila por centro poblado, con su nivel máximo y **todos sus peligros**.

    Se exporta la misma unidad que muestra la tabla —centros poblados, no clasificaciones—
    para que el Excel cuadre con lo que el usuario tenía en pantalla. Por eso los peligros no
    abren fila: van en una columna de texto legible y, además, uno por columna con su nivel,
    que es lo que permite filtrar y pivotar en Excel.

    Los peligros salen de `clasificaciones_filtradas`, el `to_attr` que arma el viewset con los
    mismos `peligros`/`niveles` de la petición y en el mismo orden con el que el visor elige el
    ícono: así el Excel de una consulta filtrada no puede hablar de lo que el mapa oculta. Se
    lee **sin repliegue** a propósito — un `getattr` con defecto convertiría la pérdida del
    prefetch en una consulta por fila, silenciosa hasta producción.
    """
    campos = queryset.select_related("distrito__provincia").iterator(chunk_size=2000)
    for c in campos:
        nivel = getattr(c, "nivel", None)
        peligros = {
            clasificacion.tipo_peligro_id: clasificacion.nivel
            for clasificacion in c.clasificaciones_filtradas
        }
        yield [
            c.codigo,
            c.nombre,
            c.categoria,
            c.distrito.provincia.nombre,
            c.distrito.nombre,
            c.distrito_id,
            nivel or "",
            NIVELES.get(nivel, "Sin dato clasificado"),
            "; ".join(
                f"{clasificacion.tipo_peligro.nombre} "
                f"({clasificacion.nivel} · {NIVELES[clasificacion.nivel]})"
                for clasificacion in c.clasificaciones_filtradas
            ),
            # Enteros, no texto: es lo que deja ordenar y filtrar la columna por nivel, que es
            # justo el motivo de que exista. Vacío = ese peligro no está clasificado aquí, o no
            # pasa los filtros.
            *(peligros.get(tipo.pk, "") for tipo in tipos),
        ]


#: Columnas fijas. Las de peligro se añaden después, **desde el catálogo**: un décimo peligro
#: aparece en el Excel sin tocar código, igual que aparece en el visor.
CABECERAS_CCPP = [
    "Código INEI", "Centro poblado", "Categoría", "Provincia", "Distrito", "Ubigeo distrito",
    "Nivel", "Nivel (descripción)", "Peligros",
]
ANCHOS_CCPP = [13, 30, 14, 18, 18, 14, 7, 20, 60]


def cabeceras_ccpp(tipos) -> list[str]:
    return CABECERAS_CCPP + [f"{tipo.nombre} (nivel)" for tipo in tipos]


def anchos_ccpp(tipos) -> list[int]:
    return ANCHOS_CCPP + [max(10, len(tipo.nombre) + 3) for tipo in tipos]


CABECERAS_INVERSION = [
    "Ejercicio", "Corte", "Fuente", "Código MEF", "Entidad", "Ámbito", "Provincia", "Distrito",
    "Ubigeo distrito", "PIA (0068)", "PIM (0068)", "Devengado (0068)", "% ejecución",
    "Saldo por ejecutar", "Variación PIA-PIM", "PIA institucional", "PIM institucional",
    "Devengado institucional", "% del 0068 sobre el total",
    "PIM proyectos", "PIM actividades", "% en proyectos",
]
ANCHOS_INVERSION = [10, 10, 26, 12, 42, 14, 18, 18, 14, 15, 15, 15, 12, 17, 17, 17, 17, 21, 20,
                    15, 15, 13]

#: Columnas que se añaden al comparar con otro ejercicio.
CABECERAS_COMPARACION = [
    "PIM comparado", "Devengado comparado", "% ejecución comparado",
    "Δ PIM", "Δ PIM (%)", "Δ devengado", "Δ % ejecución", "Comparabilidad",
]
ANCHOS_COMPARACION = [16, 20, 21, 15, 13, 16, 16, 46]

#: Texto de la columna «Comparabilidad». Es la mitigación de una decisión consciente: el Δ de
#: % de ejecución se muestra aunque uno de los dos ejercicios sea un corte parcial, y en
#: pantalla la leyenda está al lado — pero el Excel viaja solo por correo, así que la
#: advertencia tiene que ir **en la propia fila**.
AVISO_NO_COMPARABLE = (
    "Cortes distintos: el Δ de % de ejecución no es comparable (uno de los dos es parcial)"
)
AVISO_COMPARABLE = "Ejercicios del mismo tipo de corte"


def cabeceras_inversion(ejercicio_comparado=None) -> list[str]:
    if ejercicio_comparado is None:
        return list(CABECERAS_INVERSION)
    anio = ejercicio_comparado.anio
    return CABECERAS_INVERSION + [
        f"{c} ({anio})" if "comparado" in c else c for c in CABECERAS_COMPARACION
    ]


def anchos_inversion(ejercicio_comparado=None) -> list[int]:
    return ANCHOS_INVERSION + (ANCHOS_COMPARACION if ejercicio_comparado is not None else [])


def filas_inversion(filas, ejercicio, ejercicio_comparado=None):
    """Una fila por entidad, con los derivados ya calculados por `inversion.consultas`.

    Se exportan los mismos números que la pantalla —no los recalcula el export— y cada fila
    repite ejercicio, corte y fuente: el archivo viaja suelto por correo, y sin el corte nadie
    puede saber que un 47 % de ejecución es de medio año.
    """
    for f in filas:
        fila = [
            ejercicio.anio,
            ejercicio.corte,
            ejercicio.get_fuente_display(),
            f["codigo"],
            f["entidad"],
            f["ambito"],
            f["provincia"] or "",
            f["distrito"] or "",
            f["ubigeo_distrito"] or "",
            f["pia"],
            f["pim"],
            f["devengado"],
            f["pct_ejecucion"],
            f["saldo"],
            f["variacion_pia_pim"],
            f["pia_institucional"],
            f["pim_institucional"],
            f["devengado_institucional"],
            f["pct_0068_institucional"],
            f["pim_proyectos"],
            f["pim_actividades"],
            f["pct_proyectos"],
        ]
        if ejercicio_comparado is not None:
            c = f.get("comparacion") or {}
            fila += [
                c.get("pim"),
                c.get("devengado"),
                c.get("pct_ejecucion"),
                c.get("delta_pim"),
                c.get("pct_delta_pim"),
                c.get("delta_devengado"),
                c.get("delta_pct_ejecucion"),
                AVISO_COMPARABLE if c.get("comparable") else AVISO_NO_COMPARABLE,
            ]
        yield fila
