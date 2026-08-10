"""Importa el presupuesto del PP 0068 por entidad ejecutora.

Acepta **tres formas**, y las distingue por su propia cabecera en vez de por un tipo de carga
distinto, porque las tres responden a la misma pregunta y reemplazan trozos del mismo dato:

1. **Excel del cliente** (`.xlsx`, hoja `Base AAAA`) — un solo ejercicio con su corte, y dentro
   las filas del programa y las de `PRESUPUESTO INSTITUCIONAL`. Es el camino recurrente: cada
   vez que PREDES reciba una base nueva, sube ese archivo y nada más.
2. **Serie consolidada del programa** (`.csv` con `PRODUCTO_PROYECTO`) — la salida de
   `scripts/consolidar_pp0068.py`, varios ejercicios de golpe. Es la carga histórica.
3. **Serie de totales institucionales** (`.csv` sin `PRODUCTO_PROYECTO`) — la salida de
   `scripts/totales_institucionales.py`, el denominador de la 2.

El reemplazo es atómico **por ejercicio y por parte**: subir la serie del programa no borra los
totales institucionales ya cargados, ni al revés. Si lo hiciera, cargar los tres archivos en
cualquier orden dejaría siempre alguno a medias.

La entidad se resuelve por su código del MEF, que es estable. La geografía no lo es —hay 15
entidades cuya provincia o distrito cambia de grafía entre ejercicios (QUISPICANCHI /
QUISPICANCHIS, KIMBIRI / QUIMBIRI)—, así que se prueban todas las grafías vistas, de la más
reciente a la más antigua, contra `Distrito.nombre_normalizado`.
"""
import csv
import io
import re
from decimal import Decimal, InvalidOperation

from django.db import transaction

import openpyxl

from apps.inversion.catalogo import (
    ACTIVIDAD_A_PROCESO,
    PROCESO_POR_DEFECTO_PROYECTOS,
    codigo_clasificable,
    es_proyecto,
)

CERO = Decimal("0")
PERIODO = re.compile(r"^(\d{4})-(\d{2})$")
CATEGORIA_INSTITUCIONAL = "PRESUPUESTO INSTITUCIONAL"

#: Columnas mínimas de cada forma de CSV.
COLUMNAS_PROGRAMA = {"EJERCICIO", "ENTIDAD_CODIGO", "PRODUCTO_PROYECTO", "PIA", "PIM", "DEVENGADO"}
COLUMNAS_INSTITUCIONAL = {"EJERCICIO", "ENTIDAD_CODIGO", "PIA", "PIM", "DEVENGADO"}


def _texto(valor) -> str:
    return str(valor).strip() if valor is not None else ""


def _importe(valor, contexto: str) -> Decimal:
    """Importe del archivo. Acepta el `600,00` con coma decimal que trae el Excel del cliente."""
    if valor is None or valor == "":
        return CERO
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    texto = str(valor).strip()
    try:
        return Decimal(texto.replace(",", ".") if re.fullmatch(r"-?\d+,\d+", texto) else texto)
    except InvalidOperation:
        raise ValueError(f"{contexto}: no se puede interpretar el importe {valor!r}.")


def _ambito(nivel_gobierno: str, nombre: str):
    from apps.inversion.models import EntidadEjecutora

    if nivel_gobierno == "M":
        # Las 13 municipalidades provinciales de Cusco se nombran así en el MEF, sin excepción,
        # y las 4 mancomunidades también se declaran en el nombre.
        mayusculas = nombre.upper()
        if "MANCOMUNIDAD" in mayusculas:
            return EntidadEjecutora.Ambito.MANCOMUNIDAD
        return (
            EntidadEjecutora.Ambito.PROVINCIAL
            if "PROVINCIAL" in mayusculas
            else EntidadEjecutora.Ambito.DISTRITAL
        )
    return (
        EntidadEjecutora.Ambito.REGIONAL
        if nivel_gobierno == "R"
        else EntidadEjecutora.Ambito.NACIONAL
    )


class _Resolutor:
    """Casa nombres de provincia y distrito con el padrón, probando varias grafías."""

    def __init__(self):
        from apps.territorio.models import Distrito, Provincia

        self.distritos: dict[str, list] = {}
        for distrito in Distrito.objects.select_related("provincia"):
            self.distritos.setdefault(distrito.nombre_normalizado, []).append(distrito)
        self.provincias = {
            self._normalizar(p.nombre): p for p in Provincia.objects.all()
        }

    @staticmethod
    def _normalizar(nombre: str) -> str:
        from apps.territorio.utils import normalizar_nombre

        return normalizar_nombre(nombre)

    def distrito(self, nombre_provincia: str, nombre_distrito: str):
        candidatos = self.distritos.get(self._normalizar(nombre_distrito), [])
        if not candidatos:
            return None
        if len(candidatos) == 1:
            return candidatos[0]
        provincia = self._normalizar(nombre_provincia)
        acotados = [d for d in candidatos if self._normalizar(d.provincia.nombre) == provincia]
        return acotados[0] if len(acotados) == 1 else None

    def provincia(self, nombre: str):
        return self.provincias.get(self._normalizar(nombre))


# --- Lectura de cada forma -------------------------------------------------


def _leer_csv(upload):
    with upload.archivo.open("rb") as fh:
        texto = fh.read().decode("utf-8-sig")
    lector = csv.DictReader(io.StringIO(texto))
    columnas = set(lector.fieldnames or [])
    if COLUMNAS_PROGRAMA <= columnas:
        forma = "programa"
    elif COLUMNAS_INSTITUCIONAL <= columnas:
        forma = "institucional"
    else:
        faltan = sorted(COLUMNAS_INSTITUCIONAL - columnas)
        raise ValueError(
            "El CSV no se parece a ninguna de las dos series esperadas. Le faltan columnas: "
            + ", ".join(faltan)
            + f". Columnas encontradas: {', '.join(sorted(columnas))}."
        )
    return forma, list(lector)


def _leer_excel(upload):
    """Convierte el Excel del cliente a las mismas filas que producen los CSV."""
    libro = openpyxl.load_workbook(upload.archivo.path, read_only=True, data_only=True)
    hojas = [h for h in libro.sheetnames if h.lower().startswith("base")]
    if not hojas:
        raise ValueError(
            "El Excel no tiene ninguna hoja que empiece por 'Base'. "
            f"Hojas del archivo: {', '.join(libro.sheetnames)}."
        )
    hoja = libro[hojas[0]]

    filas = hoja.iter_rows(values_only=True)
    cabecera = [_texto(c) for c in next(filas)]
    indice = {nombre: i for i, nombre in enumerate(cabecera)}
    obligatorias = ["Periodo", "Pliego", "Categoría_Presupuestal", "PIA", "PIM", "Devengado"]
    if faltan := [c for c in obligatorias if c not in indice]:
        raise ValueError(f"A la hoja '{hojas[0]}' le faltan columnas: {', '.join(faltan)}.")

    def celda(fila, nombre):
        i = indice.get(nombre)
        return fila[i] if i is not None and i < len(fila) else None

    programa, institucional, periodos = [], [], set()
    for fila in filas:
        periodo = _texto(celda(fila, "Periodo"))
        if not periodo:
            continue
        periodos.add(periodo)
        categoria = _texto(celda(fila, "Categoría_Presupuestal"))
        entidad_tipo, codigo, nombre = _entidad_de_pliego(_texto(celda(fila, "Pliego")))
        comun = {
            "ENTIDAD_TIPO": entidad_tipo,
            "ENTIDAD_CODIGO": codigo,
            "ENTIDAD_NOMBRE": nombre,
            # El Excel es solo de gobiernos locales y del regional; el tipo de llave lo dice.
            "NIVEL_GOBIERNO": "M" if entidad_tipo == "SEC_EJEC" else "R",
            "PROVINCIA": _texto(celda(fila, "Provincia")),
            "DISTRITO": _texto(celda(fila, "Distrito")),
            "PIA": celda(fila, "PIA"),
            "PIM": celda(fila, "PIM"),
            "DEVENGADO": celda(fila, "Devengado"),
        }
        if categoria == CATEGORIA_INSTITUCIONAL:
            institucional.append(comun)
            continue
        producto, nombre_producto = _codigo_y_nombre(_texto(celda(fila, "Nombre_producto")))
        actividad, nombre_actividad = _codigo_y_nombre(_texto(celda(fila, "Nombre_Actividad")))
        programa.append({
            **comun,
            "PRODUCTO_PROYECTO": producto,
            "PRODUCTO_PROYECTO_NOMBRE": nombre_producto,
            "ACTIVIDAD_ACCION_OBRA": actividad,
            "ACTIVIDAD_ACCION_OBRA_NOMBRE": nombre_actividad,
        })
    libro.close()

    if len(periodos) != 1:
        raise ValueError(
            f"Se esperaba un solo Periodo en la hoja; hay {len(periodos)}: "
            f"{', '.join(sorted(periodos))}."
        )
    corte = periodos.pop()
    if not (coincide := PERIODO.match(corte)):
        raise ValueError(f"Periodo '{corte}': se esperaba el formato AAAA-MM.")
    anio = coincide.group(1)
    for fila in programa + institucional:
        fila["EJERCICIO"] = anio
        fila["CORTE"] = corte
    return programa, institucional, anio, corte


def _entidad_de_pliego(texto: str) -> tuple[str, str, str]:
    """`080201-300692: MUNICIPALIDAD PROVINCIAL DE ACOMAYO` → (SEC_EJEC, 300692, nombre).

    Misma regla que `scripts/consolidar_pp0068.py`: los gobiernos locales llevan el prefijo
    `ubigeo-secejec` y el resto solo el pliego.
    """
    prefijo, _, nombre = texto.partition(":")
    prefijo, nombre = prefijo.strip(), nombre.strip()
    if "-" in prefijo:
        return "SEC_EJEC", prefijo.split("-", 1)[1], nombre
    return "PLIEGO", prefijo, nombre


def _codigo_y_nombre(texto: str) -> tuple[str, str]:
    codigo, _, nombre = texto.partition(":")
    return codigo.strip(), nombre.strip()


# --- Escritura -------------------------------------------------------------


def _sincronizar_entidades(filas, resolutor, advertencias) -> dict[str, object]:
    """Crea o actualiza las entidades del archivo y devuelve {codigo: EntidadEjecutora}.

    Las grafías se recorren de la más reciente a la más antigua: la primera que case con el
    padrón gana. Es lo que hace que QUISPICANCHI y QUISPICANCHIS lleguen al mismo distrito sin
    una tabla de alias que haya que mantener.
    """
    from apps.inversion.models import EntidadEjecutora

    vistas: dict[str, dict] = {}
    for fila in filas:
        codigo = _texto(fila["ENTIDAD_CODIGO"])
        if not codigo:
            continue
        entrada = vistas.setdefault(
            codigo, {"nombre": "", "nivel": "", "geografias": [], "anio": ""}
        )
        anio = _texto(fila.get("EJERCICIO"))
        if anio >= entrada["anio"]:
            entrada["anio"] = anio
            entrada["nombre"] = _texto(fila.get("ENTIDAD_NOMBRE")) or entrada["nombre"]
            entrada["nivel"] = _texto(fila.get("NIVEL_GOBIERNO")) or entrada["nivel"]
        geografia = (anio, _texto(fila.get("PROVINCIA")), _texto(fila.get("DISTRITO")))
        if geografia[1] or geografia[2]:
            entrada["geografias"].append(geografia)

    entidades: dict[str, object] = {}
    sin_territorio: list[str] = []
    for codigo, datos in vistas.items():
        ambito = _ambito(datos["nivel"], datos["nombre"])
        defaults = {"nombre": datos["nombre"] or codigo, "ambito": ambito}

        # La geografía solo se toca si el archivo la trae. La serie de totales institucionales
        # no tiene columnas de provincia ni distrito, y escribir `None` desde ahí borraba el
        # territorio que había resuelto la serie del programa: 120 municipalidades se quedaban
        # «sin territorio» por el simple hecho de cargar los archivos en cierto orden.
        if datos["geografias"]:
            distrito = provincia = None
            for _anio, nombre_prov, nombre_dist in sorted(datos["geografias"], reverse=True):
                distrito = distrito or resolutor.distrito(nombre_prov, nombre_dist)
                provincia = provincia or resolutor.provincia(nombre_prov)
                if distrito and provincia:
                    break
            defaults["distrito"] = distrito
            defaults["provincia"] = provincia or (distrito.provincia if distrito else None)

        entidad, _creada = EntidadEjecutora.objects.update_or_create(
            codigo=codigo, defaults=defaults
        )
        entidades[codigo] = entidad
        if "distrito" in defaults and entidad.sin_territorio:
            sin_territorio.append(f"{entidad.nombre} ({codigo})")

    if sin_territorio:
        advertencias.append(
            f"{len(sin_territorio)} municipalidad(es) no casan con ningún distrito del padrón y "
            f"quedan sin territorio: {', '.join(sorted(sin_territorio))}. Siguen contando en los "
            f"totales, pero no pueden cruzarse con datos distritales."
        )
    return entidades


def _sincronizar_catalogo(filas, advertencias) -> dict[str, object]:
    """Descubre los códigos del archivo y los añade al catálogo sin pisar lo editado."""
    from apps.inversion.models import ClasificacionActividad, ProcesoGRD

    procesos = {p.slug: p for p in ProcesoGRD.objects.all()}
    if not procesos:
        raise ValueError(
            "No hay procesos de la GRD en la base. Corre `manage.py seed --solo-catalogos` "
            "antes de importar."
        )

    descubiertos: dict[str, dict] = {}
    for fila in filas:
        producto = _texto(fila["PRODUCTO_PROYECTO"])
        actividad = _texto(fila.get("ACTIVIDAD_ACCION_OBRA"))
        codigo = codigo_clasificable(producto, actividad)
        if not codigo:
            continue
        nombre = _texto(
            fila.get("PRODUCTO_PROYECTO_NOMBRE")
            if es_proyecto(producto)
            else fila.get("ACTIVIDAD_ACCION_OBRA_NOMBRE")
        )
        descubiertos.setdefault(codigo, {"nombre": nombre, "proyecto": es_proyecto(producto)})

    existentes = {c.codigo: c for c in ClasificacionActividad.objects.select_related("proceso")}
    nuevos = 0
    for codigo, datos in descubiertos.items():
        if codigo in existentes:
            continue
        proceso_slug = (
            PROCESO_POR_DEFECTO_PROYECTOS if datos["proyecto"] else ACTIVIDAD_A_PROCESO.get(codigo)
        )
        existentes[codigo] = ClasificacionActividad.objects.create(
            codigo=codigo,
            nombre=datos["nombre"] or codigo,
            origen=(
                ClasificacionActividad.Origen.PROYECTO
                if datos["proyecto"]
                else ClasificacionActividad.Origen.ACTIVIDAD
            ),
            proceso=procesos.get(proceso_slug),
            automatico=True,
        )
        nuevos += 1

    sin_proceso = [c for c in existentes.values() if c.proceso_id is None]
    if nuevos:
        advertencias.append(
            f"{nuevos} código(s) nuevos añadidos al catálogo de procesos de la GRD con la "
            f"clasificación propuesta. Conviene revisarlos en el admin: quedan marcados como "
            f"«asignado automáticamente»."
        )
    if sin_proceso:
        advertencias.append(
            f"{len(sin_proceso)} código(s) siguen sin proceso asignado y su importe se mostrará "
            f"como «sin clasificar»: "
            + ", ".join(sorted(c.codigo for c in sin_proceso)[:25])
            + ("…" if len(sin_proceso) > 25 else "")
            + "."
        )
    return existentes


def _ejercicios(filas, corte_por_anio, fuente_por_anio, advertencias) -> dict[str, object]:
    """Crea o actualiza los ejercicios del archivo, **sin tocar `visible`**.

    Publicar la ventana es una decisión editorial: una importación no la toma. Un ejercicio
    nuevo nace oculto y PREDES lo enciende cuando ha revisado las cifras.
    """
    from apps.inversion.models import Ejercicio

    ejercicios: dict[str, object] = {}
    for anio in sorted({_texto(f["EJERCICIO"]) for f in filas if _texto(f.get("EJERCICIO"))}):
        corte = corte_por_anio.get(anio, "anual")
        fuente = fuente_por_anio.get(anio, Ejercicio.Fuente.MEF)
        ejercicio, creado = Ejercicio.objects.get_or_create(
            anio=int(anio),
            defaults={
                "corte": corte,
                "fuente": fuente,
                "es_parcial": corte != "anual",
            },
        )
        if not creado and (ejercicio.corte, ejercicio.fuente) != (corte, fuente):
            ejercicio.corte = corte
            ejercicio.es_parcial = corte != "anual"
            ejercicio.fuente = fuente
            ejercicio.save(update_fields=["corte", "es_parcial", "fuente"])
        ejercicios[anio] = ejercicio
        if creado:
            advertencias.append(
                f"Ejercicio {anio} creado **oculto**. Se publica marcando «visible» en "
                f"Inversión → Ejercicios presupuestales."
            )
    return ejercicios


def _escribir_programa(filas, ejercicios, entidades, catalogo, upload) -> dict:
    """Reemplaza el detalle del programa de los ejercicios presentes en el archivo."""
    from apps.inversion.models import PresupuestoActividad, PresupuestoEntidad

    detalle: dict[tuple, list[Decimal]] = {}
    totales: dict[tuple, list[Decimal]] = {}
    for n, fila in enumerate(filas, start=2):
        anio = _texto(fila["EJERCICIO"])
        codigo_entidad = _texto(fila["ENTIDAD_CODIGO"])
        producto = _texto(fila["PRODUCTO_PROYECTO"])
        codigo = codigo_clasificable(producto, _texto(fila.get("ACTIVIDAD_ACCION_OBRA")))
        importes = [
            _importe(fila["PIA"], f"Fila {n}, PIA"),
            _importe(fila["PIM"], f"Fila {n}, PIM"),
            _importe(fila["DEVENGADO"], f"Fila {n}, DEVENGADO"),
        ]
        for acumulador, llave in (
            (detalle, (anio, codigo_entidad, codigo)),
            (totales, (anio, codigo_entidad)),
        ):
            actual = acumulador.setdefault(llave, [CERO, CERO, CERO])
            for i, importe in enumerate(importes):
                actual[i] += importe

    anios = set(ejercicios)
    PresupuestoActividad.objects.filter(ejercicio__anio__in=[int(a) for a in anios]).delete()
    PresupuestoActividad.objects.bulk_create(
        [
            PresupuestoActividad(
                ejercicio=ejercicios[anio],
                entidad=entidades[codigo_entidad],
                clasificacion=catalogo[codigo],
                pia=importes[0],
                pim=importes[1],
                devengado=importes[2],
            )
            for (anio, codigo_entidad, codigo), importes in detalle.items()
        ],
        batch_size=1000,
    )

    for (anio, codigo_entidad), importes in totales.items():
        PresupuestoEntidad.objects.update_or_create(
            ejercicio=ejercicios[anio],
            entidad=entidades[codigo_entidad],
            defaults={
                "pia": importes[0],
                "pim": importes[1],
                "devengado": importes[2],
                "dataset_upload": upload,
            },
        )
    return {"filas_detalle": len(detalle), "filas_entidad": len(totales)}


def _escribir_institucional(filas, ejercicios, entidades, upload) -> dict:
    """Escribe **solo** las columnas institucionales, dejando intacto el 0068 ya cargado."""
    from apps.inversion.models import PresupuestoEntidad

    escritas = 0
    for n, fila in enumerate(filas, start=2):
        anio = _texto(fila["EJERCICIO"])
        entidad = entidades[_texto(fila["ENTIDAD_CODIGO"])]
        PresupuestoEntidad.objects.update_or_create(
            ejercicio=ejercicios[anio],
            entidad=entidad,
            defaults={
                "pia_institucional": _importe(fila["PIA"], f"Fila {n}, PIA"),
                "pim_institucional": _importe(fila["PIM"], f"Fila {n}, PIM"),
                "devengado_institucional": _importe(fila["DEVENGADO"], f"Fila {n}, DEVENGADO"),
            },
        )
        escritas += 1
    return {"filas_institucional": escritas}


# --- Punto de entrada ------------------------------------------------------


def importar(upload) -> dict:
    from apps.inversion.models import Ejercicio

    nombre = (upload.archivo.name or "").lower()
    advertencias: list[str] = []
    resultado: dict = {}

    if nombre.endswith((".xlsx", ".xlsm")):
        programa, institucional, anio, corte = _leer_excel(upload)
        cortes = {anio: corte}
        fuentes = {anio: Ejercicio.Fuente.CLIENTE}
        forma = "excel"
    else:
        forma, filas = _leer_csv(upload)
        programa = filas if forma == "programa" else []
        institucional = filas if forma == "institucional" else []
        cortes, fuentes = {}, {}
        for fila in filas:
            if not (anio := _texto(fila.get("EJERCICIO"))):
                continue
            cortes[anio] = _texto(fila.get("CORTE")) or "anual"
            # La serie consolidada mezcla las dos fuentes y lo dice fila a fila: los años del
            # comparativo llevan MEF y el ejercicio nuevo, la base del cliente. Fijar una sola
            # fuente para todo el archivo etiquetaba mal el último punto de la tendencia.
            fuentes[anio] = (
                Ejercicio.Fuente.MEF
                if _texto(fila.get("FUENTE")).upper().startswith("MEF")
                else Ejercicio.Fuente.CLIENTE
            )

    todas = programa + institucional
    if not todas:
        raise ValueError("El archivo no tiene ninguna fila con datos.")

    resolutor = _Resolutor()
    if not resolutor.distritos:
        raise ValueError(
            "No hay distritos en la base: sin el padrón no se puede situar ninguna "
            "municipalidad. Importa primero el Excel de nivel de peligro."
        )

    with transaction.atomic():
        entidades = _sincronizar_entidades(todas, resolutor, advertencias)
        ejercicios = _ejercicios(todas, cortes, fuentes, advertencias)
        if programa:
            catalogo = _sincronizar_catalogo(programa, advertencias)
            resultado |= _escribir_programa(programa, ejercicios, entidades, catalogo, upload)
        if institucional:
            resultado |= _escribir_institucional(institucional, ejercicios, entidades, upload)

    visibles = Ejercicio.objects.filter(visible=True).exists()
    if not visibles:
        advertencias.append(
            "Ningún ejercicio está marcado como visible, así que /inversion sigue mostrando su "
            "estado «información en preparación». Marca el que quieras publicar."
        )

    return {
        "forma": forma,
        "filas_leidas": len(todas),
        "filas_importadas": resultado.get("filas_detalle", 0)
        + resultado.get("filas_institucional", 0),
        "ejercicios": sorted(int(a) for a in ejercicios),
        "entidades": len(entidades),
        **resultado,
        "advertencias": advertencias[:300],
    }
