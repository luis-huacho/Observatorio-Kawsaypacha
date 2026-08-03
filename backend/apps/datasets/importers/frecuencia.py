"""Importa Base_Frecuencia_Peligro_Cusco.xlsx (hoja NºEMERGENCIAS).

Formato ancho (distrito × ~25 tipos de evento con subtotales por categoría) → formato largo
`FrecuenciaEmergencia`, más `TotalDeclaradoEmergencias` para los subtotales sin desglose.

**Los TOT_* sí se guardan** (ADR-D1). El distrito de Cusco trae los cuatro subtotales llenos
(TOTAL 134) y todas las columnas de evento vacías: descartarlos dejaría a la capital regional
mostrando 0 emergencias, que es peor que no mostrar nada. Regla: si hay desglose se usa el
desglose y el declarado se ignora (registrando el descuadre en el log); si no lo hay, se guarda
el declarado y la UI dice explícitamente que la fuente no desagrega.

El Excel no trae ubigeo: el distrito se resuelve por nombre normalizado. Verificado sobre los
datos reales — 111/111 resuelven sin ambigüedad, porque en Cusco no hay distritos homónimos.
"""
from django.db import transaction

import openpyxl

from apps.peligros.catalogo import ALIAS_FUENTE, CATEGORIAS_EVENTO

HOJA = "NºEMERGENCIAS"


def _texto(valor) -> str:
    return str(valor).strip() if valor is not None else ""


def _entero(valor):
    """La columna TOTAL llega a veces como string y a veces como int, según la fila."""
    if valor is None or valor == "":
        return None
    try:
        return int(float(str(valor).strip()))
    except (TypeError, ValueError):
        return None


def _normalizar_rango(valor) -> str:
    """`2007 - 2023` y `2007-2023` son el mismo periodo escrito de dos maneras (23 variantes).

    Se guarda el texto de la fuente, solo quitando los espacios alrededor del guion: convertirlo
    a dos enteros perdería los casos que no son un rango simple.
    """
    texto = _texto(valor)
    if not texto:
        return ""
    partes = [p.strip() for p in texto.split("-")]
    return "-".join(partes) if len(partes) == 2 and all(partes) else texto


def importar(upload) -> dict:
    from apps.peligros.models import (
        CategoriaEvento,
        FrecuenciaEmergencia,
        TipoEvento,
        TotalDeclaradoEmergencias,
    )
    from apps.territorio.models import Distrito
    from apps.territorio.utils import normalizar_nombre

    wb = openpyxl.load_workbook(upload.archivo.path, read_only=True, data_only=True)
    if HOJA not in wb.sheetnames:
        raise ValueError(
            f"No se encontró la hoja '{HOJA}'. Hojas del archivo: {', '.join(wb.sheetnames)}."
        )
    ws = wb[HOJA]

    filas = ws.iter_rows(values_only=True)
    header = [_texto(c) for c in next(filas)]

    def indice(nombre: str):
        return header.index(nombre) if nombre in header else None

    idx_prov, idx_dist = indice("PROVINCIA"), indice("DISTRITO")
    if idx_prov is None or idx_dist is None:
        raise ValueError("La hoja no tiene las columnas PROVINCIA y DISTRITO.")
    idx_rango, idx_fuente, idx_link = indice("RANGO FECHA"), indice("FUENTE"), indice("LINK")

    advertencias: list[str] = []
    # Columna de cada evento y de cada subtotal, por categoría.
    idx_evento: dict[str, int] = {}
    idx_total_categoria: dict[str, int] = {}
    for cat in CATEGORIAS_EVENTO:
        for _nombre, slug, columna in cat["eventos"]:
            i = indice(columna)
            if i is None:
                advertencias.append(f"Columna de evento ausente en el archivo: '{columna}'.")
            else:
                idx_evento[slug] = i
        i = indice(cat["columna_total"])
        if i is None:
            advertencias.append(f"Columna de subtotal ausente: '{cat['columna_total']}'.")
        else:
            idx_total_categoria[cat["slug"]] = i

    categoria_de_evento = {
        slug: cat["slug"] for cat in CATEGORIAS_EVENTO for _n, slug, _c in cat["eventos"]
    }

    distritos_por_nombre: dict[str, list] = {}
    for d in Distrito.objects.select_related("provincia"):
        distritos_por_nombre.setdefault(d.nombre_normalizado, []).append(d)

    filas_leidas = 0
    desgloses: list[dict] = []
    declarados: list[dict] = []
    distritos_vistos: set[str] = set()
    #: Distritos con fila en el Excel pero sin un solo número (ni eventos ni subtotales).
    sin_dato: list[str] = []

    for n_fila, r in enumerate(filas, start=2):
        if not r or not r[idx_dist]:
            continue
        filas_leidas += 1
        nombre_dist = normalizar_nombre(_texto(r[idx_dist]))
        nombre_prov = normalizar_nombre(_texto(r[idx_prov]))

        candidatos = distritos_por_nombre.get(nombre_dist, [])
        if not candidatos:
            advertencias.append(
                f"Fila {n_fila}: no se encontró el distrito '{r[idx_dist]}' "
                f"(provincia '{r[idx_prov]}') en el padrón; la fila se omite."
            )
            continue
        if len(candidatos) > 1:
            # Hoy no ocurre en Cusco, pero desempatar por provincia es barato y evita que un
            # homónimo futuro se asigne al primero que aparezca.
            acotados = [d for d in candidatos if normalizar_nombre(d.provincia.nombre) == nombre_prov]
            if len(acotados) != 1:
                advertencias.append(
                    f"Fila {n_fila}: '{r[idx_dist]}' es ambiguo entre "
                    f"{', '.join(d.provincia.nombre for d in candidatos)}; la fila se omite."
                )
                continue
            candidatos = acotados
        distrito = candidatos[0]
        distritos_vistos.add(distrito.ubigeo)

        rango = _normalizar_rango(r[idx_rango]) if idx_rango is not None else ""
        fuente = _texto(r[idx_fuente]) if idx_fuente is not None else ""
        if fuente in ALIAS_FUENTE:
            advertencias.append(
                f"Fila {n_fila} ({distrito.nombre}): FUENTE '{fuente}' se normaliza a "
                f"'{ALIAS_FUENTE[fuente]}'."
            )
            fuente = ALIAS_FUENTE[fuente]
        link = _texto(r[idx_link]) if idx_link is not None else ""
        comun = {"distrito": distrito, "rango_fecha": rango, "fuente": fuente, "fuente_url": link}

        # --- Desglose por evento ------------------------------------------
        suma_por_categoria: dict[str, int] = {}
        filas_evento = 0
        for slug, i in idx_evento.items():
            conteo = _entero(r[i])
            if conteo is None:
                continue
            filas_evento += 1
            cat = categoria_de_evento[slug]
            suma_por_categoria[cat] = suma_por_categoria.get(cat, 0) + conteo
            desgloses.append({**comun, "evento": slug, "conteo": conteo})

        # --- Subtotales declarados (ADR-D1) -------------------------------
        # La regla es literal: si el distrito trae desglose, el declarado se ignora y solo se
        # registra el descuadre. Verificado sobre los datos reales: los 55 subtotales que
        # coexisten con desglose valen todos 0, así que guardarlos solo metería ruido y haría
        # ambiguo el `desglose_disponible` del API.
        if filas_evento:
            for cat_slug, i in idx_total_categoria.items():
                declarado = _entero(r[i])
                sumado = suma_por_categoria.get(cat_slug, 0)
                if declarado is not None and declarado != sumado:
                    advertencias.append(
                        f"{distrito.nombre}: el subtotal declarado de '{cat_slug}' es "
                        f"{declarado} pero el desglose suma {sumado}. Prevalece el desglose."
                    )
        else:
            declarados_fila = [
                {**comun, "categoria": cat_slug, "total": declarado}
                for cat_slug, i in idx_total_categoria.items()
                if (declarado := _entero(r[i])) is not None
            ]
            if declarados_fila:
                declarados.extend(declarados_fila)
                advertencias.append(
                    f"{distrito.nombre}: la fuente declara subtotales pero no desagrega por tipo "
                    f"de evento. Se muestran los totales declarados con la advertencia "
                    f"correspondiente (ADR-D1); sin esto el distrito aparecería con 0 emergencias."
                )
            else:
                # Fila presente y **enteramente vacía**: 21 distritos del archivo real. No es lo
                # mismo que declarar subtotales sin desglose (arriba) ni que declarar cero: no
                # hay ningún dato. Se cuentan aparte porque el API responderá 404 igual que para
                # ACOMAYO, y sin este aviso 21 distritos sin información quedarían indistinguibles
                # del único que ni siquiera está en el archivo.
                sin_dato.append(f"{distrito.nombre} ({distrito.ubigeo})")

    wb.close()

    # Distritos del padrón que el Excel no trae. Hoy es solo ACOMAYO (080201): el API
    # responderá 404 para ellos, que es distinto de "tiene fila y suma 0".
    sin_fila = [
        f"{d.nombre} ({d.ubigeo})"
        for d in Distrito.objects.order_by("nombre")
        if d.ubigeo not in distritos_vistos
    ]
    if sin_fila:
        advertencias.append(
            f"{len(sin_fila)} distrito(s) del padrón no tienen fila en el Excel y quedarán sin "
            f"historial de emergencias: {', '.join(sin_fila)}. Conviene pedirlas a la fuente."
        )
    if sin_dato:
        advertencias.append(
            f"{len(sin_dato)} distrito(s) tienen fila en el Excel pero ni un solo dato, ni por "
            f"evento ni como subtotal: {', '.join(sin_dato)}. Quedan sin historial igual que los "
            f"que no tienen fila, y conviene pedirlos a la fuente."
        )

    # --- Escritura atómica --------------------------------------------------
    with transaction.atomic():
        tipos = {t.slug: t for t in TipoEvento.objects.all()}
        categorias = {c.slug: c for c in CategoriaEvento.objects.all()}
        faltan = ({d["evento"] for d in desgloses} - set(tipos)) | (
            {d["categoria"] for d in declarados} - set(categorias)
        )
        if faltan:
            raise ValueError(
                "Faltan tipos o categorías de evento en la base: "
                + ", ".join(sorted(faltan))
                + ". Corre `manage.py seed --solo-catalogos` antes de importar."
            )

        FrecuenciaEmergencia.objects.all().delete()
        TotalDeclaradoEmergencias.objects.all().delete()
        FrecuenciaEmergencia.objects.bulk_create(
            [
                FrecuenciaEmergencia(
                    distrito=d["distrito"],
                    tipo_evento=tipos[d["evento"]],
                    conteo=d["conteo"],
                    rango_fecha=d["rango_fecha"],
                    fuente=d["fuente"],
                    fuente_url=d["fuente_url"],
                    dataset_upload=upload,
                )
                for d in desgloses
            ],
            batch_size=1000,
        )
        TotalDeclaradoEmergencias.objects.bulk_create(
            [
                TotalDeclaradoEmergencias(
                    distrito=d["distrito"],
                    categoria=categorias[d["categoria"]],
                    total=d["total"],
                    rango_fecha=d["rango_fecha"],
                    fuente=d["fuente"],
                    fuente_url=d["fuente_url"],
                    dataset_upload=upload,
                )
                for d in declarados
            ],
            batch_size=500,
        )

    return {
        "filas_leidas": filas_leidas,
        "filas_importadas": len(desgloses) + len(declarados),
        "distritos_en_archivo": len(distritos_vistos),
        # Los que quedan con algo que mostrar: la fila vacía no cuenta como dato.
        "distritos_con_datos": len(distritos_vistos) - len(sin_dato),
        "distritos_sin_fila": sin_fila,
        "distritos_sin_dato": sin_dato,
        "registros_desglosados": len(desgloses),
        "totales_declarados": len(declarados),
        "advertencias": advertencias[:300],
    }
