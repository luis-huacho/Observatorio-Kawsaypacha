"""Importa Base_Nivel Peligro_CCPP_Cusco.xlsx (9 hojas, una por peligro).

Verificado contra el archivo real: 8,968 CCPP × 9 hojas → 10,978 clasificaciones, de las que
solo 3,238 centros poblados tienen alguna. Las anomalías conocidas (229 filas sin nivel, 2 sin
código, SICUANI sin distrito en una hoja, dos grafías de la fuente) van al log **sin abortar**:
el log es lo que PREDES lee para corregir su Excel, así que cita hoja y fila.
"""
from django.db import transaction

import openpyxl

from apps.peligros.catalogo import ALIAS_FUENTE, PELIGRO_POR_HOJA, PELIGRO_POR_NOMBRE

COLUMNAS_ESPERADAS = [
    "DEPARTAMEN", "PROVINCIA", "DISTRITO", "CODIGO", "NOMB_CPOB", "CATEGORIA",
    "ALTITUD", "LONGITUD", "LATITUD", "POBLACION", "PELIGRO", "TIP_PELIG",
    "NIVEL_PELI", "Fuente", "Link",
]

# Índices de columna, para que el desempaquetado no dependa del orden literal del `for`.
C_PROV, C_DIST, C_CODIGO, C_NOMBRE, C_CATEGORIA = 1, 2, 3, 4, 5
C_ALTITUD, C_LON, C_LAT, C_POBLACION = 6, 7, 8, 9
C_PELIGRO, C_TIP_PELIG, C_NIVEL, C_FUENTE, C_LINK = 10, 11, 12, 13, 14


def _texto(valor) -> str:
    return str(valor).strip() if valor is not None else ""


def _entero(valor):
    if valor is None or valor == "":
        return None
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return None


def _decimal(valor):
    if valor is None or valor == "":
        return None
    try:
        return float(valor)
    except (TypeError, ValueError):
        return None


def _mejor(actual: str, nuevo: str) -> str:
    """Al deduplicar un CCPP entre las 9 hojas, gana el valor no vacío.

    Sin esto SICUANI se queda sin distrito: una de las hojas lo trae en blanco y quedarse con
    el primero que aparece es cuestión de en qué orden se leyeron las hojas.
    """
    return actual or nuevo


def importar(upload) -> dict:
    from apps.peligros.models import ClasificacionPeligro, Fuente, TipoPeligro
    from apps.territorio.models import CentroPoblado, Distrito, Provincia

    wb = openpyxl.load_workbook(upload.archivo.path, read_only=True, data_only=True)
    advertencias: list[str] = []
    filas_leidas = 0
    descartadas_sin_nivel = 0
    descartadas_sin_codigo = 0

    # --- Validación de estructura ------------------------------------------
    hojas_desconocidas = [h for h in wb.sheetnames if h not in PELIGRO_POR_HOJA]
    if hojas_desconocidas:
        advertencias.append(
            "Hojas no reconocidas, se ignoran: " + ", ".join(hojas_desconocidas)
        )
    hojas_faltantes = [h for h in PELIGRO_POR_HOJA if h not in wb.sheetnames]
    if hojas_faltantes:
        advertencias.append("Hojas esperadas ausentes: " + ", ".join(hojas_faltantes))

    hojas = [h for h in wb.sheetnames if h in PELIGRO_POR_HOJA]
    if not hojas:
        raise ValueError(
            "El archivo no tiene ninguna de las 9 hojas esperadas "
            f"({', '.join(PELIGRO_POR_HOJA)})."
        )
    for hoja in hojas:
        header = next(wb[hoja].iter_rows(max_row=1, values_only=True), None)
        leidas = [_texto(c) for c in (header or [])][:15]
        if leidas != COLUMNAS_ESPERADAS:
            raise ValueError(
                f"Hoja '{hoja}': columnas inesperadas.\n"
                f"Se esperaba: {', '.join(COLUMNAS_ESPERADAS)}\n"
                f"Se encontró: {', '.join(leidas) or '(vacío)'}"
            )

    # --- Lectura -----------------------------------------------------------
    ccpp_por_codigo: dict[str, dict] = {}
    clasificaciones: list[dict] = []
    nombres_fuente: set[str] = set()

    for hoja in hojas:
        catalogo = PELIGRO_POR_HOJA[hoja]
        for n_fila, r in enumerate(wb[hoja].iter_rows(min_row=2, values_only=True), start=2):
            if not r or all(c is None for c in r):
                continue
            filas_leidas += 1

            codigo = _texto(r[C_CODIGO])
            if not codigo:
                descartadas_sin_codigo += 1
                advertencias.append(f"{hoja}, fila {n_fila}: sin CODIGO, se descarta la fila.")
                continue
            if len(codigo) != 10 or not codigo.isdigit():
                advertencias.append(
                    f"{hoja}, fila {n_fila}: CODIGO '{codigo}' no tiene 10 dígitos, se descarta."
                )
                continue

            datos = ccpp_por_codigo.setdefault(
                codigo,
                {"nombre": "", "categoria": "", "provincia": "", "distrito": "",
                 "lat": None, "lon": None, "altitud": None, "poblacion": None},
            )
            datos["nombre"] = _mejor(datos["nombre"], _texto(r[C_NOMBRE]))
            datos["categoria"] = _mejor(datos["categoria"], _texto(r[C_CATEGORIA]))
            datos["provincia"] = _mejor(datos["provincia"], _texto(r[C_PROV]))
            datos["distrito"] = _mejor(datos["distrito"], _texto(r[C_DIST]))
            for campo, valor in (
                ("lat", _decimal(r[C_LAT])),
                ("lon", _decimal(r[C_LON])),
                ("altitud", _entero(r[C_ALTITUD])),
                ("poblacion", _entero(r[C_POBLACION])),
            ):
                if datos[campo] is None:
                    datos[campo] = valor

            peligro_fila = _texto(r[C_PELIGRO])
            if not peligro_fila:
                continue  # fila del padrón sin evaluación para este peligro: no es anomalía

            # El nombre del peligro sale de la columna, no del título de la hoja.
            if peligro_fila != catalogo["nombre"]:
                if peligro_fila in PELIGRO_POR_NOMBRE:
                    advertencias.append(
                        f"{hoja}, fila {n_fila}: la columna PELIGRO dice '{peligro_fila}' y la "
                        f"hoja corresponde a '{catalogo['nombre']}'. Se usa el valor de la columna."
                    )
                    catalogo_fila = PELIGRO_POR_NOMBRE[peligro_fila]
                else:
                    advertencias.append(
                        f"{hoja}, fila {n_fila}: PELIGRO '{peligro_fila}' no está en el catálogo, "
                        f"se descarta la clasificación."
                    )
                    continue
            else:
                catalogo_fila = catalogo

            tip_pelig = _texto(r[C_TIP_PELIG])
            if tip_pelig and tip_pelig != catalogo_fila["categoria_geo"]:
                advertencias.append(
                    f"{hoja}, fila {n_fila}: TIP_PELIG '{tip_pelig}' no coincide con el catálogo "
                    f"('{catalogo_fila['categoria_geo']}'). Prevalece el catálogo."
                )

            nivel = _entero(r[C_NIVEL])
            if nivel is None:
                # 229 filas del archivo real: traen peligro y respaldo documental pero sin
                # nivel. Sin nivel no hay semáforo, y asumir 1 sería inventar un dato.
                descartadas_sin_nivel += 1
                continue
            if not 1 <= nivel <= 4:
                advertencias.append(
                    f"{hoja}, fila {n_fila}: NIVEL_PELI '{nivel}' fuera de 1-4 en CCPP {codigo}, "
                    f"se descarta."
                )
                continue

            fuente = _texto(r[C_FUENTE])
            fuente = ALIAS_FUENTE.get(fuente, fuente)
            if fuente:
                nombres_fuente.add(fuente)

            clasificaciones.append({
                "codigo": codigo,
                "slug_peligro": catalogo_fila["slug"],
                "nivel": nivel,
                "fuente": fuente,
                "fuente_url": _texto(r[C_LINK]),
            })
    wb.close()

    if descartadas_sin_nivel:
        advertencias.append(
            f"{descartadas_sin_nivel} filas traen PELIGRO y respaldo documental pero sin "
            f"NIVEL_PELI: se descartan (sin nivel no hay semáforo, y asumir 'bajo' sería "
            f"inventar el dato). Conviene pedir a la fuente que las complete."
        )

    # --- Escritura atómica --------------------------------------------------
    # Todo-o-nada: si algo falla aquí, los datos activos previos quedan intactos.
    with transaction.atomic():
        _sincronizar_territorio(ccpp_por_codigo, Provincia, Distrito, CentroPoblado)

        tipos = {t.slug: t for t in TipoPeligro.objects.all()}
        faltan = {c["slug_peligro"] for c in clasificaciones} - set(tipos)
        if faltan:
            raise ValueError(
                "Faltan tipos de peligro en el catálogo de la base: "
                + ", ".join(sorted(faltan))
                + ". Corre `manage.py seed --solo-catalogos` antes de importar."
            )

        fuentes = {}
        for nombre in nombres_fuente:
            obj, _ = Fuente.objects.get_or_create(nombre=nombre, defaults={"sigla": nombre})
            fuentes[nombre] = obj

        pk_ccpp = dict(CentroPoblado.objects.values_list("codigo", "pk"))
        objetos = [
            ClasificacionPeligro(
                centro_poblado_id=pk_ccpp[c["codigo"]],
                tipo_peligro=tipos[c["slug_peligro"]],
                nivel=c["nivel"],
                fuente=fuentes.get(c["fuente"]),
                fuente_url=c["fuente_url"],
                dataset_upload=upload,
            )
            for c in clasificaciones
        ]
        ClasificacionPeligro.objects.all().delete()
        ClasificacionPeligro.objects.bulk_create(objetos, batch_size=2000)

    ccpp_clasificados = len({c["codigo"] for c in clasificaciones})
    return {
        "filas_leidas": filas_leidas,
        "filas_importadas": len(objetos),
        "centros_poblados": len(ccpp_por_codigo),
        "centros_poblados_clasificados": ccpp_clasificados,
        "centros_poblados_sin_dato": len(ccpp_por_codigo) - ccpp_clasificados,
        "clasificaciones": len(objetos),
        "descartadas_sin_nivel": descartadas_sin_nivel,
        "descartadas_sin_codigo": descartadas_sin_codigo,
        "fuentes": sorted(nombres_fuente),
        "advertencias": advertencias[:300],
    }


def _sincronizar_territorio(ccpp_por_codigo, Provincia, Distrito, CentroPoblado) -> None:
    """Upsert de provincias, distritos y centros poblados derivando el ubigeo del código INEI.

    No se borra nada: los CCPP son el padrón, y otras tablas (medidas, frecuencias) los
    referencian. Un CCPP que desaparezca del Excel se marcaría como no vigente, no se elimina.
    """
    provincias = {}
    for codigo, datos in ccpp_por_codigo.items():
        provincias.setdefault(codigo[:4], datos["provincia"])
    existentes = set(Provincia.objects.values_list("ubigeo", flat=True))
    Provincia.objects.bulk_create(
        [Provincia(ubigeo=u, nombre=n) for u, n in provincias.items() if u not in existentes]
    )
    pk_provincia = dict(Provincia.objects.values_list("ubigeo", "pk"))

    distritos = {}
    for codigo, datos in ccpp_por_codigo.items():
        actual = distritos.get(codigo[:6], "")
        distritos[codigo[:6]] = _mejor(actual, datos["distrito"])
    existentes = set(Distrito.objects.values_list("ubigeo", flat=True))
    nuevos = [
        Distrito(ubigeo=u, nombre=n, provincia_id=pk_provincia[u[:4]])
        for u, n in distritos.items()
        if u not in existentes
    ]
    for d in nuevos:  # bulk_create no llama a save(), que es donde se normaliza el nombre
        from apps.territorio.utils import normalizar_nombre

        d.nombre_normalizado = normalizar_nombre(d.nombre)
    Distrito.objects.bulk_create(nuevos)

    campos = ["nombre", "categoria", "lat", "lon", "altitud", "poblacion"]
    pk_por_codigo = dict(CentroPoblado.objects.values_list("codigo", "pk"))
    nuevos_ccpp, actualizar = [], []
    for codigo, d in ccpp_por_codigo.items():
        obj = CentroPoblado(
            codigo=codigo,
            distrito_id=codigo[:6],
            nombre=d["nombre"],
            categoria=d["categoria"],
            lat=d["lat"],
            lon=d["lon"],
            altitud=d["altitud"],
            poblacion=d["poblacion"],
        )
        if codigo in pk_por_codigo:
            obj.pk = pk_por_codigo[codigo]
            actualizar.append(obj)
        else:
            nuevos_ccpp.append(obj)
    CentroPoblado.objects.bulk_create(nuevos_ccpp, batch_size=2000)
    if actualizar:
        CentroPoblado.objects.bulk_update(actualizar, campos, batch_size=2000)
