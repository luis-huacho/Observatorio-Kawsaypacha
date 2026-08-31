#!/usr/bin/env python3
"""Justifica, cifra por cifra, lo que la ventana /inversion muestra para una provincia.

La pantalla enseña ocho números y ninguno está escrito en ningún sitio fuera del código. Este
script produce el libro que permite recalcularlos **desde el archivo del cliente**, sin leer
Python: cada cifra de cabecera queda apoyada en una fórmula viva que apunta a las filas de
detalle, y esas filas llevan el número de fila que ocupan en el Excel original.

Concilia **tres capas**, que es lo que distingue una auditoría de una repetición:

1. **El Excel del cliente** — se relee y se vuelve a sumar aquí, sin pasar por la plataforma.
2. **La base de datos** — vía `GET /api/inversion/entidades/`, que sirve las filas de
   `PresupuestoEntidad` tal como las escribió el importador, y `…/entidades/<codigo>/`, que
   sirve las de `PresupuestoActividad` con su proceso de la GRD.
3. **La pantalla** — vía `GET /api/inversion/`, el bloque `agregados` que alimenta los KPI.

Si las tres no coinciden, el script **termina con código distinto de cero**. Un libro que se
genera sin quejarse es un libro cuyas tres capas cuadran; sin esa guarda, la auditoría se
limitaría a copiar lo que ya dice la pantalla.

Sin `--api` el libro se genera igual, con las columnas de la plataforma en «no consultado». No
se inventa la cifra que no se pudo pedir.

Ejemplo:

    python3 scripts/auditar_inversion.py \\
        data/inversion/Base_Prespuesto_PP0068_cusco_final.xlsx \\
        _docs/auditoria-inversion-cusco-2026.xlsx \\
        --ubigeo 0801 --api http://localhost:8000
"""

from __future__ import annotations

import argparse
import json
import sys
import urllib.error
import urllib.request
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from consolidar_pp0068 import a_numero, entidad_de_pliego

#: Etiqueta con la que el Excel marca el total de la entidad. Esa fila **incluye** al 0068:
#: sumarla junto al desglose del programa cuenta el programa dos veces.
CATEGORIA_INSTITUCIONAL = "PRESUPUESTO INSTITUCIONAL"

#: Prefijo de la categoría del programa. El backend no lo comprueba (manda al bloque «programa»
#: toda fila que no sea institucional), así que comprobarlo aquí es parte de lo que se audita.
PROGRAMA = "0068"

#: Verde de la paleta PREDES, el mismo de `backend/apps/api/exports.py`.
VERDE = "0B3B26"
GRIS = "F1F5F9"

SOLES = '#,##0'
PORCENTAJE = '0.00%'

#: Tolerancia al comparar importes. Los tres orígenes son enteros de soles; un céntimo de
#: diferencia ya es una discrepancia que hay que mirar, no un redondeo.
EPSILON = 0.005


# --- Lectura del Excel del cliente -----------------------------------------


def _texto(valor) -> str:
    """Texto limpio de una celda. Quita también el NBSP y los tabuladores del archivo."""
    if valor is None:
        return ""
    return str(valor).replace("\xa0", " ").replace("\t", " ").strip()


def partes_de_pliego(texto: str) -> tuple[str, str, str]:
    """`080101-300684: MUNICIPALIDAD PROVINCIAL DEL CUZCO` → (ubigeo, código SIAF, nombre).

    El código y el nombre salen de `entidad_de_pliego`, la misma función que usan los dos
    scripts de consolidación: si el parseo del pliego tuviera un fallo, la auditoría lo
    heredaría en vez de taparlo con una segunda implementación que quizá acierte por otro lado.

    El ubigeo, en cambio, esa función lo descarta, y es lo que aquí decide la provincia.
    """
    (_, codigo), nombre = entidad_de_pliego(texto)
    prefijo = str(texto).partition(":")[0].strip()
    ubigeo = prefijo.split("-", 1)[0] if "-" in prefijo else ""
    return ubigeo, codigo, nombre


def codigo_y_nombre(texto: str) -> tuple[str, str]:
    """Parte `3000001: ACCIONES COMUNES` en código y nombre."""
    codigo, _, nombre = _texto(texto).partition(":")
    return codigo.strip(), nombre.strip()


def es_proyecto(codigo_producto: str) -> bool:
    """Un producto que empieza en 2 es un proyecto de inversión (misma regla que el backend)."""
    return codigo_producto.startswith("2")


def codigo_clasificable(codigo_producto: str, codigo_actividad: str) -> str:
    """El código cuyo nombre dice de qué proceso de la GRD es la fila.

    Réplica de `apps.inversion.catalogo.codigo_clasificable`. Se repite aquí, y no se importa,
    porque el script corre fuera de Django: importarlo obligaría a levantar el proyecto entero
    para leer un Excel. La réplica se comprueba sola —el proceso de cada fila se cruza contra
    el que sirve el API—, así que una divergencia saldría en el libro y no en silencio.
    """
    return codigo_producto if es_proyecto(codigo_producto) else codigo_actividad


def leer_excel(ruta: str, ubigeo_provincia: str):
    """Devuelve (filas del programa, filas institucionales, corte, resumen del archivo).

    Solo las de la provincia pedida. El filtro es el **ubigeo que va dentro del `Pliego`**, no
    la columna `Provincia`: el ubigeo es el mismo dato con el que el backend resuelve el
    territorio, y la columna de nombre se usa después para contrastar que las dos coinciden.
    """
    libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hojas = [h for h in libro.sheetnames if h.lower().startswith("base")]
    if not hojas:
        raise SystemExit(
            f"El Excel no tiene ninguna hoja que empiece por 'Base'. Hojas: "
            f"{', '.join(libro.sheetnames)}."
        )
    hoja = libro[hojas[0]]

    filas = hoja.iter_rows(values_only=True)
    cabecera = [_texto(c) for c in next(filas)]
    indice = {nombre: i for i, nombre in enumerate(cabecera)}
    obligatorias = ["Periodo", "Provincia", "Distrito", "Pliego", "Categoría_Presupuestal",
                    "Nivel", "Nombre_producto", "Nombre_Actividad", "PIA", "PIM", "Devengado"]
    if faltan := [c for c in obligatorias if c not in indice]:
        raise SystemExit(f"A la hoja '{hojas[0]}' le faltan columnas: {', '.join(faltan)}.")
    # La cabecera del archivo termina en espacio duro: '% Ejecución\xa0'. `_texto` ya lo
    # normalizó, pero dejarlo escrito evita que alguien "arregle" la normalización.
    col_pct = indice.get("% Ejecución")

    programa, institucional = [], []
    periodos, categorias = set(), {}
    total_filas = vacias = 0
    pliegos_todos, pliegos_provincia = set(), set()

    for n, fila in enumerate(filas, start=2):
        periodo = _texto(fila[indice["Periodo"]] if indice["Periodo"] < len(fila) else None)
        if not periodo:
            vacias += 1
            continue
        total_filas += 1
        periodos.add(periodo)

        def celda(nombre):
            i = indice.get(nombre)
            return fila[i] if i is not None and i < len(fila) else None

        crudo = _texto(celda("Pliego"))
        ubigeo, codigo, nombre = partes_de_pliego(crudo)
        pliegos_todos.add(codigo)
        categoria = _texto(celda("Categoría_Presupuestal"))
        categorias[categoria] = categorias.get(categoria, 0) + 1

        if ubigeo[:4] != ubigeo_provincia:
            continue
        pliegos_provincia.add(codigo)

        registro = {
            "fila": n,
            "pliego": crudo,
            "ubigeo": ubigeo,
            "codigo": codigo,
            "nombre": nombre,
            "provincia": _texto(celda("Provincia")),
            "distrito": _texto(celda("Distrito")),
            "categoria": categoria,
            "pia": a_numero(celda("PIA"), "PIA"),
            "pim": a_numero(celda("PIM"), "PIM"),
            "devengado": a_numero(celda("Devengado"), "Devengado"),
            "pct_archivo": (fila[col_pct] if col_pct is not None and col_pct < len(fila)
                            else None),
        }

        if categoria == CATEGORIA_INSTITUCIONAL:
            institucional.append(registro)
            continue

        producto, nombre_producto = codigo_y_nombre(celda("Nombre_producto"))
        actividad, nombre_actividad = codigo_y_nombre(celda("Nombre_Actividad"))
        registro |= {
            "nivel": _texto(celda("Nivel")),
            "producto": producto,
            "nombre_producto": nombre_producto,
            "actividad": actividad,
            "nombre_actividad": nombre_actividad,
            "clasificable": codigo_clasificable(producto, actividad),
        }
        programa.append(registro)

    libro.close()

    if len(periodos) != 1:
        raise SystemExit(
            f"Se esperaba un solo Periodo en la hoja; hay {len(periodos)}: "
            f"{', '.join(sorted(periodos))}."
        )
    corte = periodos.pop()

    programa.sort(key=lambda r: (r["codigo"], r["producto"], r["actividad"]))
    institucional.sort(key=lambda r: r["codigo"])

    resumen = {
        "hoja": hojas[0],
        "filas_con_datos": total_filas,
        "filas_vacias": vacias,
        "categorias": categorias,
        "pliegos_archivo": len(pliegos_todos),
        "pliegos_provincia": len(pliegos_provincia),
    }
    return programa, institucional, corte, resumen


def revisar_archivo(programa, institucional, advertencias: list[str]) -> None:
    """Comprobaciones sobre la porción del archivo que se audita, no sobre el archivo entero."""
    for registro in programa:
        if not registro["categoria"].startswith(PROGRAMA):
            advertencias.append(
                f"Fila {registro['fila']}: la categoría es «{registro['categoria']}», que no es "
                f"el programa {PROGRAMA}. El importador la trataría como 0068 igualmente."
            )
        if registro["devengado"] > registro["pim"]:
            advertencias.append(
                f"Fila {registro['fila']}: el devengado ({registro['devengado']:,.0f}) supera al "
                f"PIM ({registro['pim']:,.0f})."
            )
        for campo in ("pia", "pim", "devengado"):
            if registro[campo] < 0:
                advertencias.append(f"Fila {registro['fila']}: {campo.upper()} negativo.")

    por_entidad: dict[str, list[dict]] = {}
    for registro in institucional:
        por_entidad.setdefault(registro["codigo"], []).append(registro)
    for codigo, filas in sorted(por_entidad.items()):
        if len(filas) > 1:
            advertencias.append(
                f"{codigo} ({filas[0]['nombre']}) trae {len(filas)} filas de "
                f"{CATEGORIA_INSTITUCIONAL.lower()} (filas "
                f"{', '.join(str(f['fila']) for f in filas)}): el denominador sería ambiguo."
            )

    con_institucional = set(por_entidad)
    for codigo in sorted({r["codigo"] for r in programa} - con_institucional):
        nombre = next(r["nombre"] for r in programa if r["codigo"] == codigo)
        advertencias.append(
            f"{codigo} ({nombre}) no trae fila de {CATEGORIA_INSTITUCIONAL.lower()}: se queda "
            f"sin denominador y no cuenta en «sobre N municipalidad(es)»."
        )


# --- La plataforma: base de datos y pantalla -------------------------------


def _pedir(url: str):
    with urllib.request.urlopen(url, timeout=20) as respuesta:
        return json.load(respuesta)


def consultar_api(base: str, ubigeo: str, anio: int, advertencias: list[str]):
    """Trae las dos capas de la plataforma, o `None` si no responde.

    `agregados` es lo que se pinta en los KPI. `entidades` son las filas de
    `PresupuestoEntidad` sin agregar, y `actividades` las de `PresupuestoActividad`: entre las
    dos y el Excel se ve **dónde** se rompería una cifra si se rompiera — en la lectura del
    archivo, en la escritura, o en la suma de cabecera.
    """
    base = base.rstrip("/")
    try:
        tablero = _pedir(f"{base}/api/inversion/?provincia={ubigeo}&anio={anio}")
        listado = _pedir(f"{base}/api/inversion/entidades/?provincia={ubigeo}&anio={anio}"
                         f"&page_size=200")
    except (urllib.error.URLError, TimeoutError, ValueError) as error:
        advertencias.append(
            f"No se pudo consultar {base}: {error}. El libro se genera sin las columnas de la "
            f"plataforma; las cifras del Excel siguen siendo válidas por sí solas."
        )
        return None

    if not tablero.get("disponible"):
        advertencias.append(
            f"El API responde «no disponible» ({tablero.get('motivo', 'sin motivo')}): el "
            f"ejercicio {anio} no está visible. No hay pantalla contra la que conciliar."
        )
        return None

    entidades = listado.get("results", [])
    actividades = []
    for entidad in entidades:
        detalle = _pedir(f"{base}/api/inversion/entidades/{entidad['codigo']}/?anio={anio}")
        for actividad in detalle.get("actividades", []):
            actividades.append({**actividad, "entidad": entidad["codigo"]})

    return {
        "tablero": tablero,
        "entidades": {e["codigo"]: e for e in entidades},
        "actividades": {(a["entidad"], a["codigo"]): a for a in actividades},
    }


# --- Construcción del libro ------------------------------------------------


BORDE = Border(bottom=Side(style="thin", color="CBD5E1"))


def _cabecera(hoja, fila: int, titulos: list[str], anchos: list[int]) -> None:
    for i, (titulo, ancho) in enumerate(zip(titulos, anchos), start=1):
        celda = hoja.cell(row=fila, column=i, value=titulo)
        celda.fill = PatternFill("solid", fgColor=VERDE)
        celda.font = Font(color="FFFFFF", bold=True, size=10)
        celda.alignment = Alignment(vertical="center", wrap_text=True)
        hoja.column_dimensions[get_column_letter(i)].width = ancho
    hoja.row_dimensions[fila].height = 32


def _titulo(hoja, texto: str, subtitulo: str = "") -> int:
    """Escribe el encabezado de una hoja y devuelve la fila en que sigue el contenido."""
    celda = hoja.cell(row=1, column=1, value=texto)
    celda.font = Font(bold=True, size=13, color=VERDE)
    if subtitulo:
        nota = hoja.cell(row=2, column=1, value=subtitulo)
        nota.font = Font(size=9, italic=True, color="475569")
        nota.alignment = Alignment(vertical="top", wrap_text=True)
        hoja.row_dimensions[2].height = 30
        return 4
    return 3


def _totales(hoja, fila: int, columnas: str, desde: int, hasta: int, formato: str) -> None:
    for letra in columnas:
        celda = hoja[f"{letra}{fila}"]
        celda.value = f"=SUM({letra}{desde}:{letra}{hasta})"
        celda.number_format = formato
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor=GRIS)


def _marcar_total(hoja, fila: int, ultima_columna: int) -> None:
    for i in range(1, ultima_columna + 1):
        celda = hoja.cell(row=fila, column=i)
        celda.font = Font(bold=True)
        if celda.fill.fgColor.rgb in (None, "00000000"):
            celda.fill = PatternFill("solid", fgColor=GRIS)


def hoja_detalle(libro, programa, api, contexto) -> tuple[int, int]:
    """Las filas del 0068 del Excel, una a una. Devuelve (primera fila, última fila)."""
    hoja = libro.create_sheet("Detalle_0068")
    inicio = _titulo(
        hoja,
        f"Detalle del PP 0068 — {contexto['provincia']}, ejercicio {contexto['anio']}",
        f"Una fila por cada registro del archivo «{contexto['archivo']}», hoja "
        f"«{contexto['hoja']}», cuya Categoría_Presupuestal es el programa {PROGRAMA} y cuyo "
        f"Pliego lleva un ubigeo de la provincia {contexto['ubigeo']}. La columna «Fila del "
        f"Excel» es el número de fila del archivo original: ahí se comprueba cada importe.",
    )
    titulos = ["Fila del Excel", "Código SIAF", "Municipalidad", "Nivel", "Cód. producto",
               "Producto o proyecto", "Cód. actividad", "Actividad o acción de obra",
               "Cód. clasificable", "Proceso GRD (plataforma)", "PIA", "PIM", "Devengado",
               "% Ejec. del archivo", "% Ejec. recalculado", "Δ %"]
    _cabecera(hoja, inicio, titulos, [11, 11, 30, 10, 12, 42, 12, 42, 12, 20, 12, 13, 13, 13,
                                      14, 9])
    hoja.freeze_panes = f"D{inicio + 1}"

    fila = inicio + 1
    primera = fila
    for registro in programa:
        clave = (registro["codigo"], registro["clasificable"])
        proceso = api["actividades"].get(clave, {}).get("proceso") if api else None
        hoja.append([
            registro["fila"], registro["codigo"], registro["nombre"], registro["nivel"],
            registro["producto"], registro["nombre_producto"], registro["actividad"],
            registro["nombre_actividad"], registro["clasificable"],
            proceso if proceso else ("no consultado" if api is None else "SIN CORRESPONDENCIA"),
            registro["pia"], registro["pim"], registro["devengado"], registro["pct_archivo"],
        ])
        hoja[f"O{fila}"] = f"=IF(L{fila}=0,\"\",M{fila}/L{fila})"
        hoja[f"P{fila}"] = f"=IF(OR(N{fila}=\"\",O{fila}=\"\"),\"\",O{fila}-N{fila})"
        for letra in "KLM":
            hoja[f"{letra}{fila}"].number_format = SOLES
        for letra in "NOP":
            hoja[f"{letra}{fila}"].number_format = PORCENTAJE
        for i in range(1, len(titulos) + 1):
            hoja.cell(row=fila, column=i).border = BORDE
            hoja.cell(row=fila, column=i).alignment = Alignment(vertical="top", wrap_text=True)
        fila += 1

    ultima = fila - 1
    hoja[f"C{fila}"] = f"TOTAL — {len(programa)} filas del archivo"
    _totales(hoja, fila, "KLM", primera, ultima, SOLES)
    hoja[f"O{fila}"] = f"=M{fila}/L{fila}"
    hoja[f"O{fila}"].number_format = PORCENTAJE
    _marcar_total(hoja, fila, len(titulos))
    return primera, ultima


def hoja_institucional(libro, institucional, contexto) -> tuple[int, int]:
    """Las filas `PRESUPUESTO INSTITUCIONAL`, que son el denominador del 0.8 %."""
    hoja = libro.create_sheet("Institucional")
    inicio = _titulo(
        hoja,
        f"Presupuesto institucional — {contexto['provincia']}, ejercicio {contexto['anio']}",
        f"Una fila por municipalidad, la que el archivo marca «{CATEGORIA_INSTITUCIONAL}». Es "
        f"el presupuesto de la entidad ENTERA, todos los programas, e **incluye al 0068**: "
        f"sumar estas filas junto a las de «Detalle_0068» contaría el programa dos veces.",
    )
    titulos = ["Fila del Excel", "Código SIAF", "Pliego (texto del archivo)", "Municipalidad",
               "Provincia", "Distrito", "PIA", "PIM", "Devengado", "% Ejec. del archivo",
               "% Ejec. recalculado"]
    _cabecera(hoja, inicio, titulos, [11, 11, 52, 32, 12, 16, 15, 15, 15, 14, 14])
    hoja.freeze_panes = f"D{inicio + 1}"

    fila = inicio + 1
    primera = fila
    for registro in institucional:
        hoja.append([
            registro["fila"], registro["codigo"], registro["pliego"], registro["nombre"],
            registro["provincia"], registro["distrito"], registro["pia"], registro["pim"],
            registro["devengado"], registro["pct_archivo"],
        ])
        hoja[f"K{fila}"] = f"=IF(H{fila}=0,\"\",I{fila}/H{fila})"
        for letra in "GHI":
            hoja[f"{letra}{fila}"].number_format = SOLES
        for letra in "JK":
            hoja[f"{letra}{fila}"].number_format = PORCENTAJE
        for i in range(1, len(titulos) + 1):
            hoja.cell(row=fila, column=i).border = BORDE
        fila += 1

    ultima = fila - 1
    hoja[f"D{fila}"] = f"TOTAL — {len(institucional)} municipalidades"
    _totales(hoja, fila, "GHI", primera, ultima, SOLES)
    _marcar_total(hoja, fila, len(titulos))
    return primera, ultima


def hoja_municipalidades(libro, institucional, api, rango_detalle, contexto):
    """Las 8 municipalidades, con el 0068 sumado DESDE el detalle y el institucional al lado."""
    hoja = libro.create_sheet("Municipalidades")
    inicio = _titulo(
        hoja,
        f"Municipalidades de {contexto['provincia']} — ejercicio {contexto['anio']}, corte "
        f"{contexto['corte']}",
        "Las columnas del 0068 no llevan números pegados: son SUMIF sobre «Detalle_0068», así "
        "que si se corrige un importe en el detalle, esta hoja y las cifras de cabecera "
        "cambian con él. Las institucionales sí son el valor de la fila del archivo, y la "
        "columna «Fila del Excel» dice cuál.",
    )
    titulos = ["Fila del Excel", "Código SIAF", "Municipalidad", "Ubigeo distrito", "Ámbito",
               "PIA 0068", "PIM 0068", "Devengado 0068", "Saldo", "% ejecución",
               "PIA institucional", "PIM institucional", "Devengado institucional",
               "% del 0068 sobre el PIM institucional", "PIM en la base de datos",
               "Devengado en la base", "¿Cuadra con la base?"]
    _cabecera(hoja, inicio, titulos, [11, 11, 34, 13, 12, 13, 14, 15, 14, 11, 17, 17, 18, 18,
                                      16, 16, 16])
    hoja.freeze_panes = f"D{inicio + 1}"

    d0, d1 = rango_detalle
    codigos = f"Detalle_0068!$B${d0}:$B${d1}"
    fila = inicio + 1
    primera = fila
    for registro in institucional:
        entidad = (api or {}).get("entidades", {}).get(registro["codigo"], {})
        hoja.append([
            registro["fila"], registro["codigo"], registro["nombre"],
            entidad.get("ubigeo_distrito", registro["ubigeo"]),
            entidad.get("ambito", "no consultado" if api is None else "sin dato"),
        ])
        for letra, columna_detalle in (("F", "K"), ("G", "L"), ("H", "M")):
            hoja[f"{letra}{fila}"] = (
                f"=SUMIF({codigos},$B{fila},Detalle_0068!${columna_detalle}${d0}:"
                f"${columna_detalle}${d1})"
            )
        hoja[f"I{fila}"] = f"=G{fila}-H{fila}"
        hoja[f"J{fila}"] = f"=IF(G{fila}=0,\"\",H{fila}/G{fila})"
        hoja[f"K{fila}"] = registro["pia"]
        hoja[f"L{fila}"] = registro["pim"]
        hoja[f"M{fila}"] = registro["devengado"]
        hoja[f"N{fila}"] = f"=IF(L{fila}=0,\"\",G{fila}/L{fila})"
        if api:
            hoja[f"O{fila}"] = entidad.get("pim")
            hoja[f"P{fila}"] = entidad.get("devengado")
            hoja[f"Q{fila}"] = (
                f'=IF(AND(ABS(G{fila}-O{fila})<{EPSILON},ABS(H{fila}-P{fila})<{EPSILON}),'
                f'"sí","REVISAR")'
            )
        else:
            hoja[f"O{fila}"] = hoja[f"P{fila}"] = hoja[f"Q{fila}"] = "no consultado"
        for letra in "FGHIKLMOP":
            hoja[f"{letra}{fila}"].number_format = SOLES
        for letra in "JN":
            hoja[f"{letra}{fila}"].number_format = PORCENTAJE
        for i in range(1, len(titulos) + 1):
            hoja.cell(row=fila, column=i).border = BORDE
        fila += 1

    ultima = fila - 1
    hoja[f"C{fila}"] = f"TOTAL — {len(institucional)} municipalidades"
    _totales(hoja, fila, "FGHKLM" + ("OP" if api else ""), primera, ultima, SOLES)
    hoja[f"I{fila}"] = f"=G{fila}-H{fila}"
    hoja[f"J{fila}"] = f"=H{fila}/G{fila}"
    hoja[f"N{fila}"] = f"=G{fila}/L{fila}"
    if api:
        hoja[f"Q{fila}"] = (
            f'=IF(AND(ABS(G{fila}-O{fila})<{EPSILON},ABS(H{fila}-P{fila})<{EPSILON}),'
            f'"sí","REVISAR")'
        )
    hoja[f"I{fila}"].number_format = SOLES
    for letra in "JN":
        hoja[f"{letra}{fila}"].number_format = PORCENTAJE
    _marcar_total(hoja, fila, len(titulos))
    return primera, ultima, fila


def _soles(valor) -> str:
    """Igual que `formatSoles` del frontend (es-PE, sin decimales)."""
    return "—" if valor is None else f"S/ {valor:,.0f}"


def _pct(valor) -> str:
    """Igual que `formatPct` del frontend (es-PE, un decimal como máximo)."""
    return "—" if valor is None else f"{valor * 100:.1f}%".replace(".0%", "%")


def cifras_de_pantalla(agregados, total, primera, ultima) -> list[dict]:
    """Las ocho cifras del encabezado, cada una con su fórmula viva sobre «Municipalidades»."""
    M = "Municipalidades!"
    return [
        {"rotulo": "PIM del PP 0068 · {anio}", "texto": _soles(agregados.get("pim")),
         "api": agregados.get("pim"), "formula": f"={M}G{total}", "formato": SOLES,
         "como": "Suma del PIM de las filas del programa 0068 de las municipalidades de la "
                 "provincia.",
         "origen": f"{M}G{total} ← SUMIF sobre Detalle_0068!L{primera}:L{ultima}"},
        {"rotulo": "Devengado", "texto": _soles(agregados.get("devengado")),
         "api": agregados.get("devengado"), "formula": f"={M}H{total}", "formato": SOLES,
         "como": "Suma del devengado de las mismas filas. Es gasto ejecutado al corte, no del "
                 "año completo.",
         "origen": f"{M}H{total} ← SUMIF sobre Detalle_0068!M{primera}:M{ultima}"},
        {"rotulo": "% de ejecución", "texto": _pct(agregados.get("pct_ejecucion")),
         "api": agregados.get("pct_ejecucion"), "formula": f"={M}H{total}/{M}G{total}",
         "formato": PORCENTAJE,
         "como": "Devengado ÷ PIM. El numerador llega a junio y el denominador es de todo el "
                 "año: no es media ejecución perdida.",
         "origen": f"{M}H{total} ÷ {M}G{total}"},
        {"rotulo": "Saldo por ejecutar", "texto": _soles(agregados.get("saldo")),
         "api": agregados.get("saldo"), "formula": f"={M}G{total}-{M}H{total}",
         "formato": SOLES,
         "como": "PIM − devengado. Es una resta al vuelo, no un campo guardado.",
         "origen": f"{M}G{total} − {M}H{total}"},
        {"rotulo": "Presupuesto institucional (PIM)",
         "texto": _soles(agregados.get("pim_institucional")),
         "api": agregados.get("pim_institucional"), "formula": f"={M}L{total}",
         "formato": SOLES,
         "como": "Suma del PIM de las filas «PRESUPUESTO INSTITUCIONAL» de las mismas 8 "
                 "municipalidades: su presupuesto entero, todos los programas.",
         "origen": f"{M}L{total} ← Institucional, una fila por municipalidad"},
        {"rotulo": "PIA (del rótulo institucional)",
         "texto": _soles(agregados.get("pia_institucional")),
         "api": agregados.get("pia_institucional"), "formula": f"={M}K{total}",
         "formato": SOLES,
         "como": "PIA INSTITUCIONAL, no el del 0068. El PIA del programa es otro número, mucho "
                 "menor, y está en la columna F de «Municipalidades».",
         "origen": f"{M}K{total}"},
        {"rotulo": "«el PP 0068 es el …»",
         "texto": _pct(agregados.get("pct_0068_institucional")),
         "api": agregados.get("pct_0068_institucional"),
         "formula": f"={M}G{total}/{M}L{total}", "formato": PORCENTAJE,
         "como": "PIM del 0068 ÷ PIM institucional, ambos de las municipalidades que tienen "
                 "denominador.",
         "origen": f"{M}G{total} ÷ {M}L{total}"},
        {"rotulo": "«sobre N municipalidad(es)»",
         "texto": str(agregados.get("entidades_con_institucional")),
         "api": agregados.get("entidades_con_institucional"),
         "formula": f"=COUNT({M}L{primera}:L{ultima})", "formato": "0",
         "como": "Cuántas municipalidades tienen fila de presupuesto institucional. Es el "
                 "rótulo del porcentaje de arriba: dice sobre cuántas se calculó.",
         "origen": f"{M}L{primera}:L{ultima}, celdas con número"},
    ]


def hoja_resumen(libro, cifras, contexto) -> None:
    hoja = libro.create_sheet("Resumen", 0)
    inicio = _titulo(
        hoja,
        f"Las cifras de /inversion — {contexto['provincia']}, ejercicio {contexto['anio']}",
        f"Cada fila es un número que se lee en pantalla. La columna «Recalculado desde el "
        f"Excel» NO es un valor pegado: es una fórmula que sube desde las filas del archivo "
        f"del cliente. Fuente: {contexto['archivo']}, hoja «{contexto['hoja']}», corte "
        f"{contexto['corte']}. Libro generado el {contexto['generado']}.",
    )
    titulos = ["#", "Rótulo en pantalla", "Tal como se lee", "Valor del API (sin redondear)",
               "Recalculado desde el Excel", "Diferencia", "¿Coincide?", "Cómo se calcula",
               "De qué celdas sale"]
    _cabecera(hoja, inicio, titulos, [4, 30, 18, 22, 22, 13, 12, 56, 44])
    hoja.freeze_panes = f"A{inicio + 1}"

    fila = inicio + 1
    for n, cifra in enumerate(cifras, start=1):
        hoja.append([n, cifra["rotulo"].format(anio=contexto["anio"]), cifra["texto"],
                     cifra["api"]])
        hoja[f"E{fila}"] = cifra["formula"]
        if cifra["api"] is None:
            hoja[f"F{fila}"] = hoja[f"G{fila}"] = "no consultado"
        else:
            hoja[f"F{fila}"] = f"=E{fila}-D{fila}"
            hoja[f"G{fila}"] = f'=IF(ABS(F{fila})<{EPSILON},"sí","REVISAR")'
        hoja[f"H{fila}"] = cifra["como"]
        hoja[f"I{fila}"] = cifra["origen"]
        for letra in "DE":
            hoja[f"{letra}{fila}"].number_format = cifra["formato"]
        hoja[f"F{fila}"].number_format = cifra["formato"]
        hoja[f"G{fila}"].font = Font(bold=True)
        for i in range(1, len(titulos) + 1):
            hoja.cell(row=fila, column=i).border = BORDE
            hoja.cell(row=fila, column=i).alignment = Alignment(vertical="top", wrap_text=True)
        hoja.row_dimensions[fila].height = 44
        fila += 1


def hoja_conciliacion(libro, cifras, excel, base, agregados, conteos, contexto) -> None:
    """Las tres capas, una al lado de otra. Es la hoja que enseña que el código no altera nada."""
    hoja = libro.create_sheet("Conciliacion")
    inicio = _titulo(
        hoja,
        "Conciliación de las tres capas",
        "Columna 1: el Excel del cliente, releído y vuelto a sumar por este script sin pasar "
        "por la plataforma. Columna 2: lo que guarda la base de datos, servido por "
        "/api/inversion/entidades/ (filas de PresupuestoEntidad, sin agregar). Columna 3: lo "
        "que se pinta, servido por /api/inversion/ (bloque «agregados»). Si una cifra se "
        "rompiera, estas tres columnas dirían en cuál de los tres pasos.",
    )
    titulos = ["Cifra", "Excel del cliente", "Base de datos", "API / pantalla",
               "Δ Excel − base", "Δ base − API", "Veredicto", "Nota"]
    _cabecera(hoja, inicio, titulos, [42, 20, 20, 20, 15, 15, 14, 62])
    hoja.freeze_panes = f"B{inicio + 1}"

    fila = inicio + 1
    for clave, etiqueta, formato in (
        ("pia", "PIA del PP 0068", SOLES),
        ("pim", "PIM del PP 0068", SOLES),
        ("devengado", "Devengado del PP 0068", SOLES),
        ("pia_institucional", "PIA institucional", SOLES),
        ("pim_institucional", "PIM institucional", SOLES),
        ("devengado_institucional", "Devengado institucional", SOLES),
    ):
        hoja.append([etiqueta, excel.get(clave), base.get(clave) if base else None,
                     agregados.get(clave) if agregados else None])
        if base is None:
            hoja[f"C{fila}"] = hoja[f"D{fila}"] = "no consultado"
            hoja[f"E{fila}"] = hoja[f"F{fila}"] = hoja[f"G{fila}"] = "no consultado"
        else:
            hoja[f"E{fila}"] = f"=B{fila}-C{fila}"
            hoja[f"F{fila}"] = f"=C{fila}-D{fila}"
            hoja[f"G{fila}"] = (
                f'=IF(AND(ABS(E{fila})<{EPSILON},ABS(F{fila})<{EPSILON}),"cuadra","REVISAR")'
            )
        for letra in "BCDEF":
            hoja[f"{letra}{fila}"].number_format = formato
        hoja[f"G{fila}"].font = Font(bold=True)
        for i in range(1, len(titulos) + 1):
            hoja.cell(row=fila, column=i).border = BORDE
        fila += 1

    fila += 1
    hoja[f"A{fila}"] = "Recuentos"
    hoja[f"A{fila}"].font = Font(bold=True, color=VERDE)
    fila += 1
    for etiqueta, en_excel, en_base, nota in conteos:
        hoja.append([etiqueta, en_excel, en_base, "", "", "",
                     "cuadra" if en_excel == en_base else "REVISAR"])
        hoja[f"H{fila}"] = nota
        hoja[f"G{fila}"].font = Font(bold=True)
        for i in range(1, len(titulos) + 1):
            hoja.cell(row=fila, column=i).border = BORDE
        fila += 1


def hoja_universo(libro, institucional, programa, resumen, api, contexto) -> None:
    """Por qué son 8 municipalidades y no otro número."""
    hoja = libro.create_sheet("Universo")
    inicio = _titulo(
        hoja,
        "Por qué son ocho municipalidades",
        "El archivo cubre el departamento entero. Lo que la pantalla suma es un subconjunto, y "
        "cada paso del recorte está aquí con su motivo. El API llama a este recorte «ámbito "
        "municipal»: entidades de ámbito distrital o provincial, filtradas por el ubigeo de "
        "provincia (apps/inversion/consultas.py, AMBITOS y entidades()).",
    )
    titulos = ["Paso", "Cuántos", "Qué se queda fuera y por qué"]
    _cabecera(hoja, inicio, titulos, [46, 11, 84])

    entidades_ambito = None
    if api:
        entidades_ambito = api["tablero"]["agregados"].get("entidades_en_ambito")

    pasos = [
        ("Pliegos distintos en la hoja del archivo", resumen["pliegos_archivo"],
         "Todas las entidades del departamento con alguna fila, del programa o institucional."),
        ("… de ellos, el Gobierno Regional", 1,
         "«446: GOBIERNO REGIONAL DEL DEPARTAMENTO DE CUSCO» no es una municipalidad y no "
         "lleva provincia ni distrito. Concentra buena parte del presupuesto del "
         "departamento; mezclarlo compararía un pliego con una municipalidad distrital."),
        ("Municipalidades del departamento", resumen["pliegos_archivo"] - 1,
         "13 provinciales y el resto distritales."),
        (f"… con ubigeo de la provincia {contexto['ubigeo']}", resumen["pliegos_provincia"],
         "El filtro es el ubigeo que va dentro del campo «Pliego», el mismo dato con el que el "
         "backend resuelve el territorio."),
        ("Municipalidades que la pantalla suma", len(institucional),
         "Una provincial y el resto distritales. El ámbito «municipal» del API excluye además "
         "mancomunidades, gobierno regional y nacional; en esta provincia no hay ninguna."),
        ("«entidades_en_ambito» que devuelve el API",
         entidades_ambito if entidades_ambito is not None else "no consultado",
         "Tiene que coincidir con la fila anterior. Si no coincidiera, el recorte del archivo "
         "y el de la plataforma no serían el mismo."),
    ]
    fila = inicio + 1
    for etiqueta, cuantos, motivo in pasos:
        hoja.append([etiqueta, cuantos, motivo])
        for i in range(1, 4):
            hoja.cell(row=fila, column=i).border = BORDE
            hoja.cell(row=fila, column=i).alignment = Alignment(vertical="top", wrap_text=True)
        hoja.row_dimensions[fila].height = 34
        fila += 1

    fila += 1
    hoja[f"A{fila}"] = "Trampa de doble conteo"
    hoja[f"A{fila}"].font = Font(bold=True, color=VERDE)
    fila += 1
    suma_todo = sum(r["pim"] for r in programa) + sum(r["pim"] for r in institucional)
    hoja[f"A{fila}"] = (
        f"Las {len(programa) + len(institucional)} filas de la provincia suman "
        f"{suma_todo:,.0f} de PIM. Ese número NO existe: la fila "
        f"«{CATEGORIA_INSTITUCIONAL}» ya incluye al 0068, así que sumarlas juntas cuenta el "
        f"programa dos veces. Se suman por separado, y por eso hay dos hojas de detalle."
    )
    hoja[f"A{fila}"].alignment = Alignment(vertical="top", wrap_text=True)
    hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
    hoja.row_dimensions[fila].height = 46


def hoja_salvedades(libro, salvedades, advertencias) -> None:
    hoja = libro.create_sheet("Salvedades")
    inicio = _titulo(
        hoja,
        "Salvedades — leer antes de citar una cifra",
        "Ninguna de estas hace falsa una cifra del libro. Son los sitios donde un número "
        "correcto se puede citar mal, y los defectos del archivo fuente que un auditor va a "
        "encontrar por su cuenta.",
    )
    titulos = ["#", "Salvedad", "Por qué importa", "Dónde se comprueba"]
    _cabecera(hoja, inicio, titulos, [4, 52, 62, 46])

    fila = inicio + 1
    for n, (salvedad, porque, donde) in enumerate(salvedades, start=1):
        hoja.append([n, salvedad, porque, donde])
        for i in range(1, 5):
            hoja.cell(row=fila, column=i).border = BORDE
            hoja.cell(row=fila, column=i).alignment = Alignment(vertical="top", wrap_text=True)
        hoja.row_dimensions[fila].height = 50
        fila += 1

    fila += 1
    hoja[f"A{fila}"] = "Avisos generados al construir este libro"
    hoja[f"A{fila}"].font = Font(bold=True, color=VERDE)
    fila += 1
    if not advertencias:
        hoja[f"B{fila}"] = ("Ninguno. Las filas auditadas cumplen 0 ≤ devengado ≤ PIM, todas "
                            "llevan la categoría del programa 0068 y cada municipalidad trae "
                            "exactamente una fila de presupuesto institucional.")
        hoja[f"B{fila}"].alignment = Alignment(vertical="top", wrap_text=True)
        hoja.row_dimensions[fila].height = 44
        return
    for aviso in advertencias:
        hoja[f"B{fila}"] = aviso
        hoja[f"B{fila}"].alignment = Alignment(vertical="top", wrap_text=True)
        fila += 1


def salvedades_de(contexto, excel, agregados) -> list[tuple[str, str, str]]:
    pia_programa = _soles(excel["pia"])
    return [
        ("El «PIA» del cuarto KPI es el institucional, no el del programa.",
         f"El PIA del PP 0068 en esta provincia es {pia_programa}, no "
         f"{_soles(excel['pia_institucional'])}. La etiqueta es correcta dentro del KPI "
         f"«Presupuesto institucional», pero fuera de ese contexto se lee como si fuera el PIA "
         f"del programa.",
         "Municipalidades, columna F (PIA 0068) frente a la K (PIA institucional)."),
        (f"El % de ejecución es devengado al corte {contexto['corte']} contra un PIM anual.",
         "No es media ejecución perdida: el numerador cubre medio año y el denominador el año "
         "entero. La página lo advierte encima de los KPI; el KPI del porcentaje no lo repite. "
         "No se puede comparar con el de un ejercicio ya terminado.",
         "Detalle_0068, columna O, y el aviso de la propia pantalla."),
        ("El numerador del porcentaje sobre el institucional está acotado a las "
         "municipalidades que tienen denominador.",
         "El backend suma el PIM del 0068 solo de las entidades con presupuesto institucional, "
         "para que numerador y denominador salgan del mismo universo. Aquí las 8 lo tienen y "
         "coincide con el PIM de cabecera; en una provincia con alguna municipalidad sin "
         "denominador, los dos números serían distintos.",
         "apps/inversion/consultas.py, agregados(): filtro pim_institucional__isnull=False."),
        ("El denominador son estas municipalidades, no el territorio.",
         "No incluye al Gobierno Regional ni a los ministerios que también ejecutan en la "
         "provincia. El porcentaje responde a «cuánto del presupuesto de estas municipalidades "
         "va al 0068», no a «cuánto del dinero público que se gasta aquí».",
         "Universo, los pasos del recorte."),
        ("La municipalidad provincial gestiona presupuesto de toda su provincia.",
         "Su PIM no es del distrito capital. Es la razón de que la unidad de la ventana sea la "
         "municipalidad y no el distrito, y de que el mapa nunca reparta el presupuesto "
         "provincial entre los distritos de su provincia.",
         "Municipalidades, columna E (ámbito) — ADR-D4 y ADR-D6."),
        ("La fila «PRESUPUESTO INSTITUCIONAL» incluye al 0068.",
         "Es el total de la entidad, todos los programas, no una categoría más. Sumar las dos "
         "hojas de detalle juntas cuenta el programa dos veces.",
         "Universo, «Trampa de doble conteo»."),
        ("El importador no comprueba que la categoría sea el programa 0068.",
         "Manda al bloque del programa toda fila que no sea «PRESUPUESTO INSTITUCIONAL». Hoy "
         "el archivo solo trae esos dos valores —este script lo verifica fila a fila y lo "
         "avisaría—, así que la cifra publicada es correcta; un archivo futuro con otro "
         "programa entraría dentro del 0068 sin que nada lo dijera.",
         "backend/apps/datasets/importers/inversion.py, _leer_excel()."),
        ("El archivo fuente tiene defectos conocidos, todos fuera de esta provincia.",
         "Una municipalidad de Paruro trae dos filas de presupuesto institucional y otra "
         "ninguna; hay un PIA escrito como texto («600,00») en Chumbivilcas; hay tabuladores "
         "dentro de los nombres de actividad y la cabecera «% Ejecución» termina en espacio "
         "duro. Ninguno toca las filas auditadas aquí, pero un auditor los encontrará.",
         "Salvedades, «Avisos generados al construir este libro», y "
         "_specs/00-alcance-decisiones.md."),
        ("Ninguna fila auditada incumple 0 ≤ devengado ≤ PIM.",
         "Es la regla del SIAF que valida el importador antes de escribir nada. Se declara "
         "como resultado comprobado sobre estas filas, no como supuesto: un devengado por "
         "encima del PIM daría una ejecución superior al 100 % que se ve plausible en una "
         "tabla.",
         "Detalle_0068, columnas L y M — comprobado fila a fila al generar el libro."),
    ]


def analizar_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Libro de auditoría de las cifras de /inversion para una provincia."
    )
    parser.add_argument("excel", help="Excel del cliente con la base del ejercicio.")
    parser.add_argument("salida", help="Libro .xlsx a generar.")
    parser.add_argument(
        "--ubigeo", default="0801", help="Ubigeo de la provincia (por defecto: 0801, Cusco)."
    )
    parser.add_argument(
        "--provincia", default="provincia de Cusco", help="Nombre para los títulos del libro."
    )
    parser.add_argument(
        "--api",
        default="",
        help="Base del backend, p. ej. http://localhost:8000. Sin ella el libro se genera sin "
             "las columnas de la plataforma y sin conciliación.",
    )
    return parser.parse_args()


def main() -> None:
    args = analizar_argumentos()
    advertencias: list[str] = []

    programa, institucional, corte, resumen = leer_excel(args.excel, args.ubigeo)
    if not programa:
        raise SystemExit(
            f"El archivo no tiene ninguna fila del programa {PROGRAMA} con ubigeo de provincia "
            f"{args.ubigeo}. ¿Es el ubigeo correcto?"
        )
    revisar_archivo(programa, institucional, advertencias)

    anio = int(corte.split("-")[0])
    excel = {
        "pia": sum(r["pia"] for r in programa),
        "pim": sum(r["pim"] for r in programa),
        "devengado": sum(r["devengado"] for r in programa),
        "pia_institucional": sum(r["pia"] for r in institucional),
        "pim_institucional": sum(r["pim"] for r in institucional),
        "devengado_institucional": sum(r["devengado"] for r in institucional),
    }
    print(
        f"Excel: hoja «{resumen['hoja']}», corte {corte}. De la provincia {args.ubigeo}: "
        f"{len(programa)} filas del {PROGRAMA} y {len(institucional)} de "
        f"{CATEGORIA_INSTITUCIONAL.lower()}, en {resumen['pliegos_provincia']} pliegos.",
        file=sys.stderr,
    )

    api = consultar_api(args.api, args.ubigeo, anio, advertencias) if args.api else None
    if not args.api:
        advertencias.append(
            "No se pasó --api: el libro sale sin las columnas de la base de datos ni de la "
            "pantalla. Las cifras recalculadas desde el Excel siguen siendo válidas por sí "
            "solas, pero no hay conciliación."
        )

    agregados = api["tablero"]["agregados"] if api else {}
    base = None
    if api:
        base = {
            clave: sum(e[clave] or 0 for e in api["entidades"].values())
            for clave in ("pia", "pim", "devengado", "pia_institucional", "pim_institucional",
                          "devengado_institucional")
        }

    contexto = {
        "provincia": args.provincia, "ubigeo": args.ubigeo, "anio": anio, "corte": corte,
        "archivo": args.excel, "hoja": resumen["hoja"],
        "generado": date.today().strftime("%d/%m/%Y"),
    }

    libro = openpyxl.Workbook()
    libro.remove(libro.active)
    rango_detalle = hoja_detalle(libro, programa, api, contexto)
    primera, ultima, total = hoja_municipalidades(
        libro, institucional, api, rango_detalle, contexto
    )
    hoja_institucional(libro, institucional, contexto)
    cifras = cifras_de_pantalla(agregados, total, primera, ultima)
    hoja_resumen(libro, cifras, contexto)

    conteos = [
        (f"Filas del programa {PROGRAMA} en el archivo", len(programa),
         len(api["actividades"]) if api else "no consultado",
         "Registros de PresupuestoActividad de estas municipalidades. En esta provincia la "
         "correspondencia es 1:1; si una municipalidad repitiera la misma actividad bajo dos "
         "productos, el importador las sumaría y la base tendría menos filas."),
        ("Municipalidades con presupuesto institucional", len(institucional),
         agregados.get("entidades_con_institucional", "no consultado"),
         "Es el «sobre N municipalidad(es)» del cuarto KPI."),
    ]
    hoja_conciliacion(libro, cifras, excel, base, agregados, conteos, contexto)
    hoja_universo(libro, institucional, programa, resumen, api, contexto)
    hoja_salvedades(libro, salvedades_de(contexto, excel, agregados), advertencias)

    libro.save(args.salida)

    for aviso in advertencias:
        print(f"  (aviso) {aviso}", file=sys.stderr)

    if api is None:
        print(f"\n{args.salida}: generado SIN conciliar (no se consultó la plataforma).",
              file=sys.stderr)
        raise SystemExit(0)

    descuadres = []
    for clave in ("pia", "pim", "devengado", "pia_institucional", "pim_institucional",
                  "devengado_institucional"):
        for etiqueta, otro in (("base de datos", base[clave]), ("API", agregados.get(clave))):
            if otro is None or abs(excel[clave] - float(otro)) >= EPSILON:
                descuadres.append(
                    f"{clave}: el Excel dice {excel[clave]:,.2f} y la {etiqueta} {otro}."
                )
    for etiqueta, en_excel, en_plataforma, _ in conteos:
        if en_excel != en_plataforma:
            descuadres.append(f"{etiqueta}: {en_excel} en el Excel, {en_plataforma} en la base.")

    print(f"\n{args.salida}: generado.", file=sys.stderr)
    if descuadres:
        print("\nNO CUADRA — el libro está escrito, pero estas cifras discrepan:",
              file=sys.stderr)
        for descuadre in descuadres:
            print(f"  {descuadre}", file=sys.stderr)
        raise SystemExit(1)
    print(f"{len(cifras)}/{len(cifras)} cifras conciliadas entre el Excel, la base de datos y "
          f"la pantalla.", file=sys.stderr)


if __name__ == "__main__":
    main()
