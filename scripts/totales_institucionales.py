#!/usr/bin/env python3
"""Extrae el presupuesto institucional por entidad y ejercicio, para poner el 0068 en contexto.

`consolidar_pp0068.py` produce la serie del programa. Sola no responde a «¿cuánto pesa la
gestión del riesgo dentro de esta municipalidad?»: para eso hace falta el denominador, que es
el presupuesto de la entidad **entera**, todos los programas. Este script lo saca de las mismas
dos fuentes y con las mismas llaves, así que las dos salidas se cruzan por `(ENTIDAD_TIPO,
ENTIDAD_CODIGO, EJERCICIO)` sin conversiones.

De 2022 a 2025 se suma el comparativo del MEF sin filtrar por programa. De 2026 se leen las
filas `PRESUPUESTO INSTITUCIONAL` del Excel del cliente, que son exactamente ese total y que
`consolidar_pp0068.py` descarta a propósito (sumarían doble con el desglose del programa).

**Solo se emiten entidades que ejecutan en el departamento.** El recorte de `get_data_cusco.py`
se queda con las filas donde *cualquiera* de las dos geografías es Cusco, así que de un
ministerio de Lima solo están las filas cuya meta apunta aquí: su total institucional saldría
truncado y compararlo con el 0068 daría un porcentaje inventado. Los gobiernos locales y el
regional sí ejecutan desde Cusco y su total está completo.

Ejemplo:

    python3 scripts/totales_institucionales.py \\
        data/inversion/comparativo_cusco_gastos_2022_2025.csv \\
        data/inversion/Base_Prespuesto_PP0068_cusco_final.xlsx \\
        data/inversion/pp0068_cusco_institucional_2022_2026.csv
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict

import openpyxl

from consolidar_pp0068 import (
    METRICAS,
    PERIODO,
    a_numero,
    entidad_de_fila,
    entidad_de_pliego,
)

SALIDA = (
    "EJERCICIO",
    "CORTE",
    "FUENTE",
    "NIVEL_GOBIERNO",
    "ENTIDAD_TIPO",
    "ENTIDAD_CODIGO",
    "ENTIDAD_NOMBRE",
    "PIA",
    "PIM",
    "DEVENGADO",
)

#: Etiqueta con la que el Excel del cliente marca el total de la entidad.
CATEGORIA_INSTITUCIONAL = "PRESUPUESTO INSTITUCIONAL"


def analizar_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Presupuesto institucional por entidad y ejercicio (denominador del 0068)."
    )
    parser.add_argument("mef", help="CSV recortado del MEF, sin las columnas del ejercicio falso.")
    parser.add_argument("excel", help="Excel con la base del ejercicio nuevo.")
    parser.add_argument("salida", help="CSV a generar.")
    parser.add_argument(
        "--departamento",
        default="CUSCO",
        help="Departamento de la unidad ejecutora que se conserva (por defecto: CUSCO).",
    )
    parser.add_argument("--hoja", default="Base 2026", help="Hoja del Excel con los datos.")
    return parser.parse_args()


def leer_mef(ruta: str, departamento: str, ejercicios: list[str]):
    """Suma todas las filas de cada entidad, sin filtrar por programa."""
    acumulado: dict[tuple[str, str], dict[str, list[float]]] = defaultdict(
        lambda: {anio: [0.0, 0.0, 0.0] for anio in ejercicios}
    )
    niveles: dict[tuple[str, str], str] = {}
    nombres: dict[tuple[str, str], str] = {}
    filas = descartadas = 0

    with open(ruta, newline="", encoding="utf-8-sig") as f:
        for fila in csv.DictReader(f):
            if fila["DEPARTAMENTO_EJECUTORA_NOMBRE"].strip().upper() != departamento:
                descartadas += 1
                continue
            filas += 1
            entidad = entidad_de_fila(fila)
            for anio in ejercicios:
                totales = acumulado[entidad][anio]
                for i, metrica in enumerate(METRICAS):
                    totales[i] += float(fila[f"{metrica}_{anio}"])
            niveles.setdefault(entidad, fila["NIVEL_GOBIERNO"])
            nombres.setdefault(
                entidad,
                (
                    fila["EJECUTORA_NOMBRE"]
                    if entidad[0] == "SEC_EJEC"
                    else fila["PLIEGO_NOMBRE"]
                ).strip(),
            )

    print(
        f"MEF: {filas:,} filas de entidades que ejecutan en {departamento} "
        f"({descartadas:,} descartadas por ejecutar desde fuera) en {len(acumulado):,} entidades.",
        file=sys.stderr,
    )
    return acumulado, niveles, nombres


def leer_excel(ruta: str, hoja: str):
    """Devuelve (filas institucionales, entidades de la hoja, ejercicio, corte)."""
    libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    if hoja not in libro.sheetnames:
        raise SystemExit(f"El Excel no tiene la hoja {hoja!r}. Hojas: {', '.join(libro.sheetnames)}")

    filas = list(libro[hoja].iter_rows(values_only=True))
    cabecera = [str(c).strip() if c is not None else "" for c in filas[0]]
    indice = {nombre: i for i, nombre in enumerate(cabecera)}
    faltan = [
        c
        for c in ("Periodo", "Pliego", "Categoría_Presupuestal", "PIA", "PIM", "Devengado")
        if c not in indice
    ]
    if faltan:
        raise SystemExit("Al Excel le faltan columnas: " + ", ".join(faltan))

    utiles = [f for f in filas[1:] if f[indice["Periodo"]] is not None]
    institucionales = [
        f
        for f in utiles
        if str(f[indice["Categoría_Presupuestal"]]).strip() == CATEGORIA_INSTITUCIONAL
    ]
    if not institucionales:
        raise SystemExit(
            f"El Excel no trae ninguna fila con Categoría_Presupuestal = "
            f"{CATEGORIA_INSTITUCIONAL!r}; sin ellas no hay denominador que extraer."
        )

    periodos = {str(f[indice["Periodo"]]).strip() for f in institucionales}
    if len(periodos) != 1:
        raise SystemExit(
            f"Se esperaba un solo Periodo en el Excel; hay {len(periodos)}: {sorted(periodos)}"
        )
    corte = periodos.pop()
    coincide = PERIODO.match(corte)
    if not coincide:
        raise SystemExit(f"Periodo {corte!r}: se esperaba el formato AAAA-MM.")

    entidades = {entidad_de_pliego(f[indice["Pliego"]])[0] for f in utiles}

    print(
        f"Excel: {len(institucionales):,} filas de {CATEGORIA_INSTITUCIONAL.lower()} para "
        f"{len(entidades):,} entidades, corte {corte}.",
        file=sys.stderr,
    )
    return (
        [{c: f[indice[c]] for c in indice if c} for f in institucionales],
        entidades,
        coincide.group(1),
        corte,
    )


def main() -> None:
    args = analizar_argumentos()
    departamento = args.departamento.strip().upper()

    excel, entidades_hoja, ejercicio_nuevo, corte = leer_excel(args.excel, args.hoja)

    with open(args.mef, newline="", encoding="utf-8-sig") as f:
        columnas = next(csv.reader(f))
    ejercicios = sorted({c.rsplit("_", 1)[1] for c in columnas if c.startswith("PIM_")})
    if ejercicio_nuevo in ejercicios:
        raise SystemExit(
            f"El CSV del MEF todavía trae columnas de {ejercicio_nuevo}, que es el ejercicio del "
            f"Excel. Regenéralo con --hasta-ejercicio {int(ejercicio_nuevo) - 1}."
        )

    acumulado, niveles, nombres = leer_mef(args.mef, departamento, ejercicios)

    salida = []
    vacias = 0
    for entidad, por_anio in acumulado.items():
        for anio, totales in por_anio.items():
            if not any(totales):
                vacias += 1
                continue
            salida.append({
                "EJERCICIO": anio,
                "CORTE": "anual",
                "FUENTE": "MEF",
                "NIVEL_GOBIERNO": niveles.get(entidad, ""),
                "ENTIDAD_TIPO": entidad[0],
                "ENTIDAD_CODIGO": entidad[1],
                "ENTIDAD_NOMBRE": nombres.get(entidad, ""),
                "PIA": totales[0],
                "PIM": totales[1],
                "DEVENGADO": totales[2],
            })

    # El Excel debería traer una fila institucional por entidad. Cuando trae dos, el total es
    # ambiguo y **no se elige uno**: la base de 2026 tiene una entidad con dos filas y otra sin
    # ninguna, así que la explicación probable es una etiqueta corrida, pero repararla por
    # posición sería adivinar sobre datos del cliente. Se descartan las dos y se avisa, que es
    # lo que deja el hueco visible aguas abajo en vez de publicar un porcentaje falso.
    por_entidad: dict[tuple[str, str], list[dict]] = defaultdict(list)
    nombres_excel: dict[tuple[str, str], str] = {}
    for fila in excel:
        entidad, nombre_excel = entidad_de_pliego(fila["Pliego"])
        por_entidad[entidad].append(fila)
        nombres_excel.setdefault(entidad, nombre_excel)

    ambiguas = sorted(e for e, filas in por_entidad.items() if len(filas) > 1)
    for entidad in ambiguas:
        print(
            f"  (aviso) {entidad[1]} ({nombres_excel[entidad]}) trae "
            f"{len(por_entidad[entidad])} filas de {CATEGORIA_INSTITUCIONAL.lower()} en el "
            f"Excel. Se descartan todas: el total sería ambiguo y con él el % del programa "
            f"sobre el institucional. Conviene pedir el archivo corregido.",
            file=sys.stderr,
        )

    sin_historia = 0
    for entidad, filas_entidad in por_entidad.items():
        if len(filas_entidad) > 1:
            continue
        fila = filas_entidad[0]
        nombre_excel = nombres_excel[entidad]
        sin_historia += entidad not in acumulado
        salida.append({
            "EJERCICIO": ejercicio_nuevo,
            "CORTE": corte,
            "FUENTE": "BASE_INSTITUCIONAL",
            "NIVEL_GOBIERNO": niveles.get(entidad, "M" if entidad[0] == "SEC_EJEC" else "R"),
            "ENTIDAD_TIPO": entidad[0],
            "ENTIDAD_CODIGO": entidad[1],
            "ENTIDAD_NOMBRE": nombres.get(entidad, nombre_excel),
            "PIA": a_numero(fila["PIA"], "PIA"),
            "PIM": a_numero(fila["PIM"], "PIM"),
            "DEVENGADO": a_numero(fila["Devengado"], "Devengado"),
        })

    salida.sort(key=lambda f: (f["ENTIDAD_TIPO"], f["ENTIDAD_CODIGO"], f["EJERCICIO"]))

    with open(args.salida, "w", newline="", encoding="utf-8-sig") as f:
        escritor = csv.DictWriter(f, fieldnames=SALIDA)
        escritor.writeheader()
        for fila in salida:
            for metrica in METRICAS:
                fila[metrica] = f"{fila[metrica]:.2f}"
            escritor.writerow(fila)

    # Entidades que están en la hoja pero se quedan sin total institucional del ejercicio
    # nuevo: las ambiguas de arriba y las que directamente no traen la fila. Son las que en
    # pantalla tendrán que mostrar el porcentaje como «sin dato».
    con_total = {e for e, filas_entidad in por_entidad.items() if len(filas_entidad) == 1}
    sin_denominador = sorted(entidades_hoja - con_total)

    por_anio: dict[str, int] = defaultdict(int)
    for fila in salida:
        por_anio[fila["EJERCICIO"]] += 1
    print(
        f"\n{args.salida}: {len(salida):,} filas.\n"
        + "  " + "  ".join(f"{anio}: {n:,}" for anio, n in sorted(por_anio.items()))
        + f"\nCortes sin importe descartados: {vacias:,}."
        f"\nEntidades de {ejercicio_nuevo} sin historial en el comparativo: {sin_historia}."
        f"\nEntidades sin total institucional de {ejercicio_nuevo}: {len(sin_denominador)}"
        + (
            " (" + ", ".join(f"{tipo} {codigo}" for tipo, codigo in sin_denominador) + ")"
            if sin_denominador
            else ""
        )
        + ".",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
