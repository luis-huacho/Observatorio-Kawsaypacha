#!/usr/bin/env python3
"""Consolida la serie de un programa presupuestal juntando el volcado del MEF y la base del cliente.

Las columnas *_2026 del volcado del MEF son una copia literal de las de 2025, así que el
comparativo solo sirve hasta 2025. El 2026 real llega por `Base_Prespuesto_PP0068_cusco_final.xlsx`,
que cubre únicamente el programa 0068 en Cusco.

Las dos fuentes no comparten forma: el MEF trae 53 dimensiones y 7 métricas por ejercicio; el
Excel, 9 dimensiones y 3 métricas. Solo se juntan bajando el MEF a la granularidad del Excel
—entidad · producto · actividad— y quedándose con las tres métricas comunes: PIA, PIM y
DEVENGADO. La salida es larga: una fila por combinación y ejercicio.

La entidad se identifica según cómo la nombra el Excel, que usa dos formatos:

    080201-300692: MUNICIPALIDAD PROVINCIAL DE ACOMAYO   -> SEC_EJEC 300692 (gobierno local)
    446: GOBIERNO REGIONAL DEL DEPARTAMENTO DE CUSCO     -> PLIEGO 446 (agrega sus ejecutoras)

En el MEF los gobiernos locales llevan el pliego vacío, así que para ellos la llave es SEC_EJEC
y para el resto, PLIEGO.

Ojo al leer los resultados: el corte del Excel es a mitad de año (`CORTE` lo deja escrito en
cada fila), y no todas las entidades del MEF están en la base 2026 —el gobierno nacional no
está—, por eso la columna EN_BASE_2026. La serie comparable es `EN_BASE_2026 = 1`.

Ejemplo:

    python3 scripts/consolidar_pp0068.py \\
        data/inversion/comparativo_cusco_gastos_2022_2025.csv \\
        data/inversion/Base_Prespuesto_PP0068_cusco_final.xlsx \\
        data/inversion/pp0068_cusco_2022_2026_largo.csv
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict

import openpyxl

SALIDA = (
    "EJERCICIO",
    "CORTE",
    "FUENTE",
    "NIVEL_GOBIERNO",
    "ENTIDAD_TIPO",
    "ENTIDAD_CODIGO",
    "ENTIDAD_NOMBRE",
    "PROVINCIA",
    "DISTRITO",
    "PRODUCTO_PROYECTO",
    "PRODUCTO_PROYECTO_NOMBRE",
    "ACTIVIDAD_ACCION_OBRA",
    "ACTIVIDAD_ACCION_OBRA_NOMBRE",
    "TIPO",
    "PIA",
    "PIM",
    "DEVENGADO",
    "EN_BASE_2026",
)

METRICAS = ("PIA", "PIM", "DEVENGADO")
PERIODO = re.compile(r"^(\d{4})-(\d{2})$")
PUNTO_DECIMAL = re.compile(r"^-?\d+(\.\d+)?$")
COMA_DECIMAL = re.compile(r"^-?\d+,\d+$")


def a_numero(valor, columna: str) -> float:
    """Convierte un importe del Excel, que casi siempre es entero pero a veces viene como texto.

    En la base de 2026 hay una sola celda de texto (`600,00`), con la coma como separador
    decimal y sin separador de miles. Se acepta ese formato y se rechaza cualquier otro en vez
    de adivinar: si un día llega `1.234`, no hay forma de saber si son mil doscientos treinta y
    cuatro o uno coma dos tres cuatro, y equivocarse ahí falsea el importe en silencio.
    """
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if PUNTO_DECIMAL.match(texto):
        return float(texto)
    if COMA_DECIMAL.match(texto):
        return float(texto.replace(",", "."))
    raise SystemExit(
        f"Columna {columna}: no se puede interpretar el importe {valor!r}. Se esperaba un "
        "número, o un texto tipo '600,00' (coma decimal, sin separador de miles)."
    )


def analizar_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Consolida el comparativo del MEF (2022–2025) con la base del cliente (2026)."
    )
    parser.add_argument("mef", help="CSV recortado del MEF, sin las columnas del 2026 falso.")
    parser.add_argument("excel", help="Excel con la base del ejercicio nuevo.")
    parser.add_argument("salida", help="CSV consolidado a generar.")
    parser.add_argument("--programa", default="0068", help="Categoría presupuestal (por defecto: 0068).")
    parser.add_argument("--hoja", default="Base 2026", help="Hoja del Excel con los datos.")
    return parser.parse_args()


def tipo_por_codigo(producto: str) -> str:
    """PROYECTO o ACTIVIDAD según el código de producto.

    En las 19 300 filas del 0068 la correspondencia es exacta: los códigos que empiezan en 2
    son proyectos y los que empiezan en 3, actividades. Derivarlo del código evita depender
    de cómo nombra cada fuente esa distinción (el MEF dice ACTIVIDAD; el Excel, Producto).
    """
    return {"2": "PROYECTO", "3": "ACTIVIDAD"}.get(producto[:1], "")


def entidad_de_fila(fila: dict[str, str]) -> tuple[str, str]:
    """Llave de entidad de una fila del MEF, al nivel en que la nombra el Excel."""
    if fila["NIVEL_GOBIERNO"] == "M":
        return ("SEC_EJEC", fila["SEC_EJEC"].strip())
    return ("PLIEGO", fila["PLIEGO"].strip())


def entidad_de_pliego(texto: str) -> tuple[tuple[str, str], str]:
    """Interpreta el campo Pliego del Excel: devuelve (llave, nombre)."""
    prefijo, _, nombre = str(texto).partition(":")
    prefijo, nombre = prefijo.strip(), nombre.strip()
    if "-" in prefijo:
        return ("SEC_EJEC", prefijo.split("-", 1)[1]), nombre
    return ("PLIEGO", prefijo), nombre


def codigo_y_nombre(texto) -> tuple[str, str]:
    """Parte `3000001: ACCIONES COMUNES` en código y nombre."""
    codigo, _, nombre = str(texto).partition(":")
    return codigo.strip(), nombre.strip()


def leer_mef(ruta: str, programa: str, ejercicios: list[str]):
    """Agrega el CSV del MEF a (entidad, producto, actividad) × ejercicio."""
    acumulado: dict[tuple, dict[str, list[float]]] = defaultdict(
        lambda: {anio: [0.0, 0.0, 0.0] for anio in ejercicios}
    )
    dimensiones: dict[tuple, dict[str, str]] = {}
    catalogo_producto: dict[str, str] = {}
    catalogo_actividad: dict[str, str] = {}
    niveles: dict[tuple[str, str], str] = {}
    nombres_entidad: dict[tuple[str, str], str] = {}
    filas = 0

    with open(ruta, newline="", encoding="utf-8-sig") as f:
        for fila in csv.DictReader(f):
            if fila["PROGRAMA_PPTO"] != programa:
                continue
            filas += 1
            entidad = entidad_de_fila(fila)
            producto, actividad = fila["PRODUCTO_PROYECTO"], fila["ACTIVIDAD_ACCION_OBRA"]
            llave = (*entidad, producto, actividad)

            for anio in ejercicios:
                totales = acumulado[llave][anio]
                for i, metrica in enumerate(METRICAS):
                    totales[i] += float(fila[f"{metrica}_{anio}"])

            if llave not in dimensiones:
                local = fila["NIVEL_GOBIERNO"] == "M"
                dimensiones[llave] = {
                    "NIVEL_GOBIERNO": fila["NIVEL_GOBIERNO"],
                    "PROVINCIA": fila["PROVINCIA_EJECUTORA_NOMBRE"].strip() if local else "",
                    "DISTRITO": fila["DISTRITO_EJECUTORA_NOMBRE"].strip() if local else "",
                }
            catalogo_producto.setdefault(producto, fila["PRODUCTO_PROYECTO_NOMBRE"].strip())
            catalogo_actividad.setdefault(actividad, fila["ACTIVIDAD_ACCION_OBRA_NOMBRE"].strip())
            niveles.setdefault(entidad, fila["NIVEL_GOBIERNO"])
            nombres_entidad.setdefault(
                entidad,
                (fila["EJECUTORA_NOMBRE"] if entidad[0] == "SEC_EJEC" else fila["PLIEGO_NOMBRE"]).strip(),
            )

    print(f"MEF: {filas:,} filas del programa {programa} en {len(acumulado):,} combinaciones.", file=sys.stderr)
    return acumulado, dimensiones, catalogo_producto, catalogo_actividad, niveles, nombres_entidad


def leer_excel(ruta: str, hoja: str, programa: str):
    """Devuelve (filas del programa, ejercicio, corte) de la base del cliente."""
    libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    if hoja not in libro.sheetnames:
        raise SystemExit(f"El Excel no tiene la hoja {hoja!r}. Hojas: {', '.join(libro.sheetnames)}")

    filas = list(libro[hoja].iter_rows(values_only=True))
    cabecera = [str(c).strip() if c is not None else "" for c in filas[0]]
    indice = {nombre: i for i, nombre in enumerate(cabecera)}
    faltan = [c for c in ("Periodo", "Pliego", "Categoría_Presupuestal", "Nombre_producto",
                          "Nombre_Actividad", "PIA", "PIM", "Devengado") if c not in indice]
    if faltan:
        raise SystemExit("Al Excel le faltan columnas: " + ", ".join(faltan))

    utiles = [f for f in filas[1:] if f[indice["Periodo"]] is not None]
    # Las filas PRESUPUESTO INSTITUCIONAL son el total del pliego: sumarían doble.
    del_programa = [
        f for f in utiles if str(f[indice["Categoría_Presupuestal"]]).startswith(programa)
    ]
    descartadas = len(utiles) - len(del_programa)

    periodos = {str(f[indice["Periodo"]]).strip() for f in del_programa}
    if len(periodos) != 1:
        raise SystemExit(f"Se esperaba un solo Periodo en el Excel; hay {len(periodos)}: {sorted(periodos)}")
    corte = periodos.pop()
    coincide = PERIODO.match(corte)
    if not coincide:
        raise SystemExit(f"Periodo {corte!r}: se esperaba el formato AAAA-MM.")

    print(
        f"Excel: {len(del_programa):,} filas del programa {programa}, corte {corte} "
        f"({descartadas} filas de total institucional descartadas).",
        file=sys.stderr,
    )
    return [{c: f[indice[c]] for c in indice if c} for f in del_programa], coincide.group(1), corte


def main() -> None:
    args = analizar_argumentos()

    excel, ejercicio_nuevo, corte = leer_excel(args.excel, args.hoja, args.programa)
    entidades_2026 = {entidad_de_pliego(f["Pliego"])[0] for f in excel}

    with open(args.mef, newline="", encoding="utf-8-sig") as f:
        columnas = next(csv.reader(f))
    ejercicios = sorted({c.rsplit("_", 1)[1] for c in columnas if c.startswith("PIM_")})
    if ejercicio_nuevo in ejercicios:
        raise SystemExit(
            f"El CSV del MEF todavía trae columnas de {ejercicio_nuevo}, que es el ejercicio del "
            f"Excel. Regenéralo con --hasta-ejercicio {int(ejercicio_nuevo) - 1}."
        )

    acumulado, dimensiones, cat_producto, cat_actividad, niveles, nombres = leer_mef(
        args.mef, args.programa, ejercicios
    )

    salida = []
    vacias = 0
    # Llaves que llegan a escribir alguna fila. No es lo mismo que estar en `acumulado`: el
    # comparativo del MEF arrastra combinaciones con las tres métricas en cero todos los años,
    # y esas no dejan rastro en la salida ni sirven para comparar contra 2026.
    con_historia: set[tuple] = set()
    for llave, por_anio in acumulado.items():
        entidad_tipo, entidad_codigo, producto, actividad = llave
        entidad = (entidad_tipo, entidad_codigo)
        for anio, totales in por_anio.items():
            if not any(totales):
                vacias += 1
                continue
            con_historia.add(llave)
            salida.append({
                "EJERCICIO": anio,
                "CORTE": "anual",
                "FUENTE": "MEF",
                "NIVEL_GOBIERNO": dimensiones[llave]["NIVEL_GOBIERNO"],
                "ENTIDAD_TIPO": entidad_tipo,
                "ENTIDAD_CODIGO": entidad_codigo,
                "ENTIDAD_NOMBRE": nombres.get(entidad, ""),
                "PROVINCIA": dimensiones[llave]["PROVINCIA"],
                "DISTRITO": dimensiones[llave]["DISTRITO"],
                "PRODUCTO_PROYECTO": producto,
                "PRODUCTO_PROYECTO_NOMBRE": cat_producto.get(producto, ""),
                "ACTIVIDAD_ACCION_OBRA": actividad,
                "ACTIVIDAD_ACCION_OBRA_NOMBRE": cat_actividad.get(actividad, ""),
                "TIPO": tipo_por_codigo(producto),
                "PIA": totales[0],
                "PIM": totales[1],
                "DEVENGADO": totales[2],
                "EN_BASE_2026": int(entidad in entidades_2026),
            })

    sin_historia = ausentes = 0
    for fila in excel:
        entidad, nombre_excel = entidad_de_pliego(fila["Pliego"])
        producto, nombre_producto = codigo_y_nombre(fila["Nombre_producto"])
        actividad, nombre_actividad = codigo_y_nombre(fila["Nombre_Actividad"])
        llave = (*entidad, producto, actividad)
        if llave not in con_historia:
            sin_historia += 1
            ausentes += llave not in acumulado
        salida.append({
            "EJERCICIO": ejercicio_nuevo,
            "CORTE": corte,
            "FUENTE": "BASE_" + ("PP" + args.programa),
            "NIVEL_GOBIERNO": niveles.get(entidad, "M" if entidad[0] == "SEC_EJEC" else "R"),
            "ENTIDAD_TIPO": entidad[0],
            "ENTIDAD_CODIGO": entidad[1],
            "ENTIDAD_NOMBRE": nombres.get(entidad, nombre_excel),
            "PROVINCIA": (fila["Provincia"] or "").strip(),
            "DISTRITO": (fila["Distrito"] or "").strip(),
            "PRODUCTO_PROYECTO": producto,
            # El catálogo del MEF manda, para que un código no salga con dos grafías.
            "PRODUCTO_PROYECTO_NOMBRE": cat_producto.get(producto, nombre_producto),
            "ACTIVIDAD_ACCION_OBRA": actividad,
            "ACTIVIDAD_ACCION_OBRA_NOMBRE": cat_actividad.get(actividad, nombre_actividad),
            "TIPO": tipo_por_codigo(producto),
            "PIA": a_numero(fila["PIA"], "PIA"),
            "PIM": a_numero(fila["PIM"], "PIM"),
            "DEVENGADO": a_numero(fila["Devengado"], "Devengado"),
            "EN_BASE_2026": 1,
        })

    salida.sort(key=lambda f: (f["ENTIDAD_TIPO"], f["ENTIDAD_CODIGO"], f["PRODUCTO_PROYECTO"],
                               f["ACTIVIDAD_ACCION_OBRA"], f["EJERCICIO"]))

    with open(args.salida, "w", newline="", encoding="utf-8-sig") as f:
        escritor = csv.DictWriter(f, fieldnames=SALIDA)
        escritor.writeheader()
        for fila in salida:
            for metrica in METRICAS:
                fila[metrica] = f"{fila[metrica]:.2f}"
            escritor.writerow(fila)

    por_anio: dict[str, int] = defaultdict(int)
    for fila in salida:
        por_anio[fila["EJERCICIO"]] += 1
    print(
        f"\n{args.salida}: {len(salida):,} filas.\n"
        + "  " + "  ".join(f"{anio}: {n:,}" for anio, n in sorted(por_anio.items()))
        + f"\nCortes sin importe descartados: {vacias:,}."
        f"\nFilas de {ejercicio_nuevo} sin importes en ejercicios previos: {sin_historia} "
        f"(de ellas {ausentes} ni siquiera figuran en el comparativo del MEF).",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
