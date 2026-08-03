#!/usr/bin/env python3
"""Convierte los Excel canónicos de PREDES a los JSON que consume el prototipo.

Entradas (data/layers/data/):
  Base_Nivel Peligro_CCPP_Cusco.xlsx      — 9 hojas, una por peligro, 8,968 CCPP
  Base_Frecuencia_Peligro_Cusco.xlsx      — hoja NºEMERGENCIAS, 111 distritos

Salidas (prototype/public/data/):
  ccpp.json        — centros poblados deduplicados
  peligros.json    — clasificaciones de exposición (formato largo)
  frecuencia.json  — emergencias históricas por distrito

Correr desde cualquier sitio:
  python3 prototype/scripts/xlsx_to_json.py
"""
from __future__ import annotations

from pathlib import Path
import json
import sys
import unicodedata

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data" / "layers" / "data"
XLSX_NIVEL = DATA / "Base_Nivel Peligro_CCPP_Cusco.xlsx"
XLSX_FREC = DATA / "Base_Frecuencia_Peligro_Cusco.xlsx"
OUT = ROOT / "prototype" / "public" / "data"

# El nombre de la hoja NO siempre coincide con el valor de la columna PELIGRO.
# La fuente de verdad es la columna; la hoja solo ordena el archivo.
#   (hoja, nombre canónico, slug, categoría geodinámica)
PELIGROS = [
    ("Sismo", "Sismo", "sismo", "Geodinámica interna"),
    ("Friaje", "Friaje", "friaje", "Meteorológico"),
    ("Inundación", "Inundación", "inundacion", "Meteorológico"),
    ("Heladas", "Heladas", "heladas", "Meteorológico"),
    ("Bajas temperaturas", "Bajas temperaturas", "bajas_temperaturas", "Meteorológico"),
    ("Lluvias", "Lluvias intensas", "lluvias_intensas", "Meteorológico"),
    ("Sequía", "Sequía", "sequia", "Meteorológico"),
    ("Incendios Forestales", "Incendios forestales", "incendios_forestales", "Meteorológico"),
    ("Movimientos en masa", "Movimientos en masa", "movimientos_en_masa", "Geodinámica externa"),
]

# Categorías del Excel de frecuencia: (nombre, slug, columna de subtotal, columnas de evento).
# Los subtotales TOT_* no se suman: se recalculan desde el desglose y solo se usan como
# "total declarado" cuando la fuente no desagrega (ver el caso del distrito de Cusco).
CATEGORIAS = [
    (
        "Geodinámica externa", "geodinamica_externa", "TOT_GEODINAMICA EXTERNA",
        ["HUAYCO", "DESLIZAMIENTO", "ALUVIÓN", "DERRUMBE", "REPTACIÓN", "FLUJO DE DETRITOS"],
    ),
    (
        "Geodinámica interna", "geodinamica_interna", "TOT_GEODINAMICA INTERNA",
        ["SISMO"],
    ),
    (
        "Meteorológicos / oceanográficos", "meteorologico", "TOT_METEREOLÓGICOS / OCEANOGRÁFICOS",
        ["HELADA", "BAJA TEMPERATURA", "VIENTOS FUERTES", "FRIAJE", "GRANIZADAS", "INUNDACIÓN",
         "LLUVIAS INTENSAS", "NEVADA", "SEQUÍA", "DÉFICIT HÍDRICO", "TORMENTA ELECTRICA"],
    ),
    (
        "Inducidos por acción humana", "inducido_humano", "TOT_INDUCIDOS POR LA ACCIÓN HUMANA",
        ["COLAPSO POR ANTIGÜEDAD", "INCENDIO FORESTAL", "INCENDIO"],
    ),
]

# La fuente escribe la misma institución de dos maneras.
FUENTE_CANONICA = {"CENEPRED_SIGRID": "SIGRID_CENEPRED"}

avisos: list[str] = []


def aviso(msg: str) -> None:
    avisos.append(msg)


def norm(valor) -> str:
    """Normaliza para comparar nombres de distrito: sin tildes, mayúsculas, sin espacios extra."""
    if valor is None:
        return ""
    s = unicodedata.normalize("NFKD", str(valor).strip().upper())
    return " ".join("".join(c for c in s if not unicodedata.combining(c)).split())


def texto(valor) -> str:
    return str(valor).strip() if valor is not None else ""


def entero(valor):
    """Los conteos vienen como int, pero la columna TOTAL viene como string."""
    if valor is None or valor == "":
        return None
    try:
        return int(str(valor).strip())
    except (TypeError, ValueError):
        return None


def slug_evento(nombre: str) -> str:
    s = unicodedata.normalize("NFKD", nombre.strip().lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return "_".join(s.split())


def titulo_evento(nombre: str) -> str:
    """HUAYCO -> Huayco; BAJA TEMPERATURA -> Baja temperatura."""
    return nombre.strip().capitalize()


def leer_niveles():
    """Devuelve (ccpp, clasificaciones, distritos_por_ubigeo)."""
    wb = openpyxl.load_workbook(XLSX_NIVEL, read_only=True, data_only=True)

    faltantes = [h for h, *_ in PELIGROS if h not in wb.sheetnames]
    if faltantes:
        sys.exit(f"ERROR: faltan hojas en {XLSX_NIVEL.name}: {faltantes}")

    ccpp: dict[str, dict] = {}
    clasificaciones: list[dict] = []
    distritos: dict[str, dict] = {}
    sin_nivel = 0
    huerfanas = 0

    for hoja, nombre_peligro, slug, categoria in PELIGROS:
        ws = wb[hoja]
        vistos: set[str] = set()

        for fila, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not r or all(c is None for c in r):
                continue
            if r[3] is None:
                huerfanas += 1
                aviso(f"{hoja}!{fila}: fila sin CODIGO de centro poblado, descartada")
                continue

            codigo = texto(r[3])
            depto, prov, dist = texto(r[0]), texto(r[1]), texto(r[2])
            nombre, cat_ccpp = texto(r[4]), texto(r[5])
            altitud, lon, lat, poblacion = r[6], r[7], r[8], r[9]
            peligro_col, nivel = texto(r[10]), r[12]
            fuente, link = texto(r[13]), texto(r[14])

            # Un mismo CCPP aparece en las 9 hojas; nos quedamos con la variante más completa
            # (una de las hojas deja el distrito de SICUANI en blanco).
            registro = ccpp.get(codigo)
            if registro is None:
                ccpp[codigo] = {
                    "codigo": codigo,
                    "nombre": nombre,
                    "categoria": cat_ccpp,
                    "departamento": depto,
                    "provincia": prov,
                    "distrito": dist,
                    "ubigeo_distrito": codigo[:6],
                    "lat": float(lat) if lat is not None else None,
                    "lon": float(lon) if lon is not None else None,
                    "altitud": int(altitud) if altitud is not None else None,
                    "poblacion": int(poblacion) if poblacion is not None else None,
                }
            else:
                for campo, valor in (
                    ("nombre", nombre), ("categoria", cat_ccpp),
                    ("departamento", depto), ("provincia", prov), ("distrito", dist),
                ):
                    if not registro[campo] and valor:
                        registro[campo] = valor
                        aviso(f"{hoja}!{fila}: CCPP {codigo} completó '{campo}' desde otra hoja")

            if dist and codigo[:6] not in distritos:
                distritos[codigo[:6]] = {
                    "ubigeo": codigo[:6], "distrito": dist, "provincia": prov,
                }

            if codigo in vistos:
                aviso(f"{hoja}!{fila}: CODIGO {codigo} repetido dentro de la hoja")
            vistos.add(codigo)

            if nivel is None:
                # Ojo: la fuente marca ~229 filas con peligro y respaldo documental pero sin
                # nivel. "Sin dato clasificado" no es lo mismo que "nivel bajo": se descartan.
                if peligro_col:
                    sin_nivel += 1
                continue

            if not isinstance(nivel, int) or nivel not in (1, 2, 3, 4):
                aviso(f"{hoja}!{fila}: NIVEL_PELI fuera de 1-4 ({nivel!r}), descartado")
                continue

            clasificaciones.append({
                "codigo_ccpp": codigo,
                "peligro": nombre_peligro,
                "peligro_slug": slug,
                "tipo": categoria,
                "nivel": nivel,
                "fuente": FUENTE_CANONICA.get(fuente, fuente) or None,
                "fuente_url": link or None,
            })

    if sin_nivel:
        aviso(f"{sin_nivel} filas traen PELIGRO y fuente pero NIVEL_PELI vacío: no se importan")
    if huerfanas:
        aviso(f"{huerfanas} filas sin CODIGO descartadas")

    return ccpp, clasificaciones, distritos


def leer_frecuencia(distritos: dict[str, dict]):
    """Normaliza la hoja ancha de emergencias a una lista por distrito."""
    wb = openpyxl.load_workbook(XLSX_FREC, read_only=True, data_only=True)
    if "NºEMERGENCIAS" not in wb.sheetnames:
        sys.exit(f"ERROR: falta la hoja NºEMERGENCIAS en {XLSX_FREC.name}")

    filas = list(wb["NºEMERGENCIAS"].iter_rows(values_only=True))
    cabecera = [texto(c) for c in filas[0]]
    col = {nombre: i for i, nombre in enumerate(cabecera) if nombre}

    esperadas = [c for _, _, tot, evs in CATEGORIAS for c in [tot, *evs]]
    faltan = [c for c in esperadas if c not in col]
    if faltan:
        sys.exit(f"ERROR: faltan columnas en {XLSX_FREC.name}: {faltan}")

    # El Excel no trae ubigeo: se resuelve por nombre normalizado. En Cusco no hay dos
    # distritos homónimos, así que el match es unívoco.
    por_nombre = {norm(d["distrito"]): d for d in distritos.values()}

    resultado = []
    vistos: set[str] = set()

    for fila, r in enumerate(filas[1:], start=2):
        if not r or r[2] is None:
            continue
        nombre_dist = texto(r[2])
        destino = por_nombre.get(norm(nombre_dist))
        if destino is None:
            aviso(f"frecuencia!{fila}: distrito '{nombre_dist}' no existe en el padrón de CCPP")
            continue
        vistos.add(destino["ubigeo"])

        categorias = []
        total_desglosado = 0
        total_declarado = 0
        algun_desglose = False

        for nombre_cat, slug_cat, col_tot, cols_ev in CATEGORIAS:
            eventos = []
            for nombre_ev in cols_ev:
                conteo = entero(r[col[nombre_ev]])
                if conteo:
                    eventos.append({
                        "evento": titulo_evento(nombre_ev),
                        "slug": slug_evento(nombre_ev),
                        "conteo": conteo,
                    })
            suma = sum(e["conteo"] for e in eventos)
            declarado = entero(r[col[col_tot]])
            total_desglosado += suma
            total_declarado += declarado or 0
            if eventos:
                algun_desglose = True

            if declarado is not None and suma and declarado != suma:
                aviso(
                    f"frecuencia!{fila} ({nombre_dist}): '{col_tot}' declara {declarado} "
                    f"pero el desglose suma {suma}; se usa el desglose"
                )

            eventos.sort(key=lambda e: -e["conteo"])
            categorias.append({
                "categoria": nombre_cat,
                "slug": slug_cat,
                "total": suma if eventos else (declarado or 0),
                "solo_total": not eventos and bool(declarado),
                "eventos": eventos,
            })

        fuente = texto(r[col["FUENTE"]])
        registro = {
            "ubigeo": destino["ubigeo"],
            "distrito": destino["distrito"],
            "provincia": destino["provincia"],
            "rango_fecha": texto(r[col["RANGO FECHA"]]).replace(" ", "") or None,
            "fuente": FUENTE_CANONICA.get(fuente, fuente) or None,
            "fuente_url": texto(r[col["LINK"]]) or None,
            "desglose_disponible": algun_desglose,
            "total": total_desglosado if algun_desglose else total_declarado,
            "categorias": categorias,
        }

        if not algun_desglose and total_declarado:
            aviso(
                f"frecuencia!{fila} ({nombre_dist}): la fuente declara {total_declarado} "
                f"emergencias pero no las desagrega por tipo de evento"
            )

        resultado.append(registro)

    for ubigeo in sorted(set(distritos) - vistos):
        aviso(f"distrito {ubigeo} ({distritos[ubigeo]['distrito']}) sin fila en el Excel de frecuencia")

    resultado.sort(key=lambda d: d["ubigeo"])
    return resultado


def escribir(nombre: str, datos) -> None:
    ruta = OUT / nombre
    ruta.write_text(json.dumps(datos, ensure_ascii=False), encoding="utf-8")
    print(f"  {nombre}: {len(datos)} registros ({ruta.stat().st_size / 1024:.0f} KB)")


def main() -> None:
    for archivo in (XLSX_NIVEL, XLSX_FREC):
        if not archivo.exists():
            sys.exit(f"ERROR: no se encuentra {archivo}")

    OUT.mkdir(parents=True, exist_ok=True)

    ccpp, clasificaciones, distritos = leer_niveles()
    frecuencia = leer_frecuencia(distritos)

    print("Generado en prototype/public/data/:")
    escribir("ccpp.json", sorted(ccpp.values(), key=lambda c: c["codigo"]))
    escribir("peligros.json", clasificaciones)
    escribir("frecuencia.json", frecuencia)

    con_clasificacion = len({c["codigo_ccpp"] for c in clasificaciones})
    con_desglose = sum(1 for d in frecuencia if d["desglose_disponible"])
    print(
        f"\nResumen: {len(ccpp)} centros poblados ({con_clasificacion} con al menos una "
        f"clasificación), {len(clasificaciones)} clasificaciones, "
        f"{len(distritos)} distritos, {len(frecuencia)} con fila de frecuencia "
        f"({con_desglose} con desglose por evento)."
    )

    if avisos:
        print(f"\n{len(avisos)} avisos de calidad de datos:", file=sys.stderr)
        for msg in avisos:
            print(f"  - {msg}", file=sys.stderr)


if __name__ == "__main__":
    main()
